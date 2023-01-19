"""
Definition of EXCEL views for the SEEKER app.
"""


from django.apps import apps
from django.contrib import admin
from django.contrib.auth import login, authenticate
from django.contrib.auth.models import Group, User
from django.urls import reverse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction
from django.db.models import Q, Prefetch, Count, F
from django.db.models.functions import Lower
from django.db.models.query import QuerySet 
from django.forms import formset_factory, modelformset_factory, inlineformset_factory, ValidationError
from django.forms.models import model_to_dict
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect, JsonResponse, FileResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.template.loader import render_to_string
from django.template import Context
from django.views.generic.detail import DetailView
from django.views.generic.base import RedirectView
from django.views.generic import ListView, View
from django.views.decorators.csrf import csrf_exempt

# General imports
import io, sys, os
import openpyxl, json
from openpyxl.utils.cell import get_column_letter
from openpyxl.cell import Cell
from openpyxl import Workbook
from io import StringIO
import csv
import requests

# ======= imports from my own application ======
from lila.settings import APP_PREFIX, MEDIA_DIR
from lila.utils import ErrHandle

from lila.seeker.models import Manuscript, MsItem, Canwit, Profile, Report, Codico, Codhead, Location, LocationType, Library, \
    Austat, Auwork, Collection, Colwit
from lila.seeker.views import app_editor
from lila.reader.views import ReaderImport
from lila.reader.forms import UploadFileForm



class ManuscriptUploadExcel(ReaderImport):
    """Specific parameters for importing manuscripts from Excel"""

    import_type = "excel"
    sourceinfo_url = "https://www.ru.nl/lilac/upload_manuscripts"
    model = Manuscript
    template_name = "reader/import_manuscripts.html"

    def process_files(self, request, source, lResults, lHeader):
        file_list = []
        oErr = ErrHandle()
        bOkay = True
        code = ""
        col_number = {}
        col_defs = [
            {"name": "lilacode",    "def": ['key']                      },
            {"name": "idno",        "def": ['shelf mark', 'shelfmark']  },
            {"name": "lcity",       "def": ['city']                     },
            {"name": "lcountry",    "def": ['country']                  },
            {"name": "library",     "def": ['library']                  },
            {"name": "origins",     "def": ['origin']                   },
            {"name": "dates",       "def": ['date']                     },
            {"name": "provenances", "def": ['provenance']               },
            {"name": "script",      "def": ['script']                   },
            {"name": "size",        "def": ['size']                     },
            {"name": "notes",       "def": ['notes']                    },
            ]
        oStatus = self.oStatus
        lMyHeader = ["status", "msg", "manu_id", "shelfmark", "lilacode", "filename"]

        try:
            # Make sure we have the username
            username = self.username
            profile = Profile.get_user_profile(username)
            team_group = app_editor
            kwargs = {'profile': profile, 'username': username, 'team_group': team_group, 
                      'keyfield': 'path', 'source': source}

            # Initialize column numbers
            for oDef in col_defs:
                col_number[oDef['name']] = -1

            # Get the contents of the imported file
            files = request.FILES.getlist('files_field')
            if files != None:
                for data_file in files:
                    filename = data_file.name
                    file_list.append(filename)

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        self.arErr.append("No source file specified for the selected project")
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]

                        lst_manual = []
                        lst_read = []

                        # Adapt the [lHeader] to contain the correct headers for Report
                        lHeader.clear()
                        for x in lMyHeader: lHeader.append(x)

                        # Further processing depends on the extension
                        oResult = {'status': 'ok', 'count': 0, 'sermons': 0, 'msg': "", 'user': username}

                        if extension == "xlsx":
                            # This is an Excel file: read the file using openpyxl
                            # Write data temporarily to the WRITABLE dir, but with a temporary filename
                            tmp_path = os.path.abspath(os.path.join( MEDIA_DIR, filename))
                            with io.open(tmp_path, "wb") as f:
                                sData = data_file.read()
                                f.write(sData)

                            # Read string file
                            wb = openpyxl.load_workbook(tmp_path, read_only=True)
                            sheetnames = wb.sheetnames
                            ws_manu = None
                            for sname in sheetnames:
                                if "manu" in sname.lower():
                                    ws_manu = wb[sname]
                            # Do we have a manuscript worksheet?
                            if ws_manu != None:
                                # Get a list of all columns (in lower case)
                                self.get_columns_excel(ws_manu, col_defs, col_number)

                                # Walk through all rows
                                bStop = False
                                row_num = 2
                                value = ws_manu.cell(row=row_num, column=1).value
                                while not bStop and not value is None or value == "":
                                    # Collect the data from this row
                                    oManu = self.get_row_excel(ws_manu, row_num, col_number)

                                    # Check if this manuscript already exists
                                    idno = oManu.get("idno")
                                    lilacode = oManu.get("lilacode")
                                    bExists = False
                                    if idno is None:
                                        # Try to locate it by lilacode
                                        if not lilacode is None:
                                            bExists = Manuscript.objects.filter(lilacode__iexact=lilacode).exists()
                                    else:
                                        # Do we have a lilacode
                                        if lilacode is None:
                                            # There is no lilacode
                                            bExists = Manuscript.objects.filter(idno__iexact=idno).exists()
                                        else:
                                            bExists = Manuscript.objects.filter(idno__iexact=idno, lilacode__iexact=lilacode).exists()

                                    if not bExists:
                                        # Add this manuscript using custom_add()
                                        manu = Manuscript.custom_add(oManu, **kwargs)

                                        # Now get the codicological unit that has been automatically created and adapt it
                                        codico = manu.manuscriptcodicounits.first()
                                        if codico != None:
                                            oManu['manuscript'] = manu

                                            # The manuscript's 'size' could serve as 'format' for the codico
                                            if oManu.get("format") is None:
                                                oManu['format'] = oManu.get("size")
                                            # The manuscript's 'title' could serve as 'name' for the codico
                                            if oManu.get("name") is None:
                                                oManu['name'] = oManu.get("title")

                                            codico = Codico.custom_add(oManu, **kwargs)

                                        oResult['count'] += 1
                                        oResult['manu'] = manu.id
                                        oResult['idno'] = manu.idno
                                        oResult['lilacode'] = manu.lilacode

                                        # Add the result to the list of results
                                        lResults.append(oResult)

                                        # Create the [Report] results for this one
                                        oRead = dict(status="ok", msg="read", manu_id=manu.id,
                                                filename=filename, shelfmark=manu.idno, lilacode=manu.lilacode)
                                        lst_read.append(oRead)

                                    # Go to the next row
                                    row_num += 1
                                    # Get the values from there
                                    value = ws_manu.cell(row=row_num, column=1).value



                        # Create a report and add it to what we return
                        oContents = {'headers': lHeader, 'list': lst_read}
                        oReport = Report.make(username, "ixlsx", json.dumps(oContents))
                                
                    # Set the status
                    oStatus.set("finishing", msg="file={}".format(filename))
            code = "Imported using the [import_excel] function on this file: {}".format(", ".join(file_list))
        except:
            bOkay = False
            code = oErr.get_error_message()
        return bOkay, code


class ManuscriptUploadCanwits(ReaderImport):
    """Specific parameters for importing canwits into an *EXISTING* manuscript from Excel"""

    import_type = "canwits"
    sourceinfo_url = "https://www.ru.nl/lilac/upload_canwits"
    model = Manuscript
    template_name = "reader/import_canwits.html"

    def process_files(self, request, source, lResults, lHeader):
        file_list = []
        oErr = ErrHandle()
        bOkay = True
        code = ""
        col_number = {}
        col_defs = [
            {"name": "lilacode",    "def": ['lilac']                    },
            {"name": "locus",       "def": ['locus']                    },
            {"name": "caput",       "def": ['caput']                    },
            {"name": "collection",  "def": ['collection']               },
            {"name": "author",      "def": ['author']                   },
            {"name": "ftext",       "def": ['ftext', 'full text']       },
            {"name": "ftrans",      "def": ['ftrans', 'translation']    },
            {"name": "brefs",       "def": ['bibref', 'bible']          },
            {"name": "austat_link", "def": ['authoritative statement']  },
            {"name": "austat_note", "def": ['fons materialis (note)']   }
            ]
        oStatus = self.oStatus

        try:
            # Make sure we have the username
            username = self.username
            profile = Profile.get_user_profile(username)
            team_group = app_editor
            kwargs = {'profile': profile, 'username': username, 'team_group': team_group}

            # Initialize column numbers
            for oDef in col_defs:
                col_number[oDef['name']] = -1

            # Make sure it is clear which manuscript this is
            manu = self.obj
            if manu is None:
                manu_id = self.qd.get("pk")
                if not manu_id is None:
                    manu = Manuscript.objects.filter(id=manu_id).first()

            # Double checking: if there is no manuscript, we cannot import
            if manu is None:
                code = "No manuscript found to attach the Canwits to"
                bOkay = False
                return bOkay, code

            # Get a possible parent
            parent = None
            codhead_id = self.qd.get("manu-{}-headlist".format(manu.id))
            if not codhead_id is None:
                codhead = Codhead.objects.filter(id=codhead_id).first()
                if not codhead is None:
                    parent = codhead.msitem

            # Get the contents of the imported file
            file_list = request.FILES.getlist('files_field')
            if not file_list is None and len(file_list) > 0:
                data_file = file_list[0]
                if not data_file is None:
                    filename = data_file.name

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        self.arErr.append("No source file specified for the selected project")
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]

                        if extension == "xlsx":
                            # This is an Excel file: read the file using openpyxl
                            # Write data temporarily to the WRITABLE dir, but with a temporary filename
                            tmp_path = os.path.abspath(os.path.join( MEDIA_DIR, filename))
                            with io.open(tmp_path, "wb") as f:
                                sData = data_file.read()
                                f.write(sData)

                            # Read string file
                            wb = openpyxl.load_workbook(tmp_path, read_only=True)
                            sheetnames = wb.sheetnames

                            bFound = False
                            for sheetname in sheetnames:
                                if not bFound:
                                    low_sheetname = sheetname.lower()
                                    # Either something with " canon ... witness ..."
                                    if "canon" in low_sheetname and "witness" in low_sheetname:
                                        bFound = True
                                        break
                                    # Or something containing "... canwit ..."
                                    elif "canwit" in low_sheetname:
                                        bFound = True
                                        break
                            # Check if we actually found a worksheet to load CanWit specifications from
                            if bFound:
                                # Yes, we have the correct worksheet
                                ws = wb[sheetname]

                                # Get a list of all columns (in lower case)
                                # The columns that we are expecting for each CanWit are:
                                #    locus, ftext, ftrans, bibref
                                # Optional columns, once the CanWit has been established:
                                #    collection, austat link, 
                                self.get_columns_excel(ws, col_defs, col_number)
                                #row_num = 1
                                #col_num = 1
                                #bStop = False
                                #while not bStop:
                                #    k = ws.cell(row=row_num, column=col_num).value
                                #    if k is None or k == "":
                                #        bStop = True
                                #    else:
                                #        col_name = k.lower()
                                #        for oDef in col_defs:
                                #            name = oDef['name']
                                #            if col_number[name] < 0:
                                #                for str_def in oDef['def']:
                                #                    if str_def in col_name:
                                #                        # Found it!
                                #                        col_number[name] = col_num
                                #                        break
                                #    col_num += 1

                                # Make sure we at least have [ftext]
                                if col_number['ftext'] >= 1:
                                    # Get the order number by looking for the highest order number so far
                                    order = 1
                                    msitem_last = MsItem.objects.filter(codico__manuscript=manu).order_by('order').last()
                                    if not msitem_last is None:
                                        order = msitem_last.order + 1

                                    # Walk through all rows
                                    bStop = False
                                    row_num = 2
                                    while not bStop:
                                        value = ws.cell(row=row_num, column=1).value
                                        bStop = (value is None or value == "")
                                        if not bStop:
                                            # Read the values for this row
                                            oValue = self.get_row_excel(ws, row_num, col_number)
                                            #oValue = {}
                                            #for field_name, col_num in col_number.items():
                                            #    if col_num > 0:
                                            #        oValue[field_name] = ws.cell(row=row_num, column=col_num).value

                                            # We have all the values, process the essentials to create a CanWit (if it doesn't exist already)
                                            val_ftext = oValue['ftext']
                                            canwit = Canwit.objects.filter(msitem__codico__manuscript=manu, ftext__iexact=val_ftext).first()
                                            if canwit is None:
                                                # Start creating a result
                                                oResult = {}
                                                oResult['order'] = order
                                                oResult['manu'] = manu.id

                                                # Make sure to indicate that each row in the Excel is not a structural (hierarchy creating)
                                                #    type, but an actual Canwit
                                                oValue['type'] = 'canwit'
                                                canwit = Canwit.custom_add(oValue, manuscript=manu, order=order, parent=parent)
                                                order += 1

                                                oResult['canwit'] = canwit.id
                                                oResult['lilacode'] = canwit.get_lilacode()
                                                oResult['locus'] = canwit.get_locus()
                                                oResult['caput'] = canwit.get_caput()

                                                # Add the result to the list of results
                                                lResults.append(oResult)
                                        # Go to the next row
                                        row_num += 1
                                    

                            else:
                                # Actually not sure what to do
                                oErr.Status("WARNING: Cannot find a CanWit or Canon Witness worksheet in this Excel")
                        else:
                            oErr.Status("WARNING: cannot decode extension [{}]".format(extension))

        except:
            bOkay = False
            code = oErr.get_error_message()
        return bOkay, code


class ManuscriptUploadJson(ReaderImport):
    import_type = "json"
    sourceinfo_url = "https://www.ru.nl/lilac/upload_json"

    def process_files(self, request, source, lResults, lHeader):
        file_list = []
        oErr = ErrHandle()
        bOkay = True
        code = ""
        oStatus = self.oStatus
        try:
            # Make sure we have the username
            username = self.username
            profile = Profile.get_user_profile(username)
            team_group = app_editor
            kwargs = {'profile': profile, 'username': username, 'team_group': team_group, 'keyfield': 'path'}

            # Get the contents of the imported file
            files = request.FILES.getlist('files_field')
            if files != None:
                for data_file in files:
                    filename = data_file.name
                    file_list.append(filename)

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        self.arErr.append("No source file specified for the selected project")
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]

                        lst_manual = []
                        lst_read = []

                        # Further processing depends on the extension
                        oResult = {'status': 'ok', 'count': 0, 'sermons': 0, 'msg': "", 'user': username}

                        if extension == "json":
                            # This is a JSON file: Load the file into a variable
                            sData = data_file.read()
                            lst_manu = json.loads( sData.decode(encoding="utf8"))

                            # Check if this is a dictionary or a list
                            if isinstance(lst_manu, dict):
                                # It is a dictionary: turn it into a list
                                oManuList = lst_manu
                                lst_manu = [v for k,v in oManuList.items()]

                            # Walk through the manuscripts
                            for oManu in lst_manu:
                                # Each manuscript has some stuff of its own
                                # We have an object with key/value pairs: process it
                                manu = Manuscript.custom_add(oManu, **kwargs)

                                # Now get the codicological unit that has been automatically created and adapt it
                                codico = manu.manuscriptcodicounits.first()
                                if codico != None:
                                    oManu['manuscript'] = manu
                                    codico = Codico.custom_add(oManu, **kwargs)

                                oResult['count'] += 1
                                oResult['obj'] = manu
                                oResult['name'] = manu.idno

                                # Process all the MsItems into a list of sermons
                                canwit_list = []
                                for oMsItem in oManu['msitems']:
                                    # Get the sermon object
                                    oSermon = oMsItem['sermon']
                                    order = oMsItem['order']
                                    sermon = Canwit.custom_add(oSermon, manu, order, **kwargs)

                                    # Keep track of the number of sermons read
                                    oResult['sermons'] += 1

                                    # Get parent, firstchild, next
                                    parent = oMsItem['parent']
                                    firstchild = oMsItem['firstchild']
                                    nextone = oMsItem['next']

                                    # Add to list
                                    canwit_list.append({'order': order, 'parent': parent, 'firstchild': firstchild,
                                                        'next': nextone, 'sermon': sermon})

                                # Now process the parent/firstchild/next items
                                with transaction.atomic():
                                    for oSermo in canwit_list:
                                        # Get the p/f/n numbers
                                        parent_id = oSermo['parent']
                                        firstchild_id = oSermo['firstchild']
                                        next_id = oSermo['next']
                                        # Process parent
                                        if parent_id != '' and parent_id != None:
                                            # parent_id = str(parent_id)
                                            parent = next((obj['sermon'] for obj in canwit_list if obj['order'] == parent_id), None)
                                            oSermo['sermon'].msitem.parent = parent.msitem
                                            oSermo['sermon'].msitem.save()
                                        # Process firstchild
                                        if firstchild_id != '' and firstchild_id != None:
                                            # firstchild_id = str(firstchild_id)
                                            firstchild = next((obj['sermon'] for obj in canwit_list if obj['order'] == firstchild_id), None)
                                            oSermo['sermon'].msitem.firstchild = firstchild.msitem
                                            oSermo['sermon'].msitem.save()
                                        # Process next
                                        if next_id != '' and next_id != None:
                                            # next_id = str(next_id)
                                            nextone = next((obj['sermon'] for obj in canwit_list if obj['order'] == next_id), None)
                                            oSermo['sermon'].msitem.next = nextone.msitem
                                            oSermo['sermon'].msitem.save()


                        # Create a report and add it to what we return
                        oContents = {'headers': lHeader, 'list': lst_manual, 'read': lst_read}
                        oReport = Report.make(username, "ixlsx", json.dumps(oContents))
                                
                        # Determine a status code
                        statuscode = "error" if oResult == None or oResult['status'] == "error" else "completed"
                        if oResult == None:
                            self.arErr.append("There was an error. No manuscripts have been added")
                        else:
                            lResults.append(oResult)

            code = "Imported using the [import_json] function on this file list: {}".format(", ".join(file_list))
        except:
            bOkay = False
            code = oErr.get_error_message()
        return bOkay, code


class AustatUploadExcel(ReaderImport):
    """Specific parameters for importing manuscripts from Excel"""

    import_type = "excel"
    sourceinfo_url = "https://www.ru.nl/english/people/meeder-s/"
    template_name = "reader/import_austats.html"

    def process_files(self, request, source, lResults, lHeader):
        file_list = []
        oErr = ErrHandle()
        bOkay = True
        code = ""
        col_number = {}
        col_defs = [
            {"name": "key",         "def": ['key']                      },
            {"name": "author",      "def": ['author']                   },
            {"name": "opus",        "def": ['opus']                     },
            {"name": "auwork",      "def": ['work']                     },
            {"name": "date",        "def": ['date']                     },
            {"name": "ftext",       "def": ['ftext', 'full text']       },
            {"name": "ftrans",      "def": ['ftrans', 'translation']    },
            {"name": "genres",      "def": ['genre(s)']                 },
            {"name": "keywords",    "def": ['keywords']                 },
            {"name": "editions",    "def": ['editions']                 },
            {"name": "signatures",  "def": ['cpl']                      }
            ]
        oStatus = self.oStatus
        try:
            # Make sure we have the username
            username = self.username
            profile = Profile.get_user_profile(username)
            team_group = app_editor
            kwargs = {'profile': profile, 'username': username, 'team_group': team_group}

            # Initialize column numbers
            for oDef in col_defs:
                col_number[oDef['name']] = -1

            # Get the contents of the imported file
            files = request.FILES.getlist('files_field')
            if files != None:
                for data_file in files:
                    filename = data_file.name
                    file_list.append(filename)

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        self.arErr.append("No source file specified for the selected project")
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]

                        lst_manual = []
                        lst_read = []

                        # Further processing depends on the extension
                        oResult = {'status': 'ok', 'count': 0, 'austats': 0, 'msg': "", 'user': username}

                        if extension == "xlsx":
                            # This is an Excel file: read the file using openpyxl
                            # Write data temporarily to the WRITABLE dir, but with a temporary filename
                            tmp_path = os.path.abspath(os.path.join( MEDIA_DIR, filename))
                            with io.open(tmp_path, "wb") as f:
                                sData = data_file.read()
                                f.write(sData)

                            # Read string file
                            wb = openpyxl.load_workbook(tmp_path, read_only=True)

                            # Find out which sheets point to Work (AuWork) and to Austat
                            sheetnames = wb.sheetnames
                            ws_work = None
                            ws_austat = None
                            for sname in sheetnames:
                                snameL = sname.lower()
                                if snameL.find("work-") == 0:
                                    ws_work = wb[sname]
                                elif snameL.find("austat-") == 0:
                                    ws_austat = wb[sname]

                            # If there is a Auwork sheet, then read that first
                            if not ws_work is None:
                                # THere is a Auwork sheet: look for Auwork definitions
                                oAuWork = {}
                                pass

                            # Do we have a worksheet with Austat items?
                            if not ws_austat is None:
                                # Get a list of the Austat excel columns
                                self.get_columns_excel(ws_austat, col_defs, col_number)

                                # Make sure we at least have [ftext]
                                if col_number['ftext'] >= 1:

                                    # Walk through all rows
                                    bStop = False
                                    row_num = 2
                                    while not bStop:
                                        value = ws_austat.cell(row=row_num, column=1).value
                                        bStop = (value is None or value == "")
                                        if not bStop:
                                            # Read the values for this row into dictionary [oValue]
                                            oValue = self.get_row_excel(ws_austat, row_num, col_number)
                                            #oValue = {}
                                            #for field_name, col_num in col_number.items():
                                            #    if col_num > 0:
                                            #        oValue[field_name] = ws_austat.cell(row=row_num, column=col_num).value

                                            # We have all the values, process the essentials to create a CanWit (if it doesn't exist already)
                                            val_key = oValue['key']
                                            # Split into Auwork.key (short Work key) and Austat.keycode (number)
                                            work_key, austat_key = Austat.split_key(val_key)
                                            austat = Austat.objects.filter(auwork__key=work_key, keycode=austat_key).first()
                                            if austat is None:
                                                # Start creating a result
                                                oResult = {}
                                                oResult['key'] = val_key

                                                # Make sure to indicate that each row in the Excel is an Austat description
                                                oValue['type'] = 'austat'
                                                oValue['work_key'] = work_key
                                                oValue['austat_key'] = austat_key
                                                austat = Austat.custom_add(oValue)

                                                # Not sure what to do with 'order'
                                                # order += 1

                                                oResult['austat'] = austat.id
                                                oResult['keycode'] = austat.get_keycode()
                                                oResult['ftext'] = "-" if austat.ftext is None else austat.ftext

                                                # Add the result to the list of results
                                                lResults.append(oResult)
                                        # Go to the next row
                                        row_num += 1

 
                    # Set the status
                    oStatus.set("finishing", msg="file={}".format(filename))


            code = "Imported using the [import_excel] function on this file: {}".format(", ".join(file_list))
        except:
            bOkay = False
            code = oErr.get_error_message()
        return bOkay, code


class AuworkUploadExcel(ReaderImport):
    """Specific parameters for importing manuscripts from Excel"""

    import_type = "excel"
    sourceinfo_url = "https://www.ru.nl/english/people/meeder-s/"
    template_name = "reader/import_auworks.html"

    def process_files(self, request, source, lResults, lHeader):
        file_list = []
        oErr = ErrHandle()
        bOkay = True
        code = ""
        col_number = {}
        col_defs = [
            {"name": "key",         "def": ['key']                  },
            {"name": "work",        "def": ['work']                 },
            {"name": "full",        "def": ['full', 'full text']    },
            {"name": "date",        "def": ['date']                 },
            {"name": "opus",        "def": ['opus']                 },
            {"name": "genres",      "def": ['genre(s)']             },
            {"name": "keywords",    "def": ['keywords']             },
            {"name": "signatures",  "def": ['cpl', 'signature']     }
            ]
        oStatus = self.oStatus
        try:
            # Make sure we have the username
            username = self.username
            profile = Profile.get_user_profile(username)
            team_group = app_editor
            kwargs = {'profile': profile, 'username': username, 'team_group': team_group}

            # Initialize column numbers in dictionary [col_number]
            for oDef in col_defs:
                col_number[oDef['name']] = -1

            # Get the contents of the imported file
            files = request.FILES.getlist('files_field')
            if files != None:
                for data_file in files:
                    filename = data_file.name
                    file_list.append(filename)

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        self.arErr.append("No source file specified for the selected project")
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]

                        # Further processing depends on the extension
                        oResult = {'status': 'ok', 'count': 0, 'auworks': 0, 'msg': "", 'user': username}

                        if extension == "xlsx":
                            # This is an Excel file: read the file using openpyxl
                            # Write data temporarily to the WRITABLE dir, but with a temporary filename
                            tmp_path = os.path.abspath(os.path.join( MEDIA_DIR, filename))
                            with io.open(tmp_path, "wb") as f:
                                sData = data_file.read()
                                f.write(sData)

                            # Read string file
                            wb = openpyxl.load_workbook(tmp_path, read_only=True)

                            # Find out which sheets point to work (=auwork)
                            sheetnames = wb.sheetnames
                            ws_auwork = None
                            for sname in sheetnames:
                                snameL = sname.lower()
                                if snameL.find("work-") == 0:
                                    ws_auwork = wb[sname]

                            # Do we have a worksheet with Auwork items?
                            if not ws_auwork is None:
                                # Get a list of the Auwork excel columns
                                self.get_columns_excel(ws_auwork, col_defs, col_number)
 
                                # Make sure we at least have [key]
                                if col_number['key'] >= 1:

                                    # Walk through all rows
                                    bStop = False
                                    row_num = 2
                                    while not bStop:
                                        value = ws_auwork.cell(row=row_num, column=1).value
                                        bStop = (value is None or value == "")
                                        if not bStop:
                                            # Read the values for this row into dictionary [oValue]
                                            oValue = self.get_row_excel(ws_auwork, row_num, col_number)

                                            # We have all the values, process the essentials to create a CanWit (if it doesn't exist already)
                                            auwork_key = oValue['key']
                                            auwork = Auwork.objects.filter(key=auwork_key).first()
                                            # Only process if this Auwork has not yet been added!!!
                                            if auwork is None:
                                                # Start creating a result
                                                oResult = {}
                                                oResult['key'] = auwork_key

                                                # Create this work
                                                auwork = Auwork.custom_add(oValue)

                                                # Process what we get back
                                                oResult['auwork'] = auwork.id
                                                oResult['work'] = "-" if auwork.work is None else auwork.work
                                                oResult['opus'] = "-" if auwork.opus is None else auwork.opus

                                                # Add the result to the list of results
                                                lResults.append(oResult)
                                        # Go to the next row
                                        row_num += 1

                    # Set the status
                    oStatus.set("finishing", msg="file={}".format(filename))


            code = "Imported using the [import_excel] function on this file: {}".format(", ".join(file_list))
        except:
            bOkay = False
            code = oErr.get_error_message()
        return bOkay, code


class ColwitUploadExcel(ReaderImport):
    """Specific parameters for importing manuscripts from Excel"""

    import_type = "excel"
    sourceinfo_url = "https://www.ru.nl/english/people/meeder-s/"
    template_name = "reader/import_colwits.html"

    def process_files(self, request, source, lResults, lHeader):
        file_list = []
        oErr = ErrHandle()
        bOkay = True
        code = ""
        col_number = {}
        col_defs = [
            {"name": "key",         "def": ['key']                  },
            {"name": "manuscript",  "def": ['manuscript']           },
            {"name": "section",     "def": ['section']              },
            {"name": "locus",       "def": ['locus']                },
            {"name": "collection",  "def": ['collection']           },
            {"name": "notes",       "def": ['notes']                },
            {"name": "signatures",  "def": ['cpl', 'signature']     }
            ]
        oStatus = self.oStatus
        try:
            # Make sure we have the username
            username = self.username
            profile = Profile.get_user_profile(username)
            team_group = app_editor
            kwargs = {'profile': profile, 'username': username, 'team_group': team_group}

            # Initialize column numbers in dictionary [col_number]
            for oDef in col_defs:
                col_number[oDef['name']] = -1

            # Get the contents of the imported file
            files = request.FILES.getlist('files_field')
            if files != None:
                for data_file in files:
                    filename = data_file.name
                    file_list.append(filename)

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        self.arErr.append("No source file specified for the selected project")
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]

                        # Further processing depends on the extension
                        oResult = {'status': 'ok', 'count': 0, 'colwits': 0, 'msg': "", 'user': username}

                        if extension == "xlsx":
                            # This is an Excel file: read the file using openpyxl
                            # Write data temporarily to the WRITABLE dir, but with a temporary filename
                            tmp_path = os.path.abspath(os.path.join( MEDIA_DIR, filename))
                            with io.open(tmp_path, "wb") as f:
                                sData = data_file.read()
                                f.write(sData)

                            # Read string file
                            wb = openpyxl.load_workbook(tmp_path, read_only=True)

                            # Find out which sheets point to work (=colwit)
                            sheetnames = wb.sheetnames
                            ws_colwit = None
                            for sname in sheetnames:
                                snameL = sname.lower()
                                if snameL.find("colwit") == 0:
                                    ws_colwit = wb[sname]

                            # Do we have a worksheet with Colwit items?
                            if not ws_colwit is None:
                                # Get a list of the Colwit excel columns
                                self.get_columns_excel(ws_colwit, col_defs, col_number)
 
                                # Make sure we at least have [key]
                                if col_number['key'] >= 1 and col_number['manuscript'] >= 1 and col_number['collection'] >= 1:

                                    # Walk through all rows
                                    bStop = False
                                    row_num = 2
                                    while not bStop:
                                        value = ws_colwit.cell(row=row_num, column=1).value
                                        bStop = (value is None or value == "")
                                        if not bStop:
                                            # Read the values for this row into dictionary [oValue]
                                            oValue = self.get_row_excel(ws_colwit, row_num, col_number)

                                            # We have all the values, process the essentials to create a CanWit (if it doesn't exist already)
                                            colwit_key = oValue['key']
                                            colwit_manu = oValue['manuscript']
                                            colwit_coll = oValue['collection']
                                            colwit_locus = oValue['locus']
                                            if colwit_manu != "" and colwit_coll != "":

                                                # Improve on the collection code and the manuscript code
                                                colwit_manu = colwit_manu.strip('.')
                                                colwit_coll = colwit_coll.strip('.')
                                                oValue['manuscript'] = colwit_manu
                                                oValue['collection'] = colwit_coll

                                                # Calculate what the lilacode should be
                                                lilacode = "{}.{}".format(colwit_manu, colwit_coll)

                                                # Possibly adapt the key inside oValue
                                                if colwit_key != lilacode:
                                                    oValue['key'] = lilacode

                                                # Try to find this one
                                                colwit = Colwit.objects.filter(lilacodefull__iexact=lilacode).first()
                                                if colwit is None:
                                                    # Create it
                                                    colwit = Colwit.objects.filter(lilacodefull=lilacode)

                                                # Make sure we can continue safely
                                                if not colwit is None:
                                                    # Start creating a result
                                                    oResult = {}
                                                    oResult['key'] = colwit_key

                                                    # Create this work
                                                    colwit = Colwit.custom_add(oValue, **kwargs)

                                                    # Process what we get back
                                                    oResult['colwit'] = colwit.id
                                                    oResult['lilacode'] = colwit.lilacodefull
                                                    oResult['manuscript'] = colwit_manu
                                                    oResult['locus'] = colwit_locus
                                                    oResult['collection'] = colwit_coll

                                                    # Add the result to the list of results
                                                    lResults.append(oResult)
                                        # Go to the next row
                                        row_num += 1

                    # Set the status
                    oStatus.set("finishing", msg="file={}".format(filename))


            code = "Imported using the [import_excel] function on this file: {}".format(", ".join(file_list))
        except:
            bOkay = False
            code = oErr.get_error_message()
        return bOkay, code


