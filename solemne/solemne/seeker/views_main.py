"""
Definition of MAIN views for the SEEKER app.
"""

from xml.dom import DomstringSizeErr
from django.apps import apps
from django.contrib import admin
from django.contrib.auth import login, authenticate
from django.contrib.auth.models import Group
from django.urls import reverse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction
from django.db.models import Q, Prefetch, Count, F, Sum
from django.db.models.functions import Lower
from django.db.models.query import QuerySet 
from django.forms import formset_factory, modelformset_factory, inlineformset_factory, ValidationError
from django.forms.models import model_to_dict
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect, JsonResponse, FileResponse
from django.middleware.csrf import get_token
from django.shortcuts import get_object_or_404, render, redirect
from django.template.loader import render_to_string
from django.template import Context
from django.views.generic.detail import DetailView
from django.views.generic.base import RedirectView
from django.views.generic import ListView, View
from django.views.decorators.csrf import csrf_exempt

# Other basic imports
import json, csv
import copy
import re
import openpyxl
from openpyxl.utils.cell import get_column_letter
from lxml import etree as ET
from io import StringIO


# ======= imports from my own application ======
from solemne.utils import ErrHandle, is_ajax
from solemne.bible.models import Reference
from solemne.lict.models import ResearchSet, SetList
from solemne.seeker.models import Auwork, DraggingAustat, LitrefAustat, get_crpp_date, get_current_datetime, process_lib_entries, get_searchable, get_now_time, \
    add_gold2equal, add_equal2equal, add_ssg_equal2equal, get_helptext, Information, Country, City, Author, Manuscript, \
    User, Group, Origin, Canwit, MsItem, Codhead, CanwitKeyword, CanwitAustat, NewsItem, \
    SourceInfo, AustatKeyword, AustatGenre, ManuscriptExt, Colwit, Free, LitrefAustat, \
    ManuscriptKeyword, Action, Austat, AustatLink, Location, LocationName, LocationIdentifier, LocationRelation, LocationType, \
    ProvenanceMan, Provenance, Daterange, CollOverlap, BibRange, Feast, Comment, AustatDist, \
    Basket, BasketMan, BasketAustat, Litref, LitrefMan, LitrefCol, Report, \
    Visit, Profile, Keyword, Signature, CanwitSignature, ColwitSignature, \
    Status, Library, Collection, CollectionCanwit, \
    CollectionMan, Caned, UserKeyword, Template, \
    ManuscriptCorpus, ManuscriptCorpusLock, AustatCorpus, ProjectEditor, \
    Codico, ProvenanceCod, OriginCodico, CodicoKeyword, Reconstruction, \
    Project, ManuscriptProject, CollectionProject, AustatProject, CanwitProject, \
    get_reverse_spec, LINK_EQUAL, LINK_PRT, LINK_BIDIR, LINK_PARTIAL, STYPE_IMPORTED, STYPE_EDITED, STYPE_MANUAL, LINK_UNSPECIFIED
from solemne.seeker.forms import SearchCollectionForm, SearchManuscriptForm, SearchManuForm, SearchSermonForm, LibrarySearchForm, SignUpForm, \
    AuthorSearchForm, UploadFileForm, UploadFilesForm, ManuscriptForm, CanwitForm, CommentForm, \
    AuthorEditForm, BibRangeForm, FeastForm, ColwitForm, \
    CanwitSuperForm, SearchUrlForm, CodheadForm, CanedForm, \
    CanwitSignatureForm, ColwitSignatureForm, AustatLinkForm, ColForm, LitrefAustatForm, \
    ReportEditForm, SourceEditForm, ManuscriptProvForm, LocationForm, LocationRelForm, OriginForm, \
    LibraryForm, ManuscriptExtForm, ManuscriptLitrefForm, CanwitKeywordForm, KeywordForm, \
    ManuscriptKeywordForm, DaterangeForm, ProjectForm, CanwitCollectionForm, CollectionForm, \
    AustatForm, ManuscriptCollectionForm, CollectionLitrefForm, \
    SuperSermonGoldCollectionForm, ProfileForm, UserKeywordForm, ProvenanceForm, ProvenanceManForm, \
    TemplateForm, TemplateImportForm, ManuReconForm,  ManuscriptProjectForm, \
    CodicoForm, CodicoProvForm, ProvenanceCodForm, OriginCodForm, CodicoOriginForm
from solemne.reader.views import reader_uploads
from solemne.seeker.views import get_usercomments, search_generic
from solemne.seeker.views_utils import solemne_action_add, solemne_get_history
from solemne.seeker.adaptations import listview_adaptations, add_codico_to_manuscript

# ======= from RU-Basic ========================
from solemne.basic.views import BasicPart, BasicList, BasicDetails, make_search_list, add_rel_item, adapt_search, \
   adapt_m2m, adapt_m2o, app_editor, app_userplus, treat_bom, \
   user_is_ingroup, user_is_authenticated, user_is_superuser, \
   app_developer, app_moderator


# ======= Settings for the Main views ==========
PAGINATE_BY_VALUE = 100


def adapt_regex_incexp(value):
    """Widen searching for ftext and ftrans
    
    e=ae, j=i, u=v, k=c
    """

    oTranslation = str.maketrans(dict(j="[ji]", i="[ji]", u="[uv]", v="[uv]", k="[kc]", c="[kc]"))

    if value != None and len(value) > 0:
        # Make changes:
        value = value.replace("ae", "e").replace("e", "a?e").translate(oTranslation)

    return value

def get_non_editable_projects(profile, projects):
    """Get the number of projects that I do not have editing rights for"""

    oErr = ErrHandle()
    iCount = 0
    try:
        id_list = []
        current_project_ids = [x['id'] for x in projects.values('id')]
        profile_project_ids = [x['id'] for x in profile.projects.all().values('id')]
        # Walk all the projects I need to evaluate
        for prj_id in current_project_ids:
            if not prj_id in profile_project_ids:
                # I have*NO*  editing rights for this one
                id_list.append(prj_id)
        iCount = len(id_list)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("get_non_editable_projects")
        iCount = 0

    return iCount

def evaluate_projlist(profile, instance, projlist, sText):
    bBack = True
    msg = ""
    oErr = ErrHandle()
    try:
        if projlist is None or len(projlist) == 0:
            # Check how many projects the user does *NOT* have rights for
            non_editable_projects = get_non_editable_projects(profile, instance.projects.all())
            if non_editable_projects == 0:
                # The user has not selected a project (yet): try default assignment
                user_projects = profile.projects.all()
                if user_projects.count() != 1:
                    # We cannot assign the default project
                    bBack = False
                    msg = "Make sure to assign this {} to one project before saving it".format(sText)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("evaluate_projlist")
        bBack = False
    return bBack, msg

def may_edit_project(request, profile, instance):
    """Check if the user is allowed to edit this project"""

    bBack = False
    # Is the user an editor?
    if user_is_ingroup(request, app_editor):
        # Get the projects this user has authority for
        user_projects = profile.get_project_ids()
        if len(user_projects) > 0:
            # True: the user may indeed edit *some* projects
            bBack = True

            # The following is now superfluous
            use_superfluous = False
            if use_superfluous:
                # Get the projects associated with [instance']
                project_ids = [x['id'] for x in instance.projects.all().values('id')]
                # See if there is any match
                for project_id in user_projects:
                    if project_id in project_ids:
                        bBack = True
                        break
    return bBack

def project_dependant_delete(request, to_be_deleted):
    """Delete items from the linktable, provided the user has the right to"""

    oErr = ErrHandle()
    bBack = True
    try:
        # Find out who this is
        profile = Profile.get_user_profile(request.user.username)
        # Get the editing rights for this person
        project_id = [x['id'] for x in profile.projects.all().values("id")]

        # CHeck all deletables
        delete = []
        for obj in to_be_deleted:
            # Get the project id of the deletables
            obj_id = obj.id
            prj_id = obj.project.id
            if prj_id in project_id:
                # The user may delete this project relation
                # Therefore: delete the OBJ that holde this relation!
                delete.append(obj_id)
        # Is anything left?
        if len(delete) > 0:
            # Get the class of the deletables
            cls = to_be_deleted[0].__class__
            # Delete all that need to be deleted
            cls.objects.filter(id__in=delete).delete()

    except:
        msg = oErr.get_error_message()
        oErr.DoError("project_dependant_delete")
        bBack = False
    return bBack




# ============= MANUSCRIPT VIEWS AND OPERATIONS ============================


class ManuscriptEdit(BasicDetails):
    """The details of one manuscript"""

    model = Manuscript  
    mForm = ManuscriptForm
    prefix = 'manu'
    titlesg = "Manuscript identifier"
    rtype = "json"
    new_button = True
    mainitems = []
    use_team_group = True
    history_button = True
    
    McolFormSet = inlineformset_factory(Manuscript, CollectionMan,
                                       form=ManuscriptCollectionForm, min_num=0,
                                       fk_name="manuscript", extra=0)
    MlitFormSet = inlineformset_factory(Manuscript, LitrefMan,
                                         form = ManuscriptLitrefForm, min_num=0,
                                         fk_name = "manuscript",
                                         extra=0, can_delete=True, can_order=False)
    MprovFormSet = inlineformset_factory(Manuscript, ProvenanceMan,
                                         form=ManuscriptProvForm, min_num=0,
                                         fk_name = "manuscript",
                                         extra=0, can_delete=True, can_order=False)
    MextFormSet = inlineformset_factory(Manuscript, ManuscriptExt,
                                         form=ManuscriptExtForm, min_num=0,
                                         fk_name = "manuscript",
                                         extra=0, can_delete=True, can_order=False)    
    # Kan weg
    MprojFormSet = inlineformset_factory(Manuscript, ManuscriptProject,
                                         form=ManuscriptProjectForm, min_num=0,
                                         fk_name = "manuscript",
                                         extra=0, can_delete=True, can_order=False)

    formset_objects = [{'formsetClass': McolFormSet,  'prefix': 'mcol',  'readonly': False, 'noinit': True, 'linkfield': 'manuscript'},
                       {'formsetClass': MlitFormSet,  'prefix': 'mlit',  'readonly': False, 'noinit': True, 'linkfield': 'manuscript'},
                       {'formsetClass': MprovFormSet, 'prefix': 'mprov', 'readonly': False, 'noinit': True, 'linkfield': 'manuscript'},
                       {'formsetClass': MextFormSet,  'prefix': 'mext',  'readonly': False, 'noinit': True, 'linkfield': 'manuscript'},
                       {'formsetClass': MprojFormSet, 'prefix': 'mproj', 'readonly': False, 'noinit': True, 'linkfield': 'manuscript'}
                       ]
    
    form_objects = [{'form': ManuReconForm, 'prefix': 'mrec', 'readonly': False}]

    stype_edi_fields = ['library', 'lcountry', 'lcity', 'idno', 'origin', 'source', #'project', # PROJECT_MOD_HERE
                        'hierarchy',
                        'LitrefMan', 'litlist',
                        'ManuscriptExt', 'extlist']

    def custom_init(self, instance):
        if instance != None and instance.mtype == "rec":
            # Also adapt the title
            self.titlesg = "Reconstructed manuscript"
        if not instance is None:
            # Make sure to check all codico's
            for codico in instance.manuscriptcodicounits.all():
                codico.check_hierarchy()

        return None

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        oErr = ErrHandle()
        try:
            # Need to know who this user (profile) is
            profile = Profile.get_user_profile(self.request.user.username)
            may_edit = (context['is_app_editor'])

            # Check if this is creating a reconstructed manuscript
            if instance != None and "manu-codicostart" in self.qd:
                # Get the codicostart
                codicostart = self.qd.get("manu-codicostart")
                codico = Codico.objects.filter(id=codicostart).first()
                if codico != None:
                    # Set the mtype to Reconstruction
                    instance.mtype = "rec"
                    instance.save()
                    # Create a Reconstruction object
                    obj = Reconstruction.objects.filter(codico=codico, manuscript=instance).first()
                    if obj == None:
                        obj = Reconstruction.objects.create(codico=codico, manuscript=instance, order=1)

            istemplate = (instance.mtype == "tem")
            # Make sure we mark reconstructed in the context
            context['reconstructed'] = (instance.mtype == "rec")

            # Define the main items to show and edit
            context['mainitems'] = []
            # Possibly add the Template identifier
            if istemplate:
                context['mainitems'].append(
                    {'type': 'plain', 'label': "Template:", 'value': instance.get_template_link(profile)}
                    )
            # Get the main items
            mainitems_main = [
                {'type': 'plain', 'label': "Status:",       'value': instance.get_stype_light(may_edit),'field_key': 'stype'},
                {'type': 'plain', 'label': "Country:",      'value': instance.get_country(),            'field_key': 'lcountry'},
                {'type': 'plain', 'label': "City:",         'value': instance.get_city(),               'field_key': 'lcity',
                 'title': 'City, village or abbey (monastery) of the library'},
                {'type': 'safe',  'label': "Library:",      'value': instance.get_library_markdown(),   'field_key': 'library'},
                {'type': 'plain', 'label': "Shelf mark:",   'value': instance.idno,                     'field_key': 'idno'},
                {'type': 'plain', 'label': "LiLaC code:",   'value': instance.lilacode,                 'field_key': 'lilacode'},
                {'type': 'plain', 'label': "Origin(s):",    'value': instance.get_origins(),            'field_key': 'origins',
                 'title': 'More specific origin information can be entered at the Codicological Unit'},
                {'type': 'plain', 'label': "Date(s):",      'value': instance.dates,                    'field_key': 'dates'},
                {'type': 'plain', 'label': "Script:",       'value': instance.script,                   'field_key': 'script'},
                {'type': 'plain', 'label': "Size:",         'value': instance.size,                     'field_key': 'size'},
                # Project assignment: see below
                ]
            for item in mainitems_main: context['mainitems'].append(item)
            if not istemplate:
                username = profile.user.username
                team_group = app_editor
                mainitems_m2m = [
                    {'type': 'plain', 'label': "Keywords:",     'value': instance.get_keywords_markdown(),      'field_list': 'kwlist'},
                    {'type': 'plain', 'label': "Keywords (user):", 'value': instance.get_keywords_user_markdown(profile),   'field_list': 'ukwlist',
                     'title': 'User-specific keywords. If the moderator accepts these, they move to regular keywords.'},
                    {'type': 'plain', 'label': "Personal Datasets:",  'value': instance.get_collections_markdown(username, team_group, settype="pd"), 
                        'multiple': True, 'field_list': 'collist', 'fso': self.formset_objects[0] },
                    {'type': 'plain', 'label': "Literature:",   'value': instance.get_litrefs_markdown(), 
                        'multiple': True, 'field_list': 'litlist', 'fso': self.formset_objects[1], 'template_selection': 'ru.solemne.litref_template' },

                    # Project HIER
                    {'type': 'plain', 'label': "Project:", 'value': instance.get_project_markdown2()},

                    {'type': 'plain', 'label': "Provenances:",  'value': self.get_provenance_markdown(instance), 
                        'multiple': True, 'field_list': 'mprovlist', 'fso': self.formset_objects[2] }
                    ]
                for item in mainitems_m2m: context['mainitems'].append(item)

                # Possibly append notes view
                if user_is_ingroup(self.request, app_editor):
                    context['mainitems'].append(
                        {'type': 'plain', 'label': "Notes:",       'value': instance.get_notes_markdown(),  'field_key': 'notes'}  )

                # Always append external links and the buttons for codicological unites
                context['mainitems'].append({'type': 'plain', 'label': "External links:",   'value': instance.get_external_markdown(), 
                        'multiple': True, 'field_list': 'extlist', 'fso': self.formset_objects[3] })
                context['mainitems'].append(
                    {'type': 'safe', 'label': 'Codicological unit(s):', 'value': self.get_codico_buttons(instance, context)}
                    )

                # Check if this is an editor with permission for this project
                if may_edit_project(self.request, profile, instance):
                    for oItem in context['mainitems']:
                        if oItem['label'] == "Project:":
                            # Add the list
                            oItem['field_list'] = "projlist"

            # Signal that we have select2
            context['has_select2'] = True

            # Specify that the manuscript info should appear at the right
            title_right = '<span style="font-size: xx-small">{}</span>'.format(instance.get_full_name())
            context['title_right'] = title_right

            # Note: non-app editors may still add a comment
            lhtml = []
            if context['is_app_editor']:
                lbuttons = []
                template_import_button = "import_template_button"
                has_sermons = (instance.manuitems.count() > 0)

                # Action depends on template/not
                if not istemplate:
                    # Also add the manuscript download code
                    local_context = dict(
                        ajaxurl     = reverse("manuscript_download", kwargs={'pk': instance.id}),
                        is_superuser=user_is_superuser(self.request),
                        csrf        = '<input type="hidden" name="csrfmiddlewaretoken" value="{}" />'.format(
                                                get_token(self.request)))
                    lbuttons.append(dict(html=render_to_string('seeker/manuscript_download.html', local_context, self.request)))

                    if instance.mtype != "rec":
                        # Add a button so that the user can import sermons + hierarchy from an existing template
                        if not has_sermons:
                            lbuttons.append(dict(title="Import sermon manifestations from a template", 
                                        id=template_import_button, open="import_from_template", label="Import from template..."))

                        # Add a button so that the user can turn this manuscript into a `Template`
                        lbuttons.append(dict(title="Create template from this manuscript", 
                                     submit="create_new_template", label="Create template..."))
                # Some buttons are needed anyway...
                lbuttons.append(dict(title="Open a list of origins", href=reverse('origin_list'), label="Origins..."))
                lbuttons.append(dict(title="Open a list of locations", href=reverse('location_list'), label="Locations..."))

                # Build the HTML on the basis of the above
                lhtml.append("<div class='row'><div class='col-md-12 container-small' align='right'><form method='post'>")
                for item in lbuttons:
                    if 'html' in item:
                        lhtml.append(item['html'])
                    else:
                        idfield = ""
                        if 'click' in item:
                            ref = " onclick='document.getElementById(\"{}\").click();'".format(item['click'])
                        elif 'submit' in item:
                            ref = " onclick='document.getElementById(\"{}\").submit();'".format(item['submit'])
                        elif 'open' in item:
                            ref = " data-toggle='collapse' data-target='#{}'".format(item['open'])
                        else:
                            ref = " href='{}'".format(item['href'])
                        if 'id' in item:
                            idfield = " id='{}'".format(item['id'])
                        lhtml.append("  <a role='button' class='btn btn-xs jumbo-3' title='{}' {} {}>".format(item['title'], ref, idfield))
                        lhtml.append("     <span class='glyphicon glyphicon-chevron-right'></span>{}</a>".format(item['label']))
                lhtml.append("</form></div></div>")

                if not istemplate:
                    # Add HTML to allow for the *creation* of a template from this manuscript
                    local_context = dict(manubase=instance.id)
                    lhtml.append(render_to_string('seeker/template_create.html', local_context, self.request))

                    # Add HTML to allow the user to choose sermons from a template
                    local_context['frmImport'] = TemplateImportForm({'manu_id': instance.id})
                    local_context['import_button'] = template_import_button
                    lhtml.append(render_to_string('seeker/template_import.html', local_context, self.request))

            #if instance.mtype in ["rec", "man"]:
            # Add Codico items - depending on reconstructed or not
            if instance.mtype == "rec":
                # Note: we need to go through Reconstruction, 
                #       since that table has the correct 'order' values for the reconstruction
                codicos = [x.codico for x in instance.manuscriptreconstructions.all().order_by('order')]
            else:
                # Note: we need to go directly to Codico, since the order values are there
                codicos = instance.manuscriptcodicounits.all().order_by('order')
            codico_list = []
            for codico in codicos:
                # Get the codico details URL
                url = reverse("codico_details", kwargs={'pk': codico.id})
                url_manu = reverse("manuscript_details", kwargs={'pk': codico.manuscript.id})
                # Add the information to the codico list
                codico_list.append( dict(url=url, url_manu=url_manu, kvlist=self.get_kvlist(codico, instance), codico_id=codico.id) )
            context['codico_list'] = codico_list

            # Make sure to add the mtype to the context
            context['mtype'] = instance.mtype
            lhtml.append(render_to_string("seeker/codico_list.html", context, self.request))

            # Add comment modal stuff
            initial = dict(otype="manu", objid=instance.id, profile=profile)
            context['commentForm'] = CommentForm(initial=initial, prefix="com")

            context['comment_list'] = get_usercomments('manu', instance, profile)
            lhtml.append(render_to_string("seeker/comment_add.html", context, self.request))

            # Store the after_details in the context
            context['after_details'] = "\n".join(lhtml)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ManuscriptEdit/add_to_context")

        # Return the context we have made
        return context

    def get_codico_buttons(self, instance, context):
        sBack = ""
        # The number of codico's depends on the mtype of the manuscript
        if instance.mtype == "rec":
            context['codico_count'] = instance.manuscriptreconstructions.count()
        else:
            context['codico_count'] = instance.manuscriptcodicounits.count()
        lhtml = []
        lhtml.append(render_to_string("seeker/codico_buttons.html", context, self.request))
        sBack = "\n".join(lhtml)
        return sBack

    def get_kvlist(self, codico, manu):
        """Get a list of fields and values"""

        lkv = []
        oErr = ErrHandle()
        try:
            # Get a list of sermon information for this codico
            canwit_list = []
            for msitem in codico.codicoitems.all().order_by('order'):
                for canwit in msitem.itemsermons.all():
                    # Add information of this canwit to the list
                    canwit_url = reverse('canwit_details', kwargs={'pk': canwit.id})
                    canwit_html = "<span class='badge signature ot'><a href='{}'>{}</a></span>".format(canwit_url, canwit.locus)
                    canwit_list.append(canwit_html)
            # Action depends on the size of the list
            if len(canwit_list) == 0:
                canwits = "(none)"
            elif len(canwit_list) == 1:
                canwits = canwit_list[0]
            else:
                canwits = "{}...{}".format(canwit_list[0], canwit_list[-1])
            # OLD:
            # canwits = ", ".join(canwit_list)
            lkv = []
            if codico.manuscript.id == manu.id:
                lkv.append(dict(label="Order", value=codico.order))
            else:
                # Get the 'reconstruction' element
                reconstruction = Reconstruction.objects.filter(manuscript=manu, codico=codico).first()
                if reconstruction != None:
                    # sOrder = "{} (in identifier: {})".format(reconstruction.order, codico.order)
                    sOrder = "{}".format(reconstruction.order)
                    lkv.append(dict(label="Order", value=sOrder))
            lkv.append(dict(label="Canon witnesses", value=canwits))
            lkv.append(dict(label="Title", value=codico.name))
            lkv.append(dict(label="Date", value=codico.get_date_markdown()))
            lkv.append(dict(label="Support", value=codico.support))
            lkv.append(dict(label="Extent", value=codico.extent))
            lkv.append(dict(label="Format", value=codico.format))
            lkv.append(dict(label="Keywords", value=codico.get_keywords_markdown()))
            lkv.append(dict(label="Origin", value=self.get_codiorigin_markdown(codico)))
            lkv.append(dict(label="Provenances", value=self.get_codiprovenance_markdown(codico)))
            lkv.append(dict(label="Notes", value=codico.get_notes_markdown()))
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ManuscriptEdit/get_kvlist")
        return lkv

    def get_codiorigin_markdown(self, codico):
        """Calculate a collapsable table view of the origins for this codico, for Codico details view"""

        context = dict(codi=codico)
        sBack = render_to_string("seeker/codi_origins.html", context, self.request)
        return sBack

    def get_codiprovenance_markdown(self, codico):
        """Calculate a collapsable table view of the provenances for this codico, for Codico details view"""

        context = dict(codi=codico)
        sBack = render_to_string("seeker/codi_provs.html", context, self.request)
        return sBack

    def get_provenance_markdown(self, instance):
        """Calculate a collapsible table view of the provenances for this manuscript, for Manu details view"""

        sBack = ""
        oErr = ErrHandle()
        try:
            context = dict(manu=instance)
            sBack = render_to_string("seeker/manu_provs.html", context, self.request)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_provenance_markdown")
        return sBack

    def process_formset(self, prefix, request, formset):
        errors = []
        bResult = True
        instance = formset.instance
        for form in formset:
            if form.is_valid():
                cleaned = form.cleaned_data
                # Action depends on prefix

                #if prefix == "mdr":
                #    # Processing one daterange
                #    newstart = cleaned.get('newstart', None)
                #    newfinish = cleaned.get('newfinish', None)
                #    oneref = cleaned.get('oneref', None)
                #    newpages = cleaned.get('newpages', None)

                #    if newstart:
                #        # Possibly set newfinish equal to newstart
                #        if newfinish == None or newfinish == "":
                #            newfinish = newstart
                #        # Double check if this one already exists for the current instance
                #        obj = instance.manuscript_dateranges.filter(yearstart=newstart, yearfinish=newfinish).first()
                #        if obj == None:
                #            form.instance.yearstart = int(newstart)
                #            form.instance.yearfinish = int(newfinish)
                #        # Do we have a reference?
                #        if oneref != None:
                #            form.instance.reference = oneref
                #            if newpages != None:
                #                form.instance.pages = newpages
                #        # Note: it will get saved with formset.save()
                if prefix == "mcol":
                    # Collection processing
                    newcol = cleaned.get('newcol', None)
                    if newcol != None:
                        # Find out what the profile is
                        profile = Profile.get_user_profile(request.user.username)
                        # Is the COL already existing?
                        obj = Collection.objects.filter(name=newcol).first()
                        if obj == None:
                            # TODO: add profile here
                            obj = Collection.objects.create(name=newcol, type='manu', owner=profile)
                        # once a collection has been created, make sure it gets assigned to a project
                        if not profile is None and obj.projects.count() == 0:
                            # Assign the default projects
                            projects = profile.get_defaults()
                            obj.set_projects(projects)
                        # Make sure we set the keyword
                        form.instance.collection = obj
                        # Note: it will get saved with formset.save()
                elif prefix == "mlit":
                    # Literature reference processing
                    newpages = cleaned.get("newpages")
                    # Also get the litref
                    oneref = cleaned.get("oneref")
                    if oneref:
                        litref = cleaned['oneref']
                        # Check if all is in order
                        if litref:
                            form.instance.reference = litref
                            if newpages:
                                form.instance.pages = newpages
                    # Note: it will get saved with form.save()
                elif prefix == "mprov":
                    # ========= OLD (issue #289) =======
                    #name = cleaned.get("name")
                    #note = cleaned.get("note")
                    #location = cleaned.get("location")
                    #prov_new = cleaned.get("prov_new")proccess
                    #if name:
                    #    obj = Provenance.objects.filter(name=name, note=note, location=location).first()
                    #    if obj == None:
                    #        obj = Provenance.objects.create(name=name)
                    #        if note: obj.note = note
                    #        if location: obj.location = location
                    #        obj.save()
                    #    if obj:
                    #        form.instance.provenance = obj
                    # New method, issue #289 (last part)
                    note = cleaned.get("note")
                    prov_new = cleaned.get("prov_new")
                    if prov_new != None:
                        form.instance.provenance = prov_new
                        form.instance.note = note

                    # Note: it will get saved with form.save()
                elif prefix == "mext":
                    newurl = cleaned.get('newurl')
                    if newurl:
                        form.instance.url = newurl
                elif prefix == "mproj":
                    proj_new = cleaned.get("proj_new")
                    if proj_new != None:
                        form.instance.project = proj_new

                    # Note: it will get saved with [sub]form.save()


            else:
                errors.append(form.errors)
                bResult = False
        return None

    def before_save(self, form, instance):
        oErr = ErrHandle()
        bBack = True
        msg = ""
        try:
            if instance != None:
                # Need to know who is 'talking'...
                username = self.request.user.username
                profile = Profile.get_user_profile(username)

                # If there is no source, then create a source for this one
                if instance.source == None:
                    source = SourceInfo.objects.create(
                        code="Manually created",
                        collector=username, 
                        profile=profile)
                    instance.source = source

                # Issue #473: automatic assignment of project for particular editor(s)
                projlist = form.cleaned_data.get("projlist")
                bBack, msg = evaluate_projlist(profile, instance, projlist, "Manuscript")
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ManuscriptEdit/before_save")
            bBack = False
        return bBack, msg

    def after_save(self, form, instance):
        msg = ""
        bResult = True
        oErr = ErrHandle()
        
        try:
            # Process many-to-many changes: Add and remove relations in accordance with the new set passed on by the user
            # (1) 'collections'
            collist_m = form.cleaned_data['collist']
            adapt_m2m(CollectionMan, instance, "manuscript", collist_m, "collection")

            # (2) 'keywords'
            kwlist = form.cleaned_data['kwlist']
            adapt_m2m(ManuscriptKeyword, instance, "manuscript", kwlist, "keyword")

            # (3) user-specific 'keywords'
            ukwlist = form.cleaned_data['ukwlist']
            profile = Profile.get_user_profile(self.request.user.username)
            adapt_m2m(UserKeyword, instance, "manu", ukwlist, "keyword", qfilter = {'profile': profile}, extrargs = {'profile': profile, 'type': 'manu'})

            # (4) 'literature'
            litlist = form.cleaned_data['litlist']
            adapt_m2m(LitrefMan, instance, "manuscript", litlist, "reference", extra=['pages'], related_is_through = True)

            # (5) 'provenances'Select a provenance...
            mprovlist = form.cleaned_data['mprovlist']
            adapt_m2m(ProvenanceMan, instance, "manuscript", mprovlist, "provenance", extra=['note'], related_is_through = True)

            # (6) 'projects'
            projlist = form.cleaned_data['projlist']
            manu_proj_deleted = []
            adapt_m2m(ManuscriptProject, instance, "manuscript", projlist, "project", deleted=manu_proj_deleted)
            project_dependant_delete(self.request, manu_proj_deleted)

            # When projects have been added to the manuscript, the sermons need to be updated too 
            # or vice versa
            # Issue #412: do *NOT* call this any more
            #             when the project of a manuscript changes, underlying sermons are *not* automatically affected
            # instance.adapt_projects() 

            # Issue #412 + #473: default project assignment
            if instance.projects.count() == 0:
                # Need to know who is 'talking'...
                username = self.request.user.username
                profile = Profile.get_user_profile(username)

                # The user has not selected a project (yet): try default assignment
                user_projects = profile.projects.all()
                if user_projects.count() == 1:
                    project = profile.projects.first()
                    ManuscriptProject.objects.create(manuscript=instance, project=project)
            
            # Process many-to-ONE changes
            # (1) links from SG to SSG
            datelist = form.cleaned_data['datelist']
            adapt_m2o(Daterange, instance, "manuscript", datelist)

            # (2) external URLs
            extlist = form.cleaned_data['extlist']
            adapt_m2o(ManuscriptExt, instance, "manuscript", extlist)

        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg

    def after_new(self, form, instance):
        """When a Manuscript has been created, it needs to get a Codico"""

        bResult = True
        msg = ""
        oErr = ErrHandle()
        try:
            bResult, msg = add_codico_to_manuscript(instance)
        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg 

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        solemne_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return solemne_get_history(instance)
    

class ManuscriptDetails(ManuscriptEdit):
    rtype = "html"

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        template_sermon = 'seeker/sermon_view.html'

        def sermon_object(obj ):
            # Initialisations
            html = ""
            label = ""
            # Check if this points to a sermon
            sermon = obj.itemsermons.first()
            if sermon != None:
                # Calculate the HTML for this sermon
                context = dict(msitem=sermon)
                html = treat_bom( render_to_string(template_sermon, context))
                # Determine what the label is going to be
                label = obj.locus
            else:
                # Determine what the label is going to be
                head = obj.itemheads.first()
                if head != None:
                    label = head.locus
            # Determine the parent, if any
            parent = 1 if obj.parent == None else obj.parent.order + 1
            id = obj.order + 1
            # Create a sermon representation
            oSermon = dict(id=id, parent=parent, pos=label, child=[], f = dict(order=obj.order, html=html))
            return oSermon

        # Start by executing the standard handling
        super(ManuscriptDetails, self).add_to_context(context, instance)

        oErr = ErrHandle()
        method = "one_canwit_list"
        method = "canwit_per_codico"
        try:
            # Additional sections
            context['sections'] = []

            # Lists of related objects
            context['related_objects'] = []

            # Need to know who this user (profile) is
            username = self.request.user.username
            team_group = app_editor

            # Construct the hierarchical list
            canwit_list = instance.get_canwit_list(username, team_group)
 
            # The following only goes for the correct mtype
            if instance.mtype in ["man", "tem"]:

                # Action depends on the method of processing
                if method == "one_canwit_list":
                    # Traditional (Passim) Method: one global canwit_list

                    # Add instances to the list, noting their childof and order
                    context['canwit_list'] = canwit_list
                    context['sermon_count'] = len(canwit_list)
                    # List of codicological unites that are not yet linked to data
                    codi_empty_list = []
                    for codi in instance.manuscriptcodicounits.all().order_by('order'):
                        if codi.codicoitems.count() == 0:
                            codi_empty_list.append(codi)
                    context['codi_empty_list'] = codi_empty_list

                elif method == "canwit_per_codico":
                    # New method for Lilac: canwit_list per codico
                    codi_list = []
                    # Iterate over the codicological units
                    for codico in instance.manuscriptcodicounits.all().order_by('order'):
                        oCodico = dict(codico=codico)
                        oCodico['canwit_list'] = codico.get_canwit_list(username, team_group)
                        codi_list.append(oCodico)
                    context['codi_list'] = codi_list

                    # Get the overal sermon count
                    context['sermon_count'] = Canwit.objects.filter(msitem__codico__manuscript=instance).count()
                    # The empty list is no longer used for this method
                    context['codi_empty_list'] = []
                    # There must also be a canwit list
                    context['canwit_list'] = canwit_list

                # Add the list of sermons and the comment button
                context['add_to_details'] = render_to_string("seeker/manuscript_sermons.html", context, self.request)
            elif instance.mtype == "rec":
                # THis is a reconstruction: show hierarchy view-only
                context['canwit_list'] = canwit_list
                context['sermon_count'] = len(canwit_list)

                # Note: do *NOT* give this list for reconstructions!!
                context['codi_empty_list'] = []

                # Add the list of sermons and the comment button
                context_special = copy.copy(context)
                context_special['is_app_editor'] = False
                context['add_to_details'] = render_to_string("seeker/manuscript_sermons.html", context_special, self.request)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ManuscriptDetails/add_to_context")

        # Return the context we have made
        return context

    def before_save(self, form, instance):
        return True, ""

    def process_formset(self, prefix, request, formset):
        return None

    def after_save(self, form, instance):
        if instance != None and instance.id != None:
            # If no project has been selected, then select the default project(s) - see issue #479
            count = instance.projects.count()
            if count == 0:
                # Set the default projects
                profile = Profile.get_user_profile(self.request.user.username)
                projects = profile.get_defaults()
                instance.set_projects(projects)
        return True, ""


class ManuscriptHierarchy(ManuscriptDetails):
    newRedirect = True

    def custom_init(self, instance):
        errHandle = ErrHandle()

        def getid(item, key, mapping):
            id_value = item[key]
            if "new" in id_value:
                id_value = mapping[id_value]
            obj = MsItem.objects.filter(id=id_value).first()
            id = None if obj == None else obj.id
            return id

        # Note: use [errHandle]
        try:
            # Make sure to set the correct redirect page
            if instance:
                self.redirectpage = reverse("manuscript_details", kwargs={'pk': instance.id})
            # Make sure we are not saving
            self.do_not_save = True
            # Get the [hlist] value
            if 'manu-hlist' in self.qd:
                # Interpret the list of information that we receive
                hlist = json.loads(self.qd['manu-hlist'])
                # Debugging:
                str_hlist = json.dumps(hlist, indent=2)

                # Step 1: Convert any new hierarchical elements into [MsItem] with Codhead
                head_to_id = {}
                deletables = []
                with transaction.atomic():
                    for idx, item in enumerate(hlist):
                        if 'new' in item['id']:
                            # This is a new structural element - Create an MsItem
                            msitem = MsItem.objects.create(manu=instance)
                            # Create a new MsHead item
                            shead = Codhead.objects.create(msitem=msitem,title=item['title'])
                            # Make a mapping
                            head_to_id[item['id']] = msitem.id
                            # Testing
                            id=getid(item, "id", head_to_id)
                        elif 'action' in item and item['action'] == "delete":
                            # This one must be deleted
                            deletables.append(item['id'])

                # Step 1b: adapt the list with deletables
                hlist[:] = [x for x in hlist if x.get("action") != "delete"]

                # Step 1c: delete those that need it
                MsItem.objects.filter(id__in=deletables).delete()

                # Step 2: store the hierarchy
                changes = {}
                hierarchy = []
                codi = None
                with transaction.atomic():
                    for idx, item in enumerate(hlist):
                        bNeedSaving = False

                        # Figure out whether this is the start of a Codi or an MsItem
                        if item.get("id") == "codi":
                            # THis is the start of a codicological unit

                            # Get a possible codi id
                            codi_id = item.get("codi")
                            if codi_id != None:
                                if codi == None or codi.id != codi_id:
                                    codi = Codico.objects.filter(id=codi_id).first()

                            # Safe guarding
                            if codi is None:
                                errHandle.Status("ManuscriptHierarchy: codi is none")
                                x = msitem.itemsermons.first()

                        else:
                            # This must be an MsItem definition

                            # Get the msitem of this item
                            msitem = MsItem.objects.filter(id=getid(item, "id", head_to_id)).first()
                            # Get the next if any
                            next = None if item['nextid'] == "" else MsItem.objects.filter(id=getid(item, "nextid", head_to_id)).first()
                            # Get the first child
                            firstchild = None if item['firstchild'] == "" else MsItem.objects.filter(id=getid(item, "firstchild", head_to_id)).first()
                            # Get the parent
                            parent = None if item['parent'] == "" else MsItem.objects.filter(id=getid(item, "parent", head_to_id)).first()

                            # Possibly set the msitem codi
                            if msitem.codico != codi:
                                msitem.codico = codi
                                bNeedSaving = True
                            elif codi == None and msitem.codico == None:
                                # This MsItem is inserted before something that may already have a codico
                                codi = instance.manuscriptcodicounits.order_by('order').first()
                                if codi != None:
                                    msitem.codico = codi
                                    bNeedSaving = True

                            # Possibly adapt the [shead] title and locus
                            itemhead = msitem.itemheads.first()
                            if itemhead and 'title' in item and 'locus' in item:
                                title= item['title'].strip()
                                locus = item['locus']
                                if itemhead.title != title or itemhead.locus != locus:
                                    itemhead.title = title.strip()
                                    itemhead.locus = locus
                                    # Save the itemhead
                                    itemhead.save()
                            
                            order = idx + 1

                            sermon_id = "none"
                            if msitem.itemsermons.count() > 0:
                                sermon_id = msitem.itemsermons.first().id
                            sermonlog = dict(sermon=sermon_id)
                            bAddSermonLog = False

                            # Check if anytyhing changed
                            if msitem.order != order:
                                # Implement the change
                                msitem.order = order
                                bNeedSaving =True
                            if msitem.parent is not parent:
                                # Track the change
                                old_parent_id = "none" if msitem.parent == None else msitem.parent.id
                                new_parent_id = "none" if parent == None else parent.id
                                if old_parent_id != new_parent_id:
                                    # Track the change
                                    sermonlog['parent_new'] = new_parent_id
                                    sermonlog['parent_old'] = old_parent_id
                                    bAddSermonLog = True

                                    # Implement the change
                                    msitem.parent = parent
                                    bNeedSaving = True
                                else:
                                    no_change = 1

                            if msitem.firstchild != firstchild:
                                # Implement the change
                                msitem.firstchild = firstchild
                                bNeedSaving =True
                            if msitem.next != next:
                                # Track the change
                                old_next_id = "none" if msitem.next == None else msitem.next.id
                                new_next_id = "none" if next == None else next.id
                                sermonlog['next_new'] = new_next_id
                                sermonlog['next_old'] = old_next_id
                                bAddSermonLog = True

                                # Implement the change
                                msitem.next = next
                                bNeedSaving =True
                            # Do we need to save this one?
                            if bNeedSaving:
                                msitem.save()
                                if bAddSermonLog:
                                    # Store the changes
                                    hierarchy.append(sermonlog)

                details = dict(id=instance.id, savetype="change", changes=dict(hierarchy=hierarchy))
                solemne_action_add(self, instance, details, "save")

            return True
        except:
            msg = errHandle.get_error_message()
            errHandle.DoError("ManuscriptHierarchy")
            return False


class ManuscriptCodico(ManuscriptDetails):
    """Link a codico to a manuscript"""
    
    initRedirect = True

    def custom_init(self, instance):
        errHandle = ErrHandle()

        try:
            # Check if the right parameters have been passed on
            if "mrec-rmanu" in self.qd and "mrec-codicostart" in self.qd:
                manu_id = self.qd.get("mrec-rmanu")
                codico_id = self.qd.get("mrec-codicostart")
                if manu_id == None or codico_id == None:
                    # Open another place
                    self.redirectpage = reverse("manuscript_list")
                else:

                    # Check if this thing is already existing
                    obj = Reconstruction.objects.filter(manuscript=manu_id, codico=codico_id).first()
                    if obj == None:
                        # Doesn't exist (yet), so create it
                        order = Reconstruction.objects.filter(manuscript=manu_id).count() + 1
                        obj = Reconstruction.objects.create(manuscript_id=manu_id, codico_id=codico_id, order=order)

                    # Make sure to set the correct redirect page
                    self.redirectpage = reverse("manuscript_details", kwargs={'pk': manu_id})
                    # Make sure we set the object to the reconstruction manuscript
                    self.object = obj.manuscript
            elif "mrec-rcodico" in self.qd and "mrec-manuscript" in self.qd:
                manu_id = self.qd.get("mrec-manuscript")
                codico_id = self.qd.get("mrec-rcodico")
                if manu_id == None or codico_id == None:
                    # Open another place
                    self.redirectpage = reverse("manuscript_list")
                else:
                    # Check if this thing is already existing
                    obj = Reconstruction.objects.filter(manuscript=manu_id, codico=codico_id).first()
                    if obj == None:
                        # Doesn't exist (yet), so create it
                        order = Reconstruction.objects.filter(manuscript=manu_id).count() + 1
                        obj = Reconstruction.objects.create(manuscript_id=manu_id, codico_id=codico_id, order=order)

                    # Make sure to set the correct redirect page
                    self.redirectpage = reverse("manuscript_details", kwargs={'pk': manu_id})
                    # Make sure we set the object to the reconstruction manuscript
                    self.object = obj.manuscript
            elif "mrec-codicolist" in self.qd and "mrec-manuscript" in self.qd:
                manu_id = self.qd.get("mrec-manuscript")
                codico_str = self.qd.get("mrec-codicolist")
                if manu_id == None:
                    # Open another place
                    self.redirectpage = reverse("manuscript_list")
                elif codico_str == None or codico_str == "[]":
                    # Make sure to set the correct redirect page
                    self.redirectpage = reverse("manuscript_details", kwargs={'pk': manu_id})
                else:
                    # Get the actual manuscript
                    manu = Manuscript.objects.filter(id=manu_id).first()
                    # Get the list of codico id's (in their proper order)
                    codico_lst = json.loads(codico_str)
                    # Action depends on the manuscript type
                    if manu.mtype == "rec":
                        # This is a reconstructed manuscript
                        delete_lst = []
                        current_lst = Reconstruction.objects.filter(manuscript=manu_id).order_by("order")
                        for obj in current_lst:
                            if obj.codico.id not in codico_lst:
                                delete_lst.append(obj.id)
                        # Remove those that need deletion
                        if len(delete_lst) > 0:
                            Reconstruction.objects.filter(id__in=delete_lst).delete()
                        # Add and re-order
                        order = 1
                        with transaction.atomic():
                            for id in codico_lst:
                                # Check if this one is there
                                obj = Reconstruction.objects.filter(manuscript=manu_id, codico=id).first()
                                if obj == None:
                                    # Add it
                                    obj = Reconstruction.objects.create(manuscript_id=manu_id, codico_id=id)
                                obj.order = order
                                obj.save()
                                order += 1
                        # Make sure to set the correct redirect page
                        self.redirectpage = reverse("manuscript_details", kwargs={'pk': manu_id})
                    else:
                        # This is a common manuscript (or a template, but I'm not sure that should be allowed)
                        order = 1
                        # (1) Put the codicological unites in the correct order
                        with transaction.atomic():
                            for id in codico_lst:
                                # Get the codico
                                codi = Codico.objects.filter(id=id).first()
                                # Set the correct order
                                codi.order = order
                                codi.save()
                                # Go to the next order count
                                order += 1
                        order = 1
                        # (2) Put the MsItem-s in the correct order
                        with transaction.atomic():
                            for msitem in MsItem.objects.filter(manu=manu).order_by('codico__order', 'order'):
                                msitem.order = order
                                msitem.save()
                                order += 1

                        # Make sure to set the correct redirect page
                        self.redirectpage = reverse("manuscript_details", kwargs={'pk': manu_id})
                    
                    # FOr debugging purposes
                    x = manu.manuscriptcodicounits.all()
            # Return positively
            return True
        except:
            msg = errHandle.get_error_message()
            errHandle.DoError("ManuscriptCodico")
            return False


class ManuscriptListView(BasicList):
    """Search and list manuscripts"""
    
    model = Manuscript
    listform = SearchManuForm
    has_select2 = True
    use_team_group = True
    paginate_by = PAGINATE_BY_VALUE
    bUseFilter = True
    prefix = "manu"
    basketview = False
    template_help = "seeker/filter_help.html"

    order_cols = ['library__lcity__name;library__location__name', 'library__name', 'lilacode', 'idno;name', '', 'yearstart','yearfinish', 'stype','']
    order_default = order_cols
    order_heads = [{'name': 'City/Location',    'order': 'o=1', 'type': 'str', 'custom': 'city',
                    'title': 'City or other location, such as monastery'},
                   {'name': 'Library',  'order': 'o=2', 'type': 'str', 'custom': 'library'},
                   {'name': 'LiLaC',    'order': 'o=3', 'type': 'str', 'custom': 'lilacode', 'linkdetails': True},
                   {'name': 'Name',     'order': 'o=4', 'type': 'str', 'custom': 'name',    'main': True, 'linkdetails': True},
                   {'name': 'Items',    'order': '',    'type': 'int', 'custom': 'count',   'align': 'right'},
                   {'name': 'From',     'order': 'o=6', 'type': 'int', 'custom': 'from',    'align': 'right'},
                   {'name': 'Until',    'order': 'o=7', 'type': 'int', 'custom': 'until',   'align': 'right'},
                   {'name': 'Status',   'order': 'o=8', 'type': 'str', 'custom': 'status'},
                   {'name': '',         'order': '',    'type': 'str', 'custom': 'links'}]
    filters = [ 
        {"name": "Shelfmark",       "id": "filter_manuid",           "enabled": False},
        {"name": "Country",         "id": "filter_country",          "enabled": False},
        {"name": "City/Location",   "id": "filter_city",             "enabled": False},
        {"name": "Library",         "id": "filter_library",          "enabled": False},
        {"name": "Origin",          "id": "filter_origin",           "enabled": False},
        {"name": "Provenance",      "id": "filter_provenance",       "enabled": False},
        {"name": "Date range",      "id": "filter_daterange",        "enabled": False},
        {"name": "Keyword",         "id": "filter_keyword",          "enabled": False},
        {"name": "Status",          "id": "filter_stype",            "enabled": False},
        {"name": "Manuscript",      "id": "filter_manutype",         "enabled": False},
        {"name": "solemne code",     "id": "filter_code",             "enabled": False},
        {"name": "Project",         "id": "filter_project",          "enabled": False},
        {"name": "Sermon...",       "id": "filter_sermon",           "enabled": False, "head_id": "none"},
        {"name": "Collection/Dataset...",   "id": "filter_collection",          "enabled": False, "head_id": "none"},
        {"name": "Gryson or Clavis: manual",    "id": "filter_signature_m",     "enabled": False, "head_id": "filter_sermon"},
        {"name": "Gryson or Clavis: automatic", "id": "filter_signature_a",     "enabled": False, "head_id": "filter_sermon"},
        # issue #10: undefined for any model
        # {"name": "Bible reference",         "id": "filter_bibref",              "enabled": False, "head_id": "filter_sermon"},
        {"name": "Manuscript comparison",   "id": "filter_collection_manuidno", "enabled": False, "head_id": "filter_collection"},
        {"name": "Historical Collection",   "id": "filter_collection_hc",       "enabled": False, "head_id": "filter_collection"},
        {"name": "HC/Manu overlap",         "id": "filter_collection_hcptc",    "enabled": False, "head_id": "filter_collection"},
        {"name": "PD: Manuscript",          "id": "filter_collection_manu",     "enabled": False, "head_id": "filter_collection"},
        {"name": "PD: Sermon",              "id": "filter_collection_sermo",    "enabled": False, "head_id": "filter_collection"},
        {"name": "PD: Authoritative statement",   "id": "filter_collection_super",    "enabled": False, "head_id": "filter_collection"},
      ]

    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'manuid',        'dbfield': 'idno',                   'keyS': 'idno',          'keyList': 'manuidlist', 'infield': 'id'},
            {'filter': 'country',       'fkfield': 'library__lcountry',      'keyS': 'country_ta',    'keyId': 'country',     'keyFk': "name"},
            {'filter': 'city',          'fkfield': 'library__lcity|library__location',         
                                                                             'keyS': 'city_ta',       'keyId': 'city',        'keyFk': "name"},
            {'filter': 'library',       'fkfield': 'library',                'keyS': 'libname_ta',    'keyId': 'library',     'keyFk': "name"},
            {'filter': 'provenance',    'fkfield': 'manuscriptcodicounits__provenances__location|manuscriptcodicounits__provenances',  
             'keyS': 'prov_ta',       'keyId': 'prov',        'keyFk': "name"},
            {'filter': 'origin',        'fkfield': 'manuscriptcodicounits__origins__location|manuscriptcodicounits__origins', 
             # Issue #427. This was: 'manuscriptcodicounits__origin',                 
             'keyS': 'origin_ta',     'keyId': 'origin',      'keyFk': "name"},
            {'filter': 'keyword',       'fkfield': 'keywords',               'keyFk': 'name', 'keyList': 'kwlist', 'infield': 'name' },
            {'filter': 'project',       'fkfield': 'projects',               'keyFk': 'name', 'keyList': 'projlist', 'infield': 'name'},
            {'filter': 'daterange',     'dbfield': 'manuscriptcodicounits__codico_dateranges__yearstart__gte',         'keyS': 'date_from'},
            {'filter': 'daterange',     'dbfield': 'manuscriptcodicounits__codico_dateranges__yearfinish__lte',        'keyS': 'date_until'},
            {'filter': 'code',          'fkfield': 'manuitems__itemsermons__canwit_austat__austat',    'help': 'lilacode',
             'keyS': 'lilacode', 'keyFk': 'code', 'keyList': 'lilalist', 'infield': 'id'},
            {'filter': 'manutype',      'dbfield': 'mtype',                  'keyS': 'manutype', 'keyType': 'fieldchoice', 'infield': 'abbr'},
            {'filter': 'stype',         'dbfield': 'stype',                  'keyList': 'stypelist', 'keyType': 'fieldchoice', 'infield': 'abbr'}
            ]},
        {'section': 'collection', 'filterlist': [
            # === Overlap with a specific manuscript ===
            {'filter': 'collection_manuidno',  'keyS': 'cmpmanu', 'dbfield': 'dbcmpmanu', 'keyList': 'cmpmanuidlist', 'infield': 'id' },
            #{'filter': 'collection_manuptc', 'keyS': 'overlap', 'dbfield': 'hcptc',
            # 'title': 'Percentage overlap between the "Comparison manuscript" SSGs and the SSGs referred to in other manuscripts'},

            # === Historical Collection ===
            {'filter': 'collection_hc',  'fkfield': 'manuitems__itemsermons__austats__collections',                            
             'keyS': 'collection',    'keyFk': 'name', 'keyList': 'collist_hist', 'infield': 'name' },
            {'filter': 'collection_hcptc', 'keyS': 'overlap', 'dbfield': 'hcptc',
             'title': 'Percentage overlap between the Historical Collection SSGs and the SSGs referred to in the manuscripts'},

            # === Personal Dataset ===
            {'filter': 'collection_manu',  'fkfield': 'collections',                            
             'keyS': 'collection',    'keyFk': 'name', 'keyList': 'collist_m', 'infield': 'name' },
            {'filter': 'collection_sermo', 'fkfield': 'manuitems__itemsermons__collections',               
             'keyS': 'collection_s',  'keyFk': 'name', 'keyList': 'collist_s', 'infield': 'name' },
            # Issue #416: Delete the option to search for a GoldSermon dataset 
            #{'filter': 'collection_gold',  'fkfield': 'manuitems__itemsermons__goldsermons__collections',  
            # 'keyS': 'collection_sg', 'keyFk': 'name', 'keyList': 'collist_sg', 'infield': 'name' },
            {'filter': 'collection_super', 'fkfield': 'manuitems__itemsermons__austats__collections', 
             'keyS': 'collection_ssg','keyFk': 'name', 'keyList': 'collist_ssg', 'infield': 'name' },
            # ===================
            ]},
        {'section': 'sermon', 'filterlist': [
            {'filter': 'signature_m', 'fkfield': 'manuitems__itemsermons__sermonsignatures',     'help': 'signature',
             'keyS': 'signature', 'keyFk': 'code', 'keyId': 'signatureid', 'keyList': 'siglist', 'infield': 'code' },
            {'filter': 'signature_a', 'fkfield': 'manuitems__itemsermons__austats__equal_goldsermons__goldsignatures',     'help': 'signature',
             'keyS': 'signaturea', 'keyFk': 'code', 'keyId': 'signatureaid', 'keyList': 'siglist_a', 'infield': 'code' },
            # issue #10: undefined for any model
            #{'filter': 'bibref',    'dbfield': '$dummy', 'keyS': 'bibrefbk'},
            #{'filter': 'bibref',    'dbfield': '$dummy', 'keyS': 'bibrefchvs'}
            ]},
        {'section': 'other', 'filterlist': [
            #{'filter': 'other_project',   'fkfield': 'project',  'keyS': 'project', 'keyFk': 'id', 'keyList': 'prjlist', 'infield': 'name' },
            {'filter': 'source',    'fkfield': 'source',   'keyS': 'source',  'keyFk': 'id', 'keyList': 'srclist', 'infield': 'id' },
            {'filter': 'atype',     'dbfield': 'manuitems__itemsermons__canwit_austat__austat__atype',    'keyS': 'atype'},
            {'filter': 'mtype', 'dbfield': 'mtype', 'keyS': 'mtype'}
            ]}
         ]
    uploads = reader_uploads
    custombuttons = []

    def initializations(self):
        # Possibly add to 'uploads'
        bHasExcel = False
        bHasJson = False
        for item in self.uploads:
            if item['title'] == "excel":
                bHasExcel = True
            elif item['title'] == "json":
                bHasJson = True

        # Should excel be added?
        if not bHasExcel:
            # Add a reference to the Excel upload method
            html = []
            html.append("Import manuscripts from one or more Excel files.")
            html.append("<b>Note 1:</b> this OVERWRITES a manuscript/sermon if it exists!")
            html.append("<b>Note 2:</b> default PROJECT assignment according to Mylila!")
            msg = "<br />".join(html)
            oExcel = dict(title="excel", label="Excel",
                          url=reverse('manuscript_upload_excel'),
                          type="multiple", msg=msg)
            self.uploads.append(oExcel)

        # Should json be added?
        if not bHasJson:
            # Add a reference to the Json upload method
            html = []
            html.append("Import manuscripts from one or more JSON files.")
            html.append("<b>Note 1:</b> this OVERWRITES a manuscript/sermon if it exists!")
            html.append("<b>Note 2:</b> default PROJECT assignment according to Mylila!")
            msg = "<br />".join(html)
            oJson = dict(title="json", label="Json",
                          url=reverse('manuscript_upload_json'),
                          type="multiple", msg=msg)
            self.uploads.append(oJson)

        # Possibly *NOT* show the downloads
        if not user_is_ingroup(self.request, app_developer):
            self.downloads = []
        if not user_is_authenticated(self.request) or not (user_is_superuser(self.request) or user_is_ingroup(self.request, app_moderator)):
            # Do *not* unnecessarily show the custombuttons
            self.custombuttons = []

        # ======== One-time adaptations ==============
        listview_adaptations("manuscript_list")

        return None

    def add_to_context(self, context, initial):
        # Add a files upload form
        context['uploadform'] = UploadFilesForm()

        # Add a form to enter a URL
        context['searchurlform'] = SearchUrlForm()
        
        # Find out who the user is
        profile = Profile.get_user_profile(self.request.user.username)
        context['basketsize'] = 0 if profile == None else profile.basketsize_manu
        context['basket_show'] = reverse('basket_show_manu')
        context['basket_update'] = reverse('basket_update_manu')

        context['colltype'] = "manu"

        return context

    def get_basketqueryset(self):
        if self.basketview:
            profile = Profile.get_user_profile(self.request.user.username)
            qs = profile.basketitems_manu.all()
        else:
            qs = Manuscript.objects.all()
        return qs
    
    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []
        if custom == "city":
            if instance.library:
                city = None
                if instance.library.lcity:
                    city = instance.library.lcity.name
                elif instance.library.location:
                    city = instance.library.location.name
                if city == None:
                    html.append("??")
                    sTitle = "City or location unclear"
                else:
                    html.append("<span>{}</span>".format(city[:12]))        
                    sTitle = city
        elif custom == "library":
            if instance.library:
                lib = instance.library.name
                html.append("<span>{}</span>".format(lib[:12]))  
                sTitle = lib      
        elif custom == "name":
            html.append("<span class='manuscript-idno'>{}</span>".format(instance.idno))
            # THe name should come from the codico unit!!!
            codico = Codico.objects.filter(manuscript=instance).first()
            if codico != None and codico.name != None:
                html.append("<span class='manuscript-title'>| {}</span>".format(codico.name[:100]))
                sTitle = codico.name
        elif custom == "lilacode":
            html.append(instance.get_lilacode())
        elif custom == "count":
            # html.append("{}".format(instance.manusermons.count()))
            html.append("{}".format(instance.get_canwit_count()))
        elif custom == "from":
            # Walk all codico's
            for item in Daterange.objects.filter(codico__manuscript=instance):
                html.append("<div>{}</div>".format(item.yearstart))
        elif custom == "until":
            for item in Daterange.objects.filter(codico__manuscript=instance):
                html.append("<div>{}</div>".format(item.yearfinish))
        elif custom == "status":
            # html.append("<span class='badge'>{}</span>".format(instance.stype[:1]))
            html.append(instance.get_stype_light())
            sTitle = instance.get_stype_display()
        elif custom == "links":
            sLinks = ""
            if instance.url:
                sLinks = "<a role='button' class='btn btn-xs jumbo-1' href='{}'><span class='glyphicon glyphicon-link'><span></a>".format(instance.url)
                sTitle = "External link"
            html.append(sLinks)
        # Combine the HTML code
        sBack = "\n".join(html)
        return sBack, sTitle

    def adapt_search(self, fields):
        
        def get_overlap_ptc(base_ssgs, comp_ssgs):
            """Calculate the overlap percentage between base and comp"""

            total = len(base_ssgs)
            count = 0
            for ssg_id in comp_ssgs:
                if ssg_id in base_ssgs: count += 1
            result = 100 * count / total
            return result

        # Adapt the search to the keywords that *may* be shown
        lstExclude=None
        qAlternative = None

        #prjlist = None # old
        projlist = None

        # Check if a list of keywords is given
        if 'kwlist' in fields and fields['kwlist'] != None and len(fields['kwlist']) > 0:
            # Get the list
            kwlist = fields['kwlist']
            # Get the user
            username = self.request.user.username
            user = User.objects.filter(username=username).first()
            # Check on what kind of user I am
            if not user_is_ingroup(self.request, app_editor):
                # Since I am not an app-editor, I may not filter on keywords that have visibility 'edi'
                kwlist = Keyword.objects.filter(id__in=kwlist).exclude(Q(visibility="edi")).values('id')
                fields['kwlist'] = kwlist
       
        # Check if a list of projects is given
        if 'projlist' in fields and fields['projlist'] != None and len(fields['projlist']) > 0:
            # Get the list
            projlist = fields['projlist']

        ## Check if the prjlist is identified
        #if fields['prjlist'] == None or len(fields['prjlist']) == 0:
        #    # Get the default project
        #    qs = Project.objects.all()
        #    if qs.count() > 0:
        #        prj_default = qs.first()
        #        qs = Project.objects.filter(id=prj_default.id)
        #        fields['prjlist'] = qs
        #        prjlist = qs

        # Check if an overlap percentage is specified
        if 'overlap' in fields and fields['overlap'] != None:
            # Get the overlap
            overlap = fields.get('overlap', "0")
            # Use an overt truth 
            fields['overlap'] = Q(mtype="man")
            if 'collist_hist' in fields and fields['collist_hist'] != None:
                coll_list = fields['collist_hist']
                if len(coll_list) > 0:
                    # Yes, overlap specified
                    if isinstance(overlap, int):
                        # Make sure the string is interpreted as an integer
                        overlap = int(overlap)
                        # Now add a Q expression
                        fields['overlap'] = Q(manu_colloverlaps__overlap__gte=overlap)

                        # Make sure to actually *calculate* the overlap between the different collections and manuscripts
                
                        # (1) Possible manuscripts only filter on: mtype=man, prjlist
                        lstQ = []
                        # if prjlist != None: lstQ.append(Q(project__in=prjlist))
                        lstQ.append(Q(mtype="man"))
                        lstQ.append(Q(manuitems__itemsermons__austats__collections__in=coll_list))
                        manu_list = Manuscript.objects.filter(*lstQ)

                        # We also need to have the profile
                        profile = Profile.get_user_profile(self.request.user.username)
                        # Now calculate the overlap for all
                        with transaction.atomic():
                            for coll in coll_list:
                                for manu in manu_list:
                                    ptc = CollOverlap.get_overlap(profile, coll, manu)
                if 'cmpmanuidlist' in fields and fields['cmpmanuidlist'] != None:
                    # The base manuscripts with which the comparison goes
                    base_manu_list = fields['cmpmanuidlist']
                    if len(base_manu_list) > 0:
                        # Yes, overlap specified
                        if isinstance(overlap, int):
                            # Make sure the string is interpreted as an integer
                            overlap = int(overlap)
                            # Now add a Q expression
                            # fields['overlap'] = Q(manu_colloverlaps__overlap__gte=overlap)
                            # Make sure to actually *calculate* the overlap between the different collections and manuscripts

                            # (1) Get a list of SSGs associated with these manuscripts
                            base_ssg_list = Austat.objects.filter(canwit_austat__canwit__msitem__manu__in=base_manu_list).values('id')
                            base_ssg_list = [x['id'] for x in base_ssg_list]
                            base_count = len(base_ssg_list)
                
                            # (2) Possible overlapping manuscripts only filter on: mtype=man, prjlist and the SSG list
                            lstQ = []
                            # if prjlist != None: lstQ.append(Q(project__in=prjlist))
                            lstQ.append(Q(mtype="man"))
                            lstQ.append(Q(manuitems__itemsermons__austats__id__in=base_ssg_list))
                            manu_list = Manuscript.objects.filter(*lstQ)

                            # We also need to have the profile
                            profile = Profile.get_user_profile(self.request.user.username)
                            # Now calculate the overlap for all
                            manu_include = []
                            with transaction.atomic():
                                for manu in manu_list:
                                    # Get a list of SSG id's associated with this particular manuscript
                                    manu_ssg_list = [x['id'] for x in Austat.objects.filter(canwit_austat__canwit__msitem__manu__id=manu.id).values('id')]
                                    if get_overlap_ptc(base_ssg_list, manu_ssg_list) >= overlap:
                                        # Add this manuscript to the list 
                                        if not manu.id in manu_include:
                                            manu_include.append(manu.id)
                            fields['cmpmanuidlist'] = None
                            fields['cmpmanu'] = Q(id__in=manu_include)


        # Adapt the bible reference list
        bibrefbk = fields.get("bibrefbk", "")
        if bibrefbk != None and bibrefbk != "":
            bibrefchvs = fields.get("bibrefchvs", "")

            # Get the start and end of this bibref
            start, einde = Reference.get_startend(bibrefchvs, book=bibrefbk)

            # Find out which manuscripts have sermons having references in this range
            lstQ = []
            lstQ.append(Q(manuitems__itemsermons__sermonbibranges__bibrangeverses__bkchvs__gte=start))
            lstQ.append(Q(manuitems__itemsermons__sermonbibranges__bibrangeverses__bkchvs__lte=einde))
            manulist = [x.id for x in Manuscript.objects.filter(*lstQ).order_by('id').distinct()]

            fields['bibrefbk'] = Q(id__in=manulist)

        # Make sure we only show manifestations
        # fields['mtype'] = 'man'
        # Make sure we show MANUSCRIPTS (identifiers) as well as reconstructions

        # Make sure we only use the Authoritative statements with accepted modifications
        # This means that atype should be 'acc' (and not: 'mod', 'rej' or 'def')        
        # With this condition we make sure ALL manuscripts are in de unfiltered listview
        print (fields['lilacode'])
        if fields['lilacode'] != '':
            fields['atype'] = 'acc'
       
        lstExclude = [ Q(mtype='tem') ]
        
        return fields, lstExclude, qAlternative

    def view_queryset(self, qs):
        search_id = [x['id'] for x in qs.values('id')]
        profile = Profile.get_user_profile(self.request.user.username)
        profile.search_manu = json.dumps(search_id)
        profile.save()
        return None

    def get_helptext(self, name):
        """Use the get_helptext function defined in models.py"""
        return get_helptext(name)
  

class ManuscriptDownload(BasicPart):
    MainModel = Manuscript
    template_name = "seeker/download_status.html"
    action = "download"
    dtype = "excel"       # downloadtype

    def custom_init(self):
        """Calculate stuff"""
        
        dt = self.qd.get('downloadtype', "")
        if dt != None and dt != '':
            self.dtype = dt

    def get_func(self, instance, path, profile, username, team_group):
        sBack = ""
        #if path == "dateranges":
        #    qs = instance.manuscript_dateranges.all().order_by('yearstart')
        #    dates = []
        #    for obj in qs:
        #        dates.append(obj.__str__())
        #    sBack = ", ".join(dates)
        if path == "keywords":
            sBack = instance.get_keywords_markdown(plain=True)
        elif path == "keywordsU":
            sBack =  instance.get_keywords_user_markdown(profile, plain=True)
        elif path == "datasets":
            sBack = instance.get_collections_markdown(username, team_group, settype="pd", plain=True)
        elif path == "literature":
            sBack = instance.get_litrefs_markdown(plain=True)
        elif path == "origin":
            sBack = instance.get_origins()
        elif path == "provenances":
            sBack = instance.get_provenance_markdown(plain=True)
        elif path == "external":
            sBack = instance.get_external_markdown(plain=True)
        elif path == "brefs":
            sBack = instance.get_bibleref(plain=True)
        elif path == "signaturesM":
            sBack = instance.get_sermonsignatures_markdown(plain=True)
        elif path == "signaturesA":
            sBack = instance.get_eqsetsignatures_markdown(plain=True)
        elif path == "ssglinks":
            sBack = instance.get_eqset()
        return sBack

    def get_data(self, prefix, dtype, response=None):
        """Gather the data as CSV, including a header line and comma-separated"""

        # Initialize
        lData = []
        sData = ""
        manu_fields = []
        oErr = ErrHandle()

        try:
            # Need to know who this user (profile) is
            profile = Profile.get_user_profile(self.request.user.username)
            username = profile.user.username
            team_group = app_editor

            # Make sure we only look at lower-case Dtype
            dtype = dtype.lower()

            # Is this Excel?
            if dtype == "excel" or dtype == "xlsx":
                # Start workbook
                wb = openpyxl.Workbook()

                # First worksheet: MANUSCRIPT itself
                ws = wb.get_active_sheet()
                ws.title="Manuscript"

                # Read the header cells and make a header row in the MANUSCRIPT worksheet
                headers = ["Field", "Value"]
                for col_num in range(len(headers)):
                    c = ws.cell(row=1, column=col_num+1)
                    c.value = headers[col_num]
                    c.font = openpyxl.styles.Font(bold=True)
                    # Set width to a fixed size
                    ws.column_dimensions[get_column_letter(col_num+1)].width = 5.0        

                # Walk the mainitems
                row_num = 2
                kwargs = {'profile': profile, 'username': username, 'team_group': team_group}
                for item in Manuscript.specification:
                    key, value = self.obj.custom_getkv(item, kwargs=kwargs)
                    # Add the K/V row
                    ws.cell(row=row_num, column = 1).value = key
                    ws.cell(row=row_num, column = 2).value = value
                    row_num += 1

                # Second worksheet: ALL SERMONS in the manuscript
                ws = wb.create_sheet("Sermons")

                # Read the header cells and make a header row in the SERMON worksheet
                headers = [x['name'] for x in Canwit.specification ]
                for col_num in range(len(headers)):
                    c = ws.cell(row=1, column=col_num+1)
                    c.value = headers[col_num]
                    c.font = openpyxl.styles.Font(bold=True)
                    # Set width to a fixed size
                    ws.column_dimensions[get_column_letter(col_num+1)].width = 5.0        

                row_num = 1
                # Walk all msitems of this manuscript
                for msitem in self.obj.manuitems.all().order_by('order'):
                    row_num += 1
                    col_num = 1
                    ws.cell(row=row_num, column=col_num).value = msitem.order
                    # Get other stuff
                    parent = "" if msitem.parent == None else msitem.parent.order
                    firstchild = "" if msitem.firstchild == None else msitem.firstchild.order
                    next = "" if msitem.next == None else msitem.next.order

                    # Process the structural elements
                    col_num += 1
                    ws.cell(row=row_num, column=col_num).value = parent
                    col_num += 1
                    ws.cell(row=row_num, column=col_num).value = firstchild
                    col_num += 1
                    ws.cell(row=row_num, column=col_num).value = next

                    # What kind of item is this?
                    col_num += 1
                    if msitem.itemheads.count() > 0:
                        codhead = msitem.itemheads.first()
                        # This is a Codhead
                        ws.cell(row=row_num, column=col_num).value = "Structural"
                        col_num += 2
                        ws.cell(row=row_num, column=col_num).value = codhead.locus
                        col_num += 4
                        ws.cell(row=row_num, column=col_num).value = codhead.title.strip()
                    else:
                        # This is a Canwit
                        ws.cell(row=row_num, column=col_num).value = "Plain"
                        col_num += 1
                        sermon = msitem.itemsermons.first()
                        # Walk the items
                        for item in Canwit.specification:
                            if item['type'] != "":
                                key, value = sermon.custom_getkv(item, kwargs=kwargs)
                                ws.cell(row=row_num, column=col_num).value = value
                                col_num += 1
                

                # Save it
                wb.save(response)
                sData = response
            elif dtype == "json":
                # Start a *list* of manuscripts
                #  (so that we have one generic format for both a single as well as a number of manuscripts)
                lst_manu = []

                # Start one object for this particular manuscript
                oManu = dict(msitems=[])

                # Walk the mainitems
                kwargs = {'profile': profile, 'username': username, 'team_group': team_group, 'keyfield': 'path'}
                for item in Manuscript.specification:
                    # Only skip key_id items
                    if item['type'] != "fk_id":
                        key, value = self.obj.custom_getkv(item, **kwargs)
                        # Add the K/V row
                        oManu[key] = value

                # Walk all msitems of this manuscript
                for msitem in self.obj.manuitems.all().order_by('order'):
                    # Create an object for this sermon
                    oMsItem = {}

                    # Add the order of this item as well as he parent, firstchild, next
                    oMsItem['order'] = msitem.order
                    oMsItem['parent'] = "" if msitem.parent == None else msitem.parent.order
                    oMsItem['firstchild'] = "" if msitem.firstchild == None else msitem.firstchild.order
                    oMsItem['next'] = "" if msitem.next == None else msitem.next.order

                    # Create an object for this sermon
                    oSermon = {}

                    # What kind of item is this?
                    if msitem.itemheads.count() > 0:
                        codhead = msitem.itemheads.first()
                        # This is a Codhead
                        oSermon['type'] = "Structural"
                        oSermon['locus'] = codhead.locus
                        oSermon['title'] = codhead.title.strip()
                    else:
                        # This is a Canwit
                        oSermon['type'] = "Plain"

                        # Get the actual sermon
                        sermon = msitem.itemsermons.first()
                        # Walk the items of this sermon (defined in specification)
                        for item in Canwit.specification:
                            if item['type'] != "" and item['type'] != "fk_id":
                                key, value = sermon.custom_getkv(item, **kwargs)
                                oSermon[key] = value
                    # Add sermon to msitem
                    oMsItem['sermon'] = oSermon
                    # Add this sermon to the list of sermons within the manuscript
                    oManu['msitems'].append(oMsItem)

                # Add object to the list
                lst_manu.append(oManu)
                # Make sure to return this list
                sData = json.dumps( lst_manu, indent=2)
            elif dtype == "tei" or dtype== "xml-tei":
                # Prepare a context for the XML creation
                context = dict(details_id=self.obj.id, download_person=username)
                context['details_url'] = 'https://solemne.rich.ru.nl{}'.format(reverse('manuscript_details', kwargs={'pk': self.obj.id}))
                context['download_date_ymd'] = get_current_datetime().strftime("%Y-%m-%d")
                context['download_date'] = get_current_datetime().strftime("%d/%b/%Y")
                context['manu'] = self.obj

                # Convert into string
                sData = render_to_string("seeker/tei-template.xml", context, self.request)

                # Perform pretty printing
                tree = ET.fromstring(sData, parser=ET.XMLParser(encoding='utf-8', remove_blank_text=True))
                pretty = ET.tostring(tree, encoding="utf-8", pretty_print=True, xml_declaration=True)
                sData = pretty
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ManuscriptDownload/get_data")


        return sData


# ============= CODICOLOGICAL UNIT VIEWS ============================


class CodicoEdit(BasicDetails):
    """The details of one codicological unit"""

    model = Codico  
    mForm = CodicoForm
    prefix = 'codi'
    title = "codicological unit"
    rtype = "json"
    prefix_type = "simple"
    new_button = True
    backbutton = False
    mainitems = []
    use_team_group = True
    history_button = True
    
    CdrFormSet = inlineformset_factory(Codico, Daterange,
                                         form=DaterangeForm, min_num=0,
                                         fk_name = "codico",
                                         extra=0, can_delete=True, can_order=False)
    CprovFormSet = inlineformset_factory(Codico, ProvenanceCod,
                                         form=CodicoProvForm, min_num=0,
                                         fk_name = "codico",
                                         extra=0, can_delete=True, can_order=False)
    CoriFormSet = inlineformset_factory(Codico, OriginCodico,
                                         form=CodicoOriginForm, min_num=0, max_num=1,
                                         fk_name = "codico",
                                         extra=0, can_delete=True, can_order=False)

    formset_objects = [{'formsetClass': CdrFormSet,   'prefix': 'cdr',   'readonly': False, 'noinit': True, 'linkfield': 'codico'},
                       {'formsetClass': CprovFormSet, 'prefix': 'cprov', 'readonly': False, 'noinit': True, 'linkfield': 'codico'},
                       {'formsetClass': CoriFormSet,  'prefix': 'cori',  'readonly': False, 'noinit': True, 'linkfield': 'codico'}]

    stype_edi_fields = ['name', 'order', 'origin', 'support', 'extent', 'format', 
                        'Daterange', 'datelist',
                        'ProvenanceCod', 'cprovlist',
                        'OriginCodico', 'corilist']
    
    def custom_init(self, instance):
        if instance != None:
            manu_id = instance.manuscript.id
            # Also make sure to change the afterdelurl
            self.afterdelurl = reverse("manuscript_details", kwargs={'pk': manu_id})

        return None

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        oErr = ErrHandle()
        try:
            # Need to know who this user (profile) is
            profile = Profile.get_user_profile(self.request.user.username)
            may_edit = (context['is_app_editor'])

            # Define the main items to show and edit
            context['mainitems'] = []

            manu_id = None if instance == None else instance.manuscript.id

            # Add a button back to the Manuscript
            topleftlist = []
            if manu_id != None:
                buttonspecs = {'label': "M", 
                     'title': "Go to manuscript {}".format(instance.manuscript.idno), 
                     'url': reverse('manuscript_details', kwargs={'pk': manu_id})}
                topleftlist.append(buttonspecs)

                ## Also make sure to change the afterdelurl
                #context['afterdelurl'] = reverse("manuscript_details", kwargs={'pk': manu_id})

                # Check if this is the *first* codico of the manuscript
                if instance.manuscript.manuscriptcodicounits.all().order_by("order").first().id == instance.id:
                    # Make sure deleting is not allowed
                    context['no_delete'] = True
            context['topleftbuttons'] = topleftlist

            # Get the main items
            mainitems_main = [
                {'type': 'plain', 'label': "Status:",       'value': instance.get_stype_light(may_edit),'field_key': 'stype'},
                # -------- HIDDEN field values ---------------
                {'type': 'plain', 'label': "Manuscript id", 'value': manu_id,   'field_key': "manuscript",  'empty': 'hide'},
                # --------------------------------------------
                {'type': 'plain', 'label': "Manuscript:",   'value': instance.get_manu_markdown()},
                {'type': 'plain', 'label': "Title:",        'value': instance.name,                     'field_key': 'name'},
                {'type': 'safe',  'label': "Order:",        'value': instance.order             },
                {'type': 'line',  'label': "Date:",         'value': instance.get_date_markdown(), 
                 'multiple': True, 'field_list': 'datelist', 'fso': self.formset_objects[0]},   #, 'template_selection': 'ru.solemne.litref_template' },
                {'type': 'plain', 'label': "Support:",      'value': instance.support,                  'field_key': 'support'},
                {'type': 'plain', 'label': "Extent:",       'value': instance.extent,                   'field_key': 'extent'},
                {'type': 'plain', 'label': "Size:",         'value': instance.format,                   'field_key': 'format'},
                {'type': 'plain', 'label': "Project:",      'value': instance.get_project_markdown2()}
                ]
            for item in mainitems_main: context['mainitems'].append(item)
            username = profile.user.username
            team_group = app_editor
            mainitems_m2m = [
                {'type': 'plain', 'label': "Keywords:",     'value': instance.get_keywords_markdown(),  'field_list': 'kwlist'},
                # Was: (see issue #427)
                #      {'type': 'safe',  'label': "Origin:",       'value': instance.get_origin_markdown(),    'field_key': 'origin'},
                {'type': 'plain', 'label': "Origin:",       'value': self.get_origin_markdown(instance),    
                 'multiple': True, 'field_list': 'corilist', 'fso': self.formset_objects[2]},
                {'type': 'plain', 'label': "Provenances:",  'value': self.get_provenance_markdown(instance), 
                 'multiple': True, 'field_list': 'cprovlist', 'fso': self.formset_objects[1] }
                ]
            for item in mainitems_m2m: context['mainitems'].append(item)

            # Possibly append notes view
            if user_is_ingroup(self.request, app_editor):
                context['mainitems'].append(
                    {'type': 'plain', 'label': "Notes:",       'value': instance.get_notes_markdown(),  'field_key': 'notes'}  )

            # Signal that we have select2
            context['has_select2'] = True

            # Specify that the manuscript info should appear at the right
            title_right = '<span style="font-size: xx-small">{}</span>'.format(instance.get_full_name())
            context['title_right'] = title_right

            # Note: non-app editors may still add a comment
            lhtml = []
            if context['is_app_editor']:
                lbuttons = []

                # Some buttons are needed anyway...
                lbuttons.append(dict(title="Open a list of origins", href=reverse('origin_list'), label="Origins..."))
                lbuttons.append(dict(title="Open a list of locations", href=reverse('location_list'), label="Locations..."))

                # Build the HTML on the basis of the above
                lhtml.append("<div class='row'><div class='col-md-12' align='right'>")
                for item in lbuttons:
                    idfield = ""
                    if 'click' in item:
                        ref = " onclick='document.getElementById(\"{}\").click();'".format(item['click'])
                    elif 'submit' in item:
                        ref = " onclick='document.getElementById(\"{}\").submit();'".format(item['submit'])
                    elif 'open' in item:
                        ref = " data-toggle='collapse' data-target='#{}'".format(item['open'])
                    else:
                        ref = " href='{}'".format(item['href'])
                    if 'id' in item:
                        idfield = " id='{}'".format(item['id'])
                    lhtml.append("  <a role='button' class='btn btn-xs jumbo-3' title='{}' {} {}>".format(item['title'], ref, idfield))
                    lhtml.append("     <span class='glyphicon glyphicon-chevron-right'></span>{}</a>".format(item['label']))
                lhtml.append("</div></div>")

            # Add comment modal stuff
            initial = dict(otype="codi", objid=instance.id, profile=profile)
            context['commentForm'] = CommentForm(initial=initial, prefix="com")

            context['comment_list'] = get_usercomments('codi', instance, profile)
            lhtml.append(render_to_string("seeker/comment_add.html", context, self.request))

            # Store the after_details in the context
            context['after_details'] = "\n".join(lhtml)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("CodicoEdit/add_to_context")

        # Return the context we have made
        return context

    def get_origin_markdown(self, instance):
        """Calculate a collapsable table view of the origins for this codico, for Codico details view"""

        context = dict(codi=instance)
        sBack = render_to_string("seeker/codi_origins.html", context, self.request)
        return sBack

    def get_provenance_markdown(self, instance):
        """Calculate a collapsable table view of the provenances for this codico, for Codico details view"""

        context = dict(codi=instance)
        sBack = render_to_string("seeker/codi_provs.html", context, self.request)
        return sBack

    def process_formset(self, prefix, request, formset):
        errors = []
        bResult = True
        instance = formset.instance
        for form in formset:
            if form.is_valid():
                cleaned = form.cleaned_data
                # Action depends on prefix

                if prefix == "cdr":
                    # Processing one daterange
                    newstart = cleaned.get('newstart', None)
                    newfinish = cleaned.get('newfinish', None)
                    oneref = cleaned.get('oneref', None)
                    newpages = cleaned.get('newpages', None)

                    if newstart:
                        # Possibly set newfinish equal to newstart
                        if newfinish == None or newfinish == "":
                            newfinish = newstart
                        # Double check if this one already exists for the current instance
                        obj = instance.codico_dateranges.filter(yearstart=newstart, yearfinish=newfinish).first()
                        if obj == None:
                            form.instance.yearstart = int(newstart)
                            form.instance.yearfinish = int(newfinish)
                        # Do we have a reference?
                        if oneref != None:
                            form.instance.reference = oneref
                            if newpages != None:
                                form.instance.pages = newpages
                        # Note: it will get saved with formset.save()
                elif prefix == "cprov":
                    # New method, issue #289 (last part)
                    note = cleaned.get("note")
                    prov_new = cleaned.get("prov_new")
                    if prov_new != None:
                        form.instance.provenance = prov_new
                        form.instance.note = note
                elif prefix == "cori":
                    # Don't allow more than one origin
                    count = instance.origins.count()
                    if count < 1:
                        # See issue #427
                        note = cleaned.get("note")
                        origin_new = cleaned.get("origin_new")
                        if origin_new != None:
                            form.instance.origin = origin_new
                            form.instance.note = note
                    else:
                        errors.append("A codicological unit may not have more than one Origin")
            else:
                errors.append(form.errors)
                bResult = False
        return None

    def before_save(self, form, instance):
        if instance != None:
            # Double check for the correct 'order'
            if instance.order <= 0:
                # Calculate how many Codicos (!) there are
                codico_count = instance.manuscript.manuscriptcodicounits.count()
                # Adapt the order of this codico
                instance.order = codico_count + 1
                # The number will be automatically saved
        return True, ""

    def after_save(self, form, instance):
        msg = ""
        bResult = True
        oErr = ErrHandle()
        
        try:
            # Process many-to-many changes: Add and remove relations in accordance with the new set passed on by the user
            # (1) 'keywords'
            kwlist = form.cleaned_data['kwlist']
            adapt_m2m(CodicoKeyword, instance, "codico", kwlist, "keyword")

            # (2) 'provenances'
            cprovlist = form.cleaned_data['cprovlist']
            adapt_m2m(ProvenanceCod, instance, "codico", cprovlist, "provenance", extra=['note'], related_is_through = True)

            # (3) 'origins'
            corilist = form.cleaned_data['corilist']
            adapt_m2m(OriginCodico, instance, "codico", corilist, "origin", extra=['note'], related_is_through = True)

            # Process many-to-ONE changes
            # (1) links from Daterange to Codico
            datelist = form.cleaned_data['datelist']
            adapt_m2o(Daterange, instance, "codico", datelist)

            # Make sure to process changes
            instance.refresh_from_db()
        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        solemne_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return solemne_get_history(instance)


class CodicoDetails(CodicoEdit):
    rtype = "html"

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Start by executing the standard handling
        super(CodicoDetails, self).add_to_context(context, instance)

        oErr = ErrHandle()
        try:
            # Additional sections
            context['sections'] = []

            # Lists of related objects
            context['related_objects'] = []

        except:
            msg = oErr.get_error_message()
            oErr.DoError("CodicoDetails/add_to_context")

        # Return the context we have made
        return context

    def process_formset(self, prefix, request, formset):
        return None

    def after_save(self, form, instance):
        return True, ""


class CodicoListView(BasicList):
    """Search and list manuscripts"""
    
    model = Codico
    listform = CodicoForm
    has_select2 = True
    use_team_group = True
    paginate_by = PAGINATE_BY_VALUE
    bUseFilter = True
    prefix = "codi"
    template_help = "seeker/filter_help.html"
    order_cols = ['manuscript__idno', 'name', 'order', 'yearstart','yearfinish', 'stype']
    order_default = order_cols
    order_heads = [{'name': 'Manuscript', 'order': 'o=1', 'type': 'str', 'custom': 'manu'},
                   {'name': 'Title',     'order': 'o=2', 'type': 'str', 'custom': 'name', 'main': True, 'linkdetails': True},
                   {'name': 'Unit',     'order': 'o=3', 'type': 'int', 'custom': 'order',   'align': 'right'},
                   {'name': 'From',     'order': 'o=4', 'type': 'int', 'custom': 'from',    'align': 'right'},
                   {'name': 'Until',    'order': 'o=5', 'type': 'int', 'custom': 'until',   'align': 'right'},
                   {'name': 'Status',   'order': 'o=6', 'type': 'str', 'custom': 'status'}]
    filters = [ 
        {"name": "Shelfmark",       "id": "filter_manuid",           "enabled": False},
        {"name": "Title",           "id": "filter_title",            "enabled": False},
        {"name": "Origin",          "id": "filter_origin",           "enabled": False},
        {"name": "Provenance",      "id": "filter_provenance",       "enabled": False},
        {"name": "Date range",      "id": "filter_daterange",        "enabled": False},
        {"name": "Keyword",         "id": "filter_keyword",          "enabled": False},
        {"name": "Status",          "id": "filter_stype",            "enabled": False},
        {"name": "Project",         "id": "filter_project",          "enabled": False, "head_id": "filter_other"},
      ]

    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'manuid',        'fkfield': 'manuscript',       'keyS': 'manuidno',      
             'keyFk': 'idno', 'keyList': 'manuidlist', 'infield': 'id'},
            {'filter': 'provenance',    'fkfield': 'provenances__location',  'keyS': 'prov_ta',       'keyId': 'prov',        'keyFk': "name"},
            {'filter': 'title',         'dbfield': 'name',                   'keyS': 'name_ta'},
            {'filter': 'origin',        'fkfield': 'origin',                 'keyS': 'origin_ta',     'keyId': 'origin',      'keyFk': "name"},
            {'filter': 'keyword',       'fkfield': 'keywords',               'keyFk': 'name', 'keyList': 'kwlist', 'infield': 'name' },
            {'filter': 'daterange',     'dbfield': 'codico_dateranges__yearstart__gte',         'keyS': 'date_from'},
            {'filter': 'daterange',     'dbfield': 'codico_dateranges__yearfinish__lte',        'keyS': 'date_until'},
            {'filter': 'stype',         'dbfield': 'stype',                  'keyList': 'stypelist', 'keyType': 'fieldchoice', 'infield': 'abbr' }
            ]},
        {'section': 'other', 'filterlist': [
            #{'filter': 'project',   'fkfield': 'manuscript__project',  'keyS': 'project', 'keyFk': 'id', 'keyList': 'prjlist', 'infield': 'name' }
            ]}
         ]

    def add_to_context(self, context, initial):

        # Add a form to enter a URL
        context['searchurlform'] = SearchUrlForm()
        
        return context

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []
        if custom == "manu":
            if instance.manuscript != None:
                url = reverse("manuscript_details", kwargs={'pk': instance.manuscript.id})
                html.append("<span class='manuscript-idno'><a href='{}'>{}</a></span>".format(url,instance.manuscript.idno))
        elif custom == "name":
            if instance.name:
                html.append("<span class='manuscript-title'>{}</span>".format(instance.name[:100]))
                sTitle = instance.name
        elif custom == "order":
            html.append("{}".format(instance.order))
        elif custom == "from":
            for item in instance.codico_dateranges.all():
                html.append("<div>{}</div>".format(item.yearstart))
        elif custom == "until":
            for item in instance.codico_dateranges.all():
                html.append("<div>{}</div>".format(item.yearfinish))
        elif custom == "status":
            html.append(instance.get_stype_light())
            sTitle = instance.get_stype_display()
        # Combine the HTML code
        sBack = "\n".join(html)
        return sBack, sTitle

    def adapt_search(self, fields):
        # Adapt the search to the keywords that *may* be shown
        lstExclude=None
        qAlternative = None

        # prjlist = None
        # Check if a list of keywords is given
        if 'kwlist' in fields and fields['kwlist'] != None and len(fields['kwlist']) > 0:
            # Get the list
            kwlist = fields['kwlist']
            # Get the user
            username = self.request.user.username
            user = User.objects.filter(username=username).first()
            # Check on what kind of user I am
            if not user_is_ingroup(self.request, app_editor):
                # Since I am not an app-editor, I may not filter on keywords that have visibility 'edi'
                kwlist = Keyword.objects.filter(id__in=kwlist).exclude(Q(visibility="edi")).values('id')
                fields['kwlist'] = kwlist

        return fields, lstExclude, qAlternative

    def get_helptext(self, name):
        """Use the get_helptext function defined in models.py"""
        return get_helptext(name)


# ============= COLLECTION WITNESS VIEWS ============================


class ColwitEdit(BasicDetails):
    """The editable part of one sermon description (manifestation)"""
    
    model = Colwit
    mForm = ColwitForm
    prefix = "colw"
    title = "Collection Witness" 
    rtype = "json"
    mainitems = []
    basic_name = "colwit"
    use_team_group = True
    history_button = True
    prefix_type = "simple"

    CsigFormSet = inlineformset_factory(Colwit, ColwitSignature,
                                         form = ColwitSignatureForm, min_num=0,
                                         fk_name = "colwit",
                                         extra=0, can_delete=True, can_order=False)

    formset_objects = [
        {'formsetClass': CsigFormSet,  'prefix': 'csig',  'readonly': False, 'noinit': True, 'linkfield': 'colwit'}
        ]

    stype_edi_fields = ['codhead', 'collection', 'descr', 'notes',
                        ]

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        oErr = ErrHandle()
        try:
            # Need to know who this user (profile) is
            profile = Profile.get_user_profile(self.request.user.username)

            # Define the main items to show and edit
            coll_id = None if instance == None else instance.collection.id
            codh_id = None if instance == None else instance.codhead.id
            context['mainitems'] = [
                # -------- HIDDEN field values ---------------
                {'type': 'plain', 'label': "Codhead id",            'value': codh_id,           'field_key': "codhead",  'empty': 'hide'},
                {'type': 'plain', 'label': "Collection id",         'value': coll_id,           'field_key': "collection",  'empty': 'hide'},
                # --------------------------------------------
                {'type': 'safe',  'label': "Manuscript:",           'value': instance.get_manuscript()},
                {'type': 'safe',  'label': "Manuscript section:",   'value': instance.get_codhead()},
                {'type': 'safe',  'label': "Collection:",           'value': instance.get_collection(), 'field_key': "collone"},
                {'type': 'safe',  'label': "LiLaC code:",           'value': instance.get_lilacode()    },
                {'type': 'plain', 'label': "Description:",          'value': instance.descr,    'field_key': "descr"}, 
                {'type': 'line',  'label': "Signatures (Clavis):", 'value': instance.get_signatures(),
                 'multiple': True, 'field_list': 'siglist', 'fso': self.formset_objects[0]  },
                {'type': 'plain', 'label': "Notes:",                'value': instance.notes,    'field_key': 'notes'},
                ]

            # Add a button back to the Manuscript this ColWit is part of
            topleftlist = []
            if instance.get_manuscript_obj():
                manu = instance.get_manuscript_obj()
                buttonspecs = {'label': "M", 
                     'title': "Go to manuscript {}".format(manu.idno), 
                     'url': reverse('manuscript_details', kwargs={'pk': manu.id})}
                topleftlist.append(buttonspecs)
            context['topleftbuttons'] = topleftlist

            # Signal that we have select2
            context['has_select2'] = True

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ColwitEdit/add_to_context")

        # Return the context we have made
        return context

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        solemne_action_add(self, instance, details, actiontype)

    def after_save(self, form, instance):
        msg = ""
        bResult = True
        oErr = ErrHandle()
                
        try:
            # Process many-to-many changes: Add and remove relations in accordance with the new set passed on by the user
            # (1) 'signatures'
            siglist = form.cleaned_data['siglist']
            adapt_m2m(ColwitSignature, instance, "colwit", siglist, "signature")

        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg

    def before_save(self, form, instance):
        oErr = ErrHandle()
        bBack = True
        msg = ""
        try:
            # Process any change in choice of collection
            if not instance is None:
                # Get the cleaned data
                cleaned_data = form.cleaned_data
                # Look for [collone]
                collone = cleaned_data.get('collone')
                if not collone is None:
                    # Compare with existing
                    if collone.id != instance.collection.id:
                        form.instance.collection = collone

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ColwitEdit/before_save")
            bBack = False
        return bBack, msg

    def get_history(self, instance):
        return solemne_get_history(instance)

    def process_formset(self, prefix, request, formset):

        errors = []
        bResult = True
        instance = formset.instance
        for form in formset:
            if form.is_valid():
                cleaned = form.cleaned_data

                oErr = ErrHandle()
                try:

                    if prefix == "csig":
                        # Processing of a signature
                        newsig = cleaned.get("newsig")
                        newedi = cleaned.get("newedi")
                        if not newsig is None and not newedi is None:
                            # We have a new signature and editype
                            sig = Signature.objects.filter(editype=newedi, code=newsig).first()
                            if sig is None:
                                # Create it
                                sig = Signature.objects.create(editype=newedi, code=newsig)
                            # Make sure the correct values are set in the ColwitSignatureForm
                            form.instance.auwork = instance
                            form.instance.signature = sig
                except:
                    msg = oErr.get_error_message()
                    oErr.DoError("ColwitEdit/process_formset")
            else:
                errors.append(form.errors)
                bResult = False
        return None


class ColwitDetails(ColwitEdit):
    """The details of one sermon manifestation (Colwit)"""

    rtype = "html"

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Start by executing the standard handling
        context = super(ColwitDetails, self).add_to_context(context, instance)

        oErr = ErrHandle()
        sort_start = ""
        sort_start_mix = ""
        sort_start_int = ""
        sort_end = ""

        try:
            if instance != None and instance.id != None:
                context['sections'] = []

                # Lists of related objects
                related_objects = []
                resizable = True

                username = self.request.user.username
                team_group = app_editor
                profile = Profile.get_user_profile(username=username)

                # Authorization: only app-editors may edit!
                bMayEdit = user_is_ingroup(self.request, team_group)
            
                # All PDs: show the content
                if bMayEdit:
                    sort_start = '<span class="sortable"><span class="fa fa-sort sortshow"></span>&nbsp;'
                    sort_start_mix = '<span class="sortable mixed"><span class="fa fa-sort sortshow"></span>&nbsp;'
                    sort_start_int = '<span class="sortable integer"><span class="fa fa-sort sortshow"></span>&nbsp;'
                    sort_end = '</span>'

                ## ============= List of historical collections related to the Austat  ==============================
                #collections = dict(title="Historical collections", prefix="hist", gridclass="resizable", classes="")

                ## Get all historical collections (including private ones)
                #qs_hc = instance.collections.filter(settype="hc").order_by("name")
                #qs_hc = Caned.objects.filter(austat=instance).order_by(
                #    "collection__name")
                #rel_list = []
                #for obj in qs_hc:
                #    rel_item = []
                #    # The [obj] is a Caned. Now get to the actual Collection
                #    item = obj.collection
                    
                #    # Make sure we have the link to the HC
                #    url = reverse("collhist_details", kwargs={'pk': item.id})

                #    # HC: Order of Austat within collection
                #    add_rel_item(rel_item, obj.order, False, align="right")

                #    # HC: Name of collection
                #    add_rel_item(rel_item, self.get_field_value("collection", item, "name"), resizable, link=url)

                #    # HC: Manuscript + Canonical witness linked to collection
                #    add_rel_item(rel_item, self.get_field_value("collection", item, "canwits"), resizable, main=True, nowrap=False)

                #    # HC: Owner of collection
                #    add_rel_item(rel_item, self.get_field_value("collection", item, "owner"), resizable, link=url)

                #    # HC: Scope of collection
                #    add_rel_item(rel_item, self.get_field_value("collection", item, "scope"), resizable, link=url)

                #    # HC: Number of authors
                #    add_rel_item(rel_item, self.get_field_value("collection", item, "authnum"), resizable, link=url, align="right")


                #    # Add this line to the list
                #    rel_list.append(dict(id=item.id, cols=rel_item))

                #collections['rel_list'] = rel_list

                #collections['columns'] = [
                #    '{}<span title="Order">Order<span>{}'.format(sort_start_int, sort_end),
                #    '{}<span title="Name of the historical collection">Name</span>{}'.format(sort_start_int, sort_end), 
                #    '{}<span title="Manuscripts with canonical witnesses in this collection">Manuscripts</span>{}'.format(sort_start_int, sort_end), 
                #    '{}<span title="Owner">Owner</span>{}'.format(sort_start_int, sort_end), 
                #    '{}<span title="Scope">Scope</span>{}'.format(sort_start_int, sort_end), 
                #    '{}<span title="Number of Authoritative Statement authors">Authors</span>{}'.format(sort_start_int, sort_end), 
                #    ]

                ## Add the manuscript to the related objects
                #related_objects.append(collections)

                context['related_objects'] = related_objects

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ColwitDetails/add_to_context")

        # Return the context we have made
        return context

    def process_formset(self, prefix, request, formset):
        return None

    def after_save(self, form, instance):
        return True, ""


class ColwitListView(BasicList):
    """Search and list sermons"""
    
    model = Colwit
    listform = ColwitForm
    has_select2 = True
    use_team_group = True
    paginate_by = PAGINATE_BY_VALUE
    bUseFilter = True
    prefix = "colwit"
    new_button = False      # Don't show the [Add new sermon] button here. It is shown under the Manuscript Details view.
    basketview = False
    plural_name = "Collection witnesses"
    basic_name = "colwit"
    template_help = "seeker/filter_help.html"

    order_cols = ['lilacodefull', 'collection__name', 'codhead__msitem__manu__idno', 'siglist', '','', 'stype']
    order_default = order_cols
    order_heads = [
        {'name': 'Key',         'order': 'o=1', 'type': 'str', 'custom': 'key',         'linkdetails': True}, 
        {'name': 'Collection',  'order': 'o=2', 'type': 'str', 'custom': 'collection',  'linkdetails': True}, 
        {'name': 'Manuscript',  'order': 'o=3', 'type': 'str', 'custom': 'manuscript',  'linkdetails': True, 'main': True},
        {'name': 'Clavis',      'order': 'o=4', 'type': 'str', 'custom': 'clavis'},
        {'name': 'Locus',       'order': '',    'type': 'str', 'custom': 'locus' },
        {'name': 'Links',       'order': '',    'type': 'str', 'custom': 'links'},
        {'name': 'Status',      'order': 'o=7', 'type': 'str', 'custom': 'status'}]

    filters = [ ]
    uploads = []
    
    searches = [
        {'section': '', 'filterlist': [
            ]},
        {'section': 'collection', 'filterlist': [
            ]},
        {'section': 'manuscript', 'filterlist': [
            ]},
        {'section': 'other', 'filterlist': [
            ]}
         ]

    def initializations(self):
        oErr = ErrHandle()
        try:
            # ======== One-time adaptations ==============
            listview_adaptations("colwit_list")

            ## Check if there are any sermons not connected to a manuscript: remove these
            #delete_id = Colwit.objects.filter(Q(msitem__isnull=True)|Q(msitem__manu__isnull=True)).values('id')
            #if len(delete_id) > 0:
            #    oErr.Status("Deleting {} colwits that are not connected".format(len(delete_id)))
            #    Colwit.objects.filter(id__in=delete_id).delete()

            # Make sure to set a basic filter
            # self.basic_filter = Q(mtype="man")
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ColwitListiew/initializations")
        return None

    def add_to_context(self, context, initial):
        oErr = ErrHandle()
        try:
            # Find out who the user is
            profile = Profile.get_user_profile(self.request.user.username)
            context['basketsize'] = 0 if profile == None else profile.basketsize
            context['basket_show'] = reverse('basket_show')
            context['basket_update'] = reverse('basket_update')

            # Does this user have upload permissions?
            if context['is_app_uploader']:
                if len(self.uploads) == 0:
                    # Yes, user has upload permissions
                    html = []
                    html.append("Import Collection witnesses from one or more Excel files.")
                    msg = "<br />".join(html)
                    oExcel = dict(title="collection_witnesses", label="Excel",
                                    url=reverse('colwit_upload_excel'),
                                    type="multiple", msg=msg)
                    self.uploads.append(oExcel)

                context['uploads'] = self.uploads
            else:
                context['uploads'] = []
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ColwitListview/add_to_context")

        return context

    def get_basketqueryset(self):
        if self.basketview:
            profile = Profile.get_user_profile(self.request.user.username)
            qs = profile.basketitems.all()
        else:
            qs = Colwit.objects.all()
        return qs

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []

        if custom == "collection":
            html.append("<span class='collection'>{}</span>".format(instance.collection.name))
        elif custom == "key":
            lilacode = instance.lilacodefull
            html.append( "-" if lilacode is None else lilacode)
        elif custom == "manuscript":
            manu = instance.codhead.msitem.manu
            sHtml = manu.get_lilacode()
            html.append("<span class='manuscript'>{}</span>".format(sHtml))
        elif custom == "clavis":
            qs = instance.signatures.all().order_by('-editype', 'code')
            for obj in qs:
                editype = obj.editype
                full = obj.code
                short = full
                if len(short) > 20:
                    short = "{}...".format(full[:20])
                html.append("<span class='badge signature {}' title='{}'>{}</span>".format(
                    editype, full, short[:20]))
        elif custom == "locus":
            locus = instance.codhead.locus
            html.append(locus)
        elif custom == "links":
            html.append("-")
        elif custom == "status":
            # Provide that status badge
            html.append(instance.get_stype_light())
            
        # Combine the HTML code
        sBack = "\n".join(html)
        return sBack, sTitle

    def adapt_search(self, fields):
        # Adapt the search to the keywords that *may* be shown
        lstExclude=[]
        qAlternative = None
        oErr = ErrHandle()

        try:

            # Double check the length of the exclude list
            if len(lstExclude) == 0:
                lstExclude = None
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ColwitListView/adapt_search")
        
        return fields, lstExclude, qAlternative

    def view_queryset(self, qs):
        """View the queryset"""

        oErr = ErrHandle()
        try:
            search_id = [x['id'] for x in qs.values('id')]
            profile = Profile.get_user_profile(self.request.user.username)
            profile.search_canwit = json.dumps(search_id)
            profile.save()
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ColwitListView/view_queryset")
        return None

    def get_helptext(self, name):
        """Use the get_helptext function defined in models.py"""
        return get_helptext(name)


# ============= CODICOLOGICAL UNIT HEAD VIEWS ============================


class CodheadEdit(BasicDetails):
    """The editable part of one canwit description (manifestation)"""
    
    model = Codhead
    mForm = CodheadForm
    prefix = "chead"
    title = "Section" 
    rtype = "json"
    mainitems = []
    basic_name = "codhead"
    use_team_group = True
    history_button = True
    prefix_type = "simple"

    stype_edi_fields = ['locus', 'title']

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        solemne_action_add(self, instance, details, actiontype)

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        oErr = ErrHandle()
        try:
            # Need to know who this user (profile) is
            profile = Profile.get_user_profile(self.request.user.username)

            # Define the main items to show and edit
            manu_id = None if instance == None else instance.get_manuscript().id
            context['mainitems'] = [
                # {'type': 'plain', 'label': "Status:",               'value': instance.get_stype_light(True),'field_key': 'stype'},
                # -------- HIDDEN field values ---------------
                {'type': 'plain', 'label': "Manuscript id", 'value': manu_id,                   'field_key': "manu",        'empty': 'hide'},
                # --------------------------------------------
                {'type': 'plain', 'label': "Locus:",        'value': instance.locus,            'field_key': "locus"}, 
                {'type': 'plain', 'label': "Title:",        'value': instance.title,            'field_key': 'title'},
                {'type': 'plain', 'label': "Witness",       'value': instance.get_colwit()                      },
                 ]

            # Add a button back to the Manuscript
            topleftlist = []
            if instance.get_manuscript():
                manu = instance.get_manuscript()
                buttonspecs = {'label': "M", 
                     'title': "Go to manuscript {}".format(manu.idno), 
                     'url': reverse('manuscript_details', kwargs={'pk': manu.id})}
                topleftlist.append(buttonspecs)
                lcity = "" if manu.lcity == None else "{}, ".format(manu.lcity.name)
                lib = "" if manu.library == None else "{}, ".format(manu.library.name)
                idno = "{}{}{}".format(lcity, lib, manu.idno)
            else:
                idno = "(unknown)"
            context['topleftbuttons'] = topleftlist

            # Add the manuscript's IDNO completely right
            title_right = ["<span class='manuscript-idno' title='Manuscript'>{}</span>".format(
                idno)]
            #    ... as well as the *title* of the Codico to which I belong
            codico = instance.msitem.codico
            codi_title = "cod. unit. {}".format(codico.order)
            title_right.append("&nbsp;<span class='codico-title' title='Codicologial unit'>{}</span>".format(codi_title))
            context['title_right'] = "".join(title_right)

            # Signal that we have select2
            context['has_select2'] = True
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CodheadEdit/add_to_context")

        # Return the context we have made
        return context

    def after_new(self, form, instance):
        """Action to be performed after adding a new item"""

        ## Set the 'afternew' URL
        manu = instance.get_manuscript()
        #if manu and instance.order < 0:
        #    # Calculate how many sermons there are
        #    sermon_count = manu.get_canwit_count()
        #    # Make sure the new sermon gets changed
        #    form.instance.order = sermon_count

        # Return positively
        return True, "" 

    def after_save(self, form, instance):
        """This is for processing items from the list of available ones"""

        msg = ""
        bResult = True
        oErr = ErrHandle()
        method = "nodistance"   # Alternative: "superdist"
        
        try:
            # Process many-to-many changes: Add and remove relations in accordance with the new set passed on by the user
            if getattr(form, 'cleaned_data') != None:
                pass
        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg

    def before_save(self, form, instance):
        oErr = ErrHandle()
        bBack = True
        msg = ""
        try:
            if hasattr(form, 'cleaned_data'):

                # Issue #421: check how many projects are attached to the manuscript
                if not instance is None and not instance.msitem is None and not instance.msitem.manu is None:
                    # Need to know who is 'talking'...
                    username = self.request.user.username
                    profile = Profile.get_user_profile(username)

                    # There is a sermon and a manuscript
                    manu = instance.msitem.manu

        except:
            msg = oErr.get_error_message()
            oErr.DoError("CodheadEdit/before_save")
            bBack = False
        return bBack, msg

    def get_colwit(self, instance):
        """Get the code to possibly create a collection witness"""

        oErr = ErrHandle()
        sBack = "-"
        template = "seeker/create_colwit.html"
        try:
            # Create a form that allows a user to select one collection
            colform = ColForm(prefix="colw", username=self.request.user.username, team_group=app_editor, userplus=app_userplus)
            # Create a context
            colwit = Colwit.objects.filter(codhead=instance).first()
            context = dict(colwit=colwit, codhead=instance,colform=colform)
            # Create a template with this information
            sBack = render_to_string(template, context, self.request)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_colwit")
        return sBack

    def get_history(self, instance):
        return solemne_get_history(instance)


class CodheadDetails(CodheadEdit):
    """The details of one manuscript section (Codhead)"""

    rtype = "html"

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Start by executing the standard handling
        context = super(CodheadDetails, self).add_to_context(context, instance)

        oErr = ErrHandle()

        try:
            # Load the after_details information

            context['after_details'] = self.get_colwit(instance)

            context['sections'] = []

            # List of post-load objects
            context['postload_objects'] = []

            # Lists of related objects
            context['related_objects'] = []
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CodheadDetails/add_to_context")

        # Return the context we have made
        return context

    def before_save(self, form, instance):
        bResult = True
        msg = ""
        oErr = ErrHandle()
        try:
            # If needed, create an MsItem
            if instance.msitem == None:
                # Make sure we have the manuscript
                manu = form.cleaned_data.get("manu", None)
                msitem = MsItem.objects.create(manu=manu)
                # Now make sure to set the link from Manuscript to MsItem
                instance.msitem = msitem

            # Double check for the presence of manu and order
            if instance.msitem and instance.msitem.order < 0:
                # Calculate how many MSITEMS (!) there are
                msitem_count = instance.msitem.manu.manuitems.all().count()
                # Adapt the MsItem order
                msitem.order = msitem_count
                msitem.save()
                # Find out which is the one PRECEDING me (if any) at the HIGHEST level
                prec_list = instance.msitem.manu.manuitems.filter(parent__isnull=True, order__gt=msitem.order)
                if prec_list.count() > 0:
                    # Get the last item
                    prec_item = prec_list.last()
                    # Set the 'Next' here correctly
                    prec_item.next = msitem
                    prec_item.save()
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CodheadDetails/before_save")
            bResult = False

        return bResult, msg

    def process_formset(self, prefix, request, formset):
        return None

    def after_save(self, form, instance):
        return True, ""


class CodheadListView(BasicList):
    """Search and list sermons"""
    
    model = Codhead
    listform = CodheadForm
    has_select2 = True
    use_team_group = True
    paginate_by = PAGINATE_BY_VALUE
    bUseFilter = True
    prefix = "chead"
    new_button = False      # Don't show the [Add new sermon] button here. It is shown under the Manuscript Details view.
    basketview = False
    plural_name = "Manuscript sections"
    basic_name = "codhead"
    template_help = "seeker/filter_help.html"

    order_cols = ['msitem__manu__idno', 'title', 'locus']
    order_default = order_cols
    order_heads = [
        {'name': 'Manuscript',  'order': 'o=4', 'type': 'str', 'custom': 'manuscript'},
        {'name': 'Title',       'order': 'o=5', 'type': 'str', 'custom': 'title', 
         'allowwrap': True,           'autohide': "on", 'filter': 'filter_title'},
        {'name': 'Locus',       'order': '',    'type': 'str', 'field':  'locus' }
         ]

    filters = [ {"name": "Manuscript",       "id": "filter_manuid",         "enabled": False},
                {"name": "Title",            "id": "filter_title",          "enabled": False},
                ]
    
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'title',     'dbfield': 'title',         'keyS': 'srch_title'},
            {'filter': 'manuid',    'fkfield': 'msitem__manu',  'keyS': 'manuidno',     'keyList': 'manuidlist', 'keyFk': 'idno', 'infield': 'id'},
            ]},
        {'section': 'other', 'filterlist': [
            {'filter': 'mtype',     'dbfield': 'msitem__mtype',    'keyS': 'mtype'},
            ]}
         ]

    def initializations(self):
        oErr = ErrHandle()
        try:
            # ======== One-time adaptations ==============
            listview_adaptations("canwit_list")

            # Make sure to set a basic filter
            self.basic_filter = Q(mtype="man")
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CodheadListView/initializations")
        return None

    def add_to_context(self, context, initial):
        # Find out who the user is
        profile = Profile.get_user_profile(self.request.user.username)
        return context

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []
        if custom == "manuscript":
            manu = instance.get_manuscript()
            if manu == None:
                html.append("-")
            else:
                if manu.idno == None:
                    sIdNo = "-"
                else:
                    sIdNo = manu.idno[:20]
                html.append("<a href='{}' class='nostyle'><span style='font-size: small;'>{}</span></a>".format(
                    reverse('manuscript_details', kwargs={'pk': manu.id}),
                    sIdNo))
                sTitle = manu.idno
        elif custom == "title":
            sTitle = ""
            if instance.title != None and instance.title != "":
                sTitle = instance.title
            html.append(sTitle)
            
        # Combine the HTML code
        sBack = "\n".join(html)
        return sBack, sTitle

    def adapt_search(self, fields):
        # Adapt the search to the keywords that *may* be shown
        lstExclude=[]
        qAlternative = None
        oErr = ErrHandle()

        try:
            # Make sure we show MANUSCRIPTS (identifiers) as well as reconstructions
            lstExclude.append(Q(mtype='tem') )

            # Double check the length of the exclude list
            if len(lstExclude) == 0:
                lstExclude = None
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CodheadListView/adapt_search")
        
        return fields, lstExclude, qAlternative

    def view_queryset(self, qs):
        search_id = [x['id'] for x in qs.values('id')]
        profile = Profile.get_user_profile(self.request.user.username)
        profile.search_canwit = json.dumps(search_id)
        profile.save()
        return None

    def get_helptext(self, name):
        """Use the get_helptext function defined in models.py"""
        return get_helptext(name)


# ============= CANONICAL WITNESS VIEWS ============================


class CanwitEdit(BasicDetails):
    """The editable part of one canwit description (manifestation)"""
    
    model = Canwit
    mForm = CanwitForm
    prefix = "canwi"
    title = "Canon Witness" 
    rtype = "json"
    mainitems = []
    basic_name = "canwit"
    use_team_group = True
    history_button = True
    prefix_type = "simple"

    AustatmFormSet = inlineformset_factory(Canwit, CanwitAustat,
                                         form=CanwitSuperForm, min_num=0,
                                         fk_name = "canwit",
                                         extra=0, can_delete=True, can_order=False)
    AustatfFormSet = inlineformset_factory(Canwit, CanwitAustat,
                                         form=CanwitSuperForm, min_num=0,
                                         fk_name = "canwit",
                                         extra=0, can_delete=True, can_order=False)
    SDkwFormSet = inlineformset_factory(Canwit, CanwitKeyword,
                                       form=CanwitKeywordForm, min_num=0,
                                       fk_name="canwit", extra=0)
    SDcolFormSet = inlineformset_factory(Canwit, CollectionCanwit,
                                       form=CanwitCollectionForm, min_num=0,
                                       fk_name="canwit", extra=0)
    SDsignFormSet = inlineformset_factory(Canwit, CanwitSignature,
                                         form=CanwitSignatureForm, min_num=0,
                                         fk_name = "canwit",
                                         extra=0, can_delete=True, can_order=False)
    SbrefFormSet = inlineformset_factory(Canwit, BibRange,
                                         form=BibRangeForm, min_num=0,
                                         fk_name = "canwit",
                                         extra=0, can_delete=True, can_order=False)

    formset_objects = [{'formsetClass': AustatmFormSet, 'prefix': 'cwaum', 'readonly': False, 'noinit': True, 'linkfield': 'canwit'},
                       {'formsetClass': AustatfFormSet, 'prefix': 'cwauf', 'readonly': False, 'noinit': True, 'linkfield': 'canwit'},
                       {'formsetClass': SDkwFormSet,   'prefix': 'sdkw',   'readonly': False, 'noinit': True, 'linkfield': 'canwit'},                       
                       {'formsetClass': SDcolFormSet,  'prefix': 'sdcol',  'readonly': False, 'noinit': True, 'linkfield': 'canwit'},
                       {'formsetClass': SDsignFormSet, 'prefix': 'sdsig',  'readonly': False, 'noinit': True, 'linkfield': 'canwit'},
                       {'formsetClass': SbrefFormSet,  'prefix': 'sbref',  'readonly': False, 'noinit': True, 'linkfield': 'canwit'}] 

    stype_edi_fields = ['locus', 'author', 'sectiontitle', 'title', 'subtitle', 'ftext', 'ftrans', 'postscriptum', 'quote', 
                        'bibnotes', 'feast', 'bibleref', 'additional', 'note',  # 'manu', 
                        'CanwitSignature', 'siglist',
                        'CanwitAustat', 'superlist']

    def custom_init(self, instance):
        method = "nodistance"   # Alternative: "superdist"

        if instance:
            istemplate = (instance.mtype == "tem")
            if istemplate:
                # Need a smaller array of formset objects
                self.formset_objects = [{'formsetClass': self.AustatmFormSet, 'prefix': 'cwaum', 'readonly': False, 'noinit': True, 'linkfield': 'canwit'}]

            # Indicate where to go to after deleting
            if instance != None and instance.msitem != None and instance.msitem.manu != None:
                self.afterdelurl = reverse('manuscript_details', kwargs={'pk': instance.msitem.manu.id})

            # Then check if all distances have been calculated in AustatDist
            if method == "superdist":
                qs = AustatDist.objects.filter(sermon=instance)
                if qs.count() == 0:
                    # These distances need calculation...
                    instance.do_distance()
        return None

    def get_form_kwargs(self, prefix):
        # Determine the method
        method = "nodistance"   # Alternative: "superdist"

        oBack = None
        if prefix in ['cwaum', 'cwauf'] and method == "superdist":
            if self.object != None:
                # Make sure that the sermon is known
                oBack = dict(sermon_id=self.object.id)
        return oBack
           
    def add_to_context(self, context, instance):
        """Add to the existing context"""

        oErr = ErrHandle()
        try:
            # Need to know who this user (profile) is
            profile = Profile.get_user_profile(self.request.user.username)
            may_edit = (context['is_app_editor'])

            istemplate = (instance.mtype == "tem")

            # Define the main items to show and edit
            # manu_id = None if instance == None or instance.manu == None else instance.manu.id
            manu_id = None if instance == None else instance.get_manuscript().id
            context['mainitems'] = []
            # Possibly add the Template identifier
            if istemplate:
                context['mainitems'].append(
                    {'type': 'plain', 'label': "Template:", 'value': instance.get_template_link(profile)}
                    )
            # Get the main items
            mainitems_main = [
                {'type': 'plain', 'label': "Status:",               'value': instance.get_stype_light(may_edit),'field_key': 'stype'},
                # -------- HIDDEN field values ---------------
                {'type': 'plain', 'label': "Manuscript id",         'value': manu_id,                   'field_key': "manu",        'empty': 'hide'},
                # --------------------------------------------
                {'type': 'plain', 'label': "LiLaC code:",           'value': instance.get_lilacode(),   'field_key': "lilacode"}, 
                {'type': 'plain', 'label': "Caput:",                'value': instance.get_caput(),      'field_key': "caput"}, 
                {'type': 'plain', 'label': "Locus:",                'value': instance.locus,            'field_key': "locus"}, 
                {'type': 'safe',  'label': "Attributed author:",    'value': instance.get_author(),     'field_key': 'author'},
                {'type': 'plain', 'label': "Author certainty:",     'value': instance.get_autype(),     'field_key': 'autype', 'editonly': True},
                {'type': 'plain', 'label': "Section title:",        'value': instance.sectiontitle,     'field_key': 'sectiontitle'},

                # issue #32 
                # {'type': 'safe',  'label': "Lectio:",               'value': instance.get_quote_markdown(),'field_key': 'quote'}, 

                # Issue #237, delete subtitle
                {'type': 'plain', 'label': "Sub title:",            'value': instance.subtitle,         'field_key': 'subtitle', 
                 'editonly': True, 'title': 'The subtitle field is legacy. It is edit-only, non-viewable'},
                {'type': 'safe',  'label': "Full text:",            'value': instance.get_ftext_markdown(), 
                 'field_key': 'ftext',  'key_ta': 'srmincipit-key'}, 
                {'type': 'safe',  'label': "Translation:",          'value': instance.get_ftrans_markdown(),
                 'field_key': 'ftrans', 'key_ta': 'srmexplicit-key'}, 
                #{'type': 'safe',  'label': "Postscriptum:",         'value': instance.get_postscriptum_markdown(),
                # 'field_key': 'postscriptum'}, 
                # Issue #23: delete bibliographic notes
                {'type': 'plain', 'label': "Bibliographic notes:",  'value': instance.bibnotes,         'field_key': 'bibnotes', 
                 'editonly': True, 'title': 'The bibliographic-notes field is legacy. It is edit-only, non-viewable'},
                #{'type': 'plain', 'label': "Feast:",                'value': instance.get_feast(),      'field_key': 'feast'}
                 ]
            exclude_field_keys = ['locus']
            for item in mainitems_main: 
                # Make sure to exclude field key 'locus'
                if not istemplate or item['field_key'] not in exclude_field_keys:
                    context['mainitems'].append(item)

            # Bibref and Cod. notes can only be added to non-templates
            if not istemplate:
                #mainitems_BibRef ={'type': 'plain', 'label': "Bible reference(s):",   'value': instance.get_bibleref(),        
                # 'multiple': True, 'field_list': 'bibreflist', 'fso': self.formset_objects[5]}
                #context['mainitems'].append(mainitems_BibRef)
                mainitems_CodNotes ={'type': 'plain', 'label': "Cod. notes:",           'value': instance.additional,       
                 'field_key': 'additional',   'title': 'Codicological notes'}
                context['mainitems'].append(mainitems_CodNotes)

            mainitems_more =[
                {'type': 'plain', 'label': "Note:",        'value': instance.get_note_markdown(),             'field_key': 'note'},
                {'type': 'line',  'label': "Signatures:",  'value': instance.get_colwit_signatures(),
                 'title': 'These are signatures connected via the Collection Witness'},
                {'type': 'line',  'label': "Collection witness:", 'value': instance.get_colwit(),
                 'title': "Optional collection witness that I am under"},
                ]
            for item in mainitems_more: context['mainitems'].append(item)

            if not istemplate:
                username = profile.user.username
                team_group = app_editor
                mainitems_m2m = [
                    {'type': 'line',  'label': "Keywords:",             'value': instance.get_keywords_markdown(), 
                     # 'multiple': True,  'field_list': 'kwlist',         'fso': self.formset_objects[1]},
                     'field_list': 'kwlist',         'fso': self.formset_objects[2]},
                    {'type': 'plain', 'label': "Keywords (user):", 'value': instance.get_keywords_user_markdown(profile),   'field_list': 'ukwlist',
                     'title': 'User-specific keywords. If the moderator accepts these, they move to regular keywords.'},
                    {'type': 'line',  'label': "Keywords (related):",   'value': instance.get_keywords_ssg_markdown(),
                     'title': 'Keywords attached to the Authoritative statement(s)'},
                    #{'type': 'line',    'label': "Gryson/Clavis:",'value': instance.get_eqsetsignatures_markdown('combi'),
                    # 'title': "Gryson/Clavis codes of the Sermons Gold that are part of the same equality set + those manually linked to this manifestation Sermon"}, 
                    #{'type': 'line',    'label': "Gryson/Clavis (manual):",'value': instance.get_sermonsignatures_markdown(),
                    # 'title': "Gryson/Clavis codes manually linked to this manifestation Sermon", 'unique': True, 'editonly': True, 
                    # 'multiple': True,
                    # 'field_list': 'siglist_m', 'fso': self.formset_objects[4], 'template_selection': 'ru.solemne.sigs_template'},
                    {'type': 'plain',   'label': "Personal datasets:",  'value': instance.get_collections_markdown(username, team_group, settype="pd"), 
                     'multiple': True,  'field_list': 'collist_s',      'fso': self.formset_objects[3] },
                    {'type': 'plain',   'label': "Public datasets (link):",  'value': instance.get_collection_link("pd"), 
                     'title': "Public datasets in which an Authoritative statement is that is linked to this sermon"},
                    {'type': 'plain',   'label': "Historical collections (link):",  'value': instance.get_collection_link("hc"), 
                     'title': "Historical collections in which an Authoritative statement is that is linked to this sermon"},
                    {'type': 'line',    'label': "Editions:",           'value': instance.get_editions_markdown(),
                     'title': "Editions of the Sermons Gold that are part of the same equality set"},
                    {'type': 'line',    'label': "Literature:",         'value': instance.get_litrefs_markdown()},
                    # Project HIER
                    {'type': 'plain', 'label': "Project:",     'value': instance.get_project_markdown2(), 'field_list': 'projlist'},
                    ]
                for item in mainitems_m2m: context['mainitems'].append(item)

            # IN all cases: Fons materialis = Austat
            mainitems_AustatM = {'type': 'line',    'label': "Fons materialis:",  'value': self.get_austatlinks_markdown(instance, "mat"), 
                 'title': 'Authoritative statement links (fons materialis)',
                 'multiple': True,  'field_list': 'superlist',       'fso': self.formset_objects[0], 
                 'inline_selection': 'ru.solemne.ssglink_template',   'template_selection': 'ru.solemne.ssgdist_template'}
            context['mainitems'].append(mainitems_AustatM)

            # IN all cases: Fons formalis = 
            mainitems_AustatF = {'type': 'line',    'label': "Fons formalis:",  'value': self.get_austatlinks_markdown(instance, "for"), 
                 'title': 'Authoritative statement links (fons formalis)',
                 'multiple': True,  'field_list': 'formalislist',       'fso': self.formset_objects[1], 
                 'inline_selection': 'ru.solemne.ssglink_template',   'template_selection': 'ru.solemne.ssgdist_template'}
            context['mainitems'].append(mainitems_AustatF)
            # Notes:
            # Collections: provide a link to the Sermon-listview, filtering on those Sermons that are part of one particular collection

            # Add a button back to the Manuscript
            topleftlist = []
            if instance.get_manuscript():
                manu = instance.get_manuscript()
                buttonspecs = {'label': "M", 
                     'title': "Go to manuscript {}".format(manu.idno), 
                     'url': reverse('manuscript_details', kwargs={'pk': manu.id})}
                topleftlist.append(buttonspecs)
                lcity = "" if manu.lcity == None else "{}, ".format(manu.lcity.name)
                lib = "" if manu.library == None else "{}, ".format(manu.library.name)
                idno = "{}{}{}".format(lcity, lib, manu.idno)
            else:
                idno = "(unknown)"
            context['topleftbuttons'] = topleftlist
            # Add something right to the CanwitDetails title
            context['title_addition'] = instance.get_breadcrumb() # instance.get_austat_lilacode_markdown()

            # Add the manuscript's IDNO completely right
            title_right = ["<span class='manuscript-idno' title='Manuscript'>{}</span>".format(
                idno)]
            #    ... as well as the *title* of the Codico to which I belong
            codico = instance.msitem.codico

            # Also add the codicological unit to the right
            codi_title = "cod. unit. {}".format(codico.order)
            title_right.append("&nbsp;<span class='codico-title' title='Codicologial unit'>{}</span>".format(codi_title))
            context['title_right'] = "".join(title_right)

            # Signal that we have select2
            context['has_select2'] = True

            # Add comment modal stuff
            initial = dict(otype="sermo", objid=instance.id, profile=profile)
            context['commentForm'] = CommentForm(initial=initial, prefix="com")
            context['comment_list'] = get_usercomments('sermo', instance, profile)
            lhtml = []
            lhtml.append(render_to_string("seeker/comment_add.html", context, self.request))
            context['after_details'] = "\n".join(lhtml)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CanwitEdit/add_to_context")

        # Return the context we have made
        return context

    def get_austatlinks_markdown(self, instance, fonstype):
        context = {}
        template_name = 'seeker/canwit_austatlinks.html'
        sBack = ""
        if instance:
            # Add to context
            context['superlist'] = instance.canwit_austat.filter(fonstype=fonstype).order_by('canwit__author__name', 'canwit__siglist')
            context['is_app_editor'] = user_is_ingroup(self.request, app_editor)
            context['object_id'] = instance.id
            # Calculate with the template
            sBack = render_to_string(template_name, context)
        return sBack

    def after_new(self, form, instance):
        """Action to be performed after adding a new item"""

        ## Set the 'afternew' URL
        manu = instance.get_manuscript()
        if manu and instance.order < 0:
            # Calculate how many sermons there are
            sermon_count = manu.get_canwit_count()
            # Make sure the new sermon gets changed
            form.instance.order = sermon_count
        # Check project assignment
        if instance.projects.count() == 0:
            # Need to assign to default project
            projects = self.request.user.user_profiles.first().get_defaults()
            for project in projects:
                obj = CanwitProject.objects.filter(project=project, canwit=instance).first()
                if obj is None:
                    obj = CanwitProject.objects.create(project=project, canwit=instance)


        # Return positively
        return True, "" 

    def process_formset(self, prefix, request, formset):
        """This is for processing *NEWLY* added items (using the '+' sign)"""

        bAllowNewSignatureManually = True   # False
        errors = []
        bResult = True
        oErr = ErrHandle()
        # Determine the method
        method = "nodistance"   # Alternative: "superdist"
        try:
            instance = formset.instance
            for form in formset:
                if form.is_valid():
                    cleaned = form.cleaned_data
                    # Action depends on prefix
                    if prefix == "sdsig" and bAllowNewSignatureManually:
                        # Signature processing
                        # NOTE: this should never be reached, because we do not allow adding *new* signatures manually here
                        editype = ""
                        code = ""
                        if 'newgr' in cleaned and cleaned['newgr'] != "":
                            # Add gryson
                            editype = "gr"
                            code = cleaned['newgr']
                        elif 'newcl' in cleaned and cleaned['newcl'] != "":
                            # Add gryson
                            editype = "cl"
                            code = cleaned['newcl']
                        elif 'newot' in cleaned and cleaned['newot'] != "":
                            # Add gryson
                            editype = "ot"
                            code = cleaned['newot']
                        if editype != "":
                            # Set the correct parameters
                            form.instance.code = code
                            form.instance.editype = editype
                            # Note: it will get saved with formset.save()
                    elif prefix == "sdkw":
                        # Keyword processing
                        if 'newkw' in cleaned and cleaned['newkw'] != "":
                            newkw = cleaned['newkw']
                            # Is the KW already existing?
                            obj = Keyword.objects.filter(name=newkw).first()
                            if obj == None:
                                obj = Keyword.objects.create(name=newkw)
                            # Make sure we set the keyword
                            form.instance.keyword = obj
                            # Note: it will get saved with formset.save()
                    elif prefix == "sdcol":
                        # Collection processing
                        if 'newcol' in cleaned and cleaned['newcol'] != "":
                            newcol = cleaned['newcol']
                            profile = Profile.get_user_profile(request.user.username)
                            # Is the COL already existing?
                            obj = Collection.objects.filter(name=newcol).first()
                            if obj == None:
                                # TODO: add profile here
                                obj = Collection.objects.create(name=newcol, type='sermo', owner=profile)
                            # once a collection has been created, make sure it gets assigned to a project
                            if not profile is None and obj.projects.count() == 0:
                                # Assign the default projects
                                projects = profile.get_defaults()
                                obj.set_projects(projects)
                            # Make sure we set the keyword
                            form.instance.collection = obj
                            # Note: it will get saved with formset.save()
                    elif prefix in ["cwaum", "cwauf"]:
                        # Determine the fonstype based on prefix
                        fonstype = "mat" if prefix == "cwaum" else "for"
                        # Canwit-To-Austat processing
                        if method == "superdist":
                            # Note: nov/2 went over from 'newsuper' to 'newsuperdist'
                            if 'newsuperdist' in cleaned and cleaned['newsuperdist'] != "":
                                newsuperdist = cleaned['newsuperdist']
                                # Take the default linktype
                                linktype = "uns"

                                # Convert from newsuperdist to actual super (SSG)
                                superdist = AustatDist.objects.filter(id=newsuperdist).first()
                                if superdist != None:
                                    austat = superdist.austat

                                    # Check existence of link between S-SSG
                                    obj = CanwitAustat.objects.filter(sermon=instance, austat=austat, linktype=linktype, fonstype=fonstype).first()
                                    if obj == None:
                                        # Set the right parameters for creation later on
                                        form.instance.linktype = linktype
                                        form.instance.fonstype = fonstype
                                        form.instance.austat = austat
                        elif method == "nodistance":
                            newsuper = cleaned.get("newsuper", "")
                            newnote = cleaned.get("newnote")
                            newcreate = cleaned.get("newcreate", "")
                            if newsuper != "":
                                # Take the default linktype
                                linktype = "uns"

                                # Check existence
                                obj = CanwitAustat.objects.filter(canwit=instance, austat=newsuper, linktype=linktype, fonstype=fonstype).first()
                                if obj == None:
                                    obj_austat = Austat.objects.filter(id=newsuper).first()
                                    if obj_austat != None:
                                        # Set the right parameters for creation later on
                                        form.instance.linktype = linktype
                                        form.instance.fonstype = fonstype
                                        form.instance.austat = obj_austat
                                        if not newnote is None:
                                            form.instance.note = newnote
                            elif newcreate != "":
                                # Take the default linktype
                                linktype = "uns"

                                # We are being asked to create a new austat with lilacode [newcreate]
                                # Make sure the austat does not exist yet
                                obj_austat = Austat.objects.filter(keycodefull__iexact=newcreate).first()
                                if obj_austat is None:
                                    # Indeed, doesn't exist yet: create it
                                    parts = newcreate.rsplit(".", 1)
                                    if len(parts) == 2:
                                        arNumber = re.findall(r'\d+', parts[1])
                                        if len(parts[0]) > 0 and len(arNumber) > 0:
                                            work = parts[0]
                                            keycode = arNumber[0]
                                            # Get the AuWork
                                            auwork = Auwork.objects.filter(key__iexact=work).first()
                                            if auwork is None:
                                                # Must create this work
                                                auwork = Auwork.objects.create(key=work)
                                            # Create the Austat
                                            obj_austat = Austat.objects.create(
                                                keycode=keycode, auwork=auwork, keycodefull=newcreate,
                                                stype="imp", atype="acc")



                                # Double check
                                if not obj_austat is None:
                                    # well, it is existing, so add a link to this one
                                        form.instance.linktype = linktype
                                        form.instance.fonstype = fonstype
                                        form.instance.austat = obj_austat
                                        if not newnote is None:
                                            form.instance.note = newnote

                        # Note: it will get saved with form.save()
                    elif prefix == "sbref":
                        # Processing one BibRange
                        newintro = cleaned.get('newintro', None)
                        onebook = cleaned.get('onebook', None)
                        newchvs = cleaned.get('newchvs', None)
                        newadded = cleaned.get('newadded', None)

                        # Minimal need is BOOK
                        if onebook != None:
                            # Note: normally it will get saved with formset.save()
                            #       However, 'noinit=False' formsets must arrange their own saving

                            #bNeedSaving = False

                            ## Double check if this one already exists for the current instance
                            #obj = instance.canwitbibranges.filter(book=onebook, chvslist=newchvs, intro=newintro, added=newadded).first()
                            #if obj == None:
                            #    obj = BibRange.objects.create(sermon=instance, book=onebook, chvslist=newchvs)
                            #    bNeedSaving = True
                            #if newintro != None and newintro != "": 
                            #    obj.intro = newintro
                            #    bNeedSaving = True
                            #if newadded != None and newadded != "": 
                            #    obj.added = newadded
                            #    bNeedSaving = True
                            #if bNeedSaving:
                            #    obj.save()
                            #    x = instance.canwitbibranges.all()

                            form.instance.book = onebook
                            if newchvs != None:
                                form.instance.chvslist = newchvs
                            form.instance.intro = newintro
                            form.instance.added = newadded
                            

                else:
                    errors.append(form.errors)
                    bResult = False
        except:
            msg = oErr.get_error_message()
            iStop = 1
        return None

    def before_save(self, form, instance):
        oErr = ErrHandle()
        bBack = True
        msg = ""
        try:
            if hasattr(form, 'cleaned_data'):
                # Make sure the author type is processed correctly
                if 'autype' in form.cleaned_data and form.cleaned_data['autype'] != "":
                    autype = form.cleaned_data['autype']
                    form.instance.autype = autype

                # Issue #421: check how many projects are attached to the manuscript
                if not instance is None and not instance.msitem is None and not instance.msitem.manu is None:
                    # Need to know who is 'talking'...
                    username = self.request.user.username
                    profile = Profile.get_user_profile(username)

                    # Always get the project list
                    projlist = form.cleaned_data.get("projlist")

                    # There is a sermon and a manuscript
                    manu = instance.msitem.manu
                    # How many projects are attached to this manuscript
                    manu_project_count = manu.projects.count()
                    if manu_project_count > 1:
                        # There are multiple projects attached to the manuscript
                        # This means that the user *must* have specified one project

                        bBack, msg = evaluate_projlist(profile, instance, projlist, "Sermon manifestation")

                        #if len(projlist) == 0:
                        #    # Add a warning that the user must manually provide a project
                        #    msg = "Add a project: A sermon must belong to at least one project"
                        #    bBack = False
                    else:
                        # It would seem that this kind of check is needed anyway...
                        bBack, msg = evaluate_projlist(profile, instance, projlist, "Sermon manifestation")

        except:
            msg = oErr.get_error_message()
            oErr.DoError("SermonEdit/before_save")
            bBack = False
        return bBack, msg

    def after_save(self, form, instance):
        """This is for processing items from the list of available ones"""

        msg = ""
        bResult = True
        oErr = ErrHandle()
        method = "nodistance"   # Alternative: "superdist"
        
        try:
            # Process many-to-many changes: Add and remove relations in accordance with the new set passed on by the user
            if getattr(form, 'cleaned_data') != None:
                # (1) 'keywords'
                kwlist = form.cleaned_data['kwlist']
                adapt_m2m(CanwitKeyword, instance, "canwit", kwlist, "keyword")
            
                # (2) user-specific 'keywords'
                ukwlist = form.cleaned_data['ukwlist']
                profile = Profile.get_user_profile(self.request.user.username)
                adapt_m2m(UserKeyword, instance, "canwit", ukwlist, "keyword", qfilter = {'profile': profile}, extrargs = {'profile': profile, 'type': 'sermo'})

                # (3) 'Links to Austat': fons materialis + fons formalis
                superlist = form.cleaned_data['superlist']
                formalislist = form.cleaned_data['formalislist']
                fonslist = []
                if superlist.count() > 0:
                    for item in superlist.values('id'):
                        fonslist.append(item['id'])
                if formalislist.count() > 0:
                    for item in formalislist.values('id'):
                        fonslist.append(item['id'])
                superlist = CanwitAustat.objects.filter(id__in=fonslist)
                adapt_m2m(CanwitAustat, instance, "canwit", superlist, "austat", extra = ['linktype', 'fonstype'], related_is_through=True)

                # (5) 'collections'
                collist_s = form.cleaned_data['collist_s']
                adapt_m2m(CollectionCanwit, instance, "canwit", collist_s, "collection")

                # (6) 'projects'
                projlist = form.cleaned_data['projlist']
                sermo_proj_deleted = []
                adapt_m2m(CanwitProject, instance, "canwit", projlist, "project", deleted=sermo_proj_deleted)
                project_dependant_delete(self.request, sermo_proj_deleted)

                # When sermons have been added to the manuscript, the sermons need to be updated 
                # with the existing project names 
                # Issue #412: do *NOT* do automatic adjustment to other sermons or to manuscript
                # instance.adapt_projects() # Gaat direct naar adapt_projects in SermDescr

                # Issue #412: when a sermon doesn't yet have a project, it gets the project of the manuscript
                if instance.projects.count() == 0:
                    manu = instance.msitem.manu
                    # How many projects are attached to this manuscript
                    manu_project_count = manu.projects.count()
                    if manu_project_count == 1:
                        project = manu.projects.first()
                        CanwitProject.objects.create(canwit=instance, project=project)

                # Process many-to-ONE changes
                # (1) links from bibrange to sermon
                bibreflist = form.cleaned_data['bibreflist']
                adapt_m2o(BibRange, instance, "canwit", bibreflist)

                # (2) 'canwitsignatures'
                siglist_m = form.cleaned_data['siglist_m']
                adapt_m2o(CanwitSignature, instance, "canwit", siglist_m)

            ## Make sure the 'verses' field is adapted, if needed
            #bResult, msg = instance.adapt_verses()

            # Check if instances need re-calculation
            if method == "superdist":
                if 'incipit' in form.changed_data or 'explicit' in form.changed_data:
                    instance.do_distance(True)

        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        solemne_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return solemne_get_history(instance)


class CanwitDetails(CanwitEdit):
    """The details of one sermon manifestation (Canwit)"""

    rtype = "html"

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Start by executing the standard handling
        context = super(CanwitDetails, self).add_to_context(context, instance)

        oErr = ErrHandle()

        try:
            # Are we copying information?? (only allowed if we are the app_editor)
            if 'supercopy' in self.qd and context['is_app_editor']:
                # Get the ID of the SSG from which information is to be copied to the S
                superid = self.qd['supercopy']
                # Get the SSG instance
                equal = Austat.objects.filter(id=superid).first()

                if equal != None:
                    # Copy all relevant information to the Austat obj (which is a SSG)
                    obj = self.object
                    # (1) copy author
                    if equal.author != None: obj.author = equal.author
                    # (2) copy incipit
                    if equal.incipit != None and equal.incipit != "": obj.incipit = equal.incipit ; obj.srchftext = equal.srchftext
                    # (3) copy explicit
                    if equal.explicit != None and equal.explicit != "": obj.explicit = equal.explicit ; obj.srchftrans = equal.srchftrans

                    # Now save the adapted Austat obj
                    obj.save()

                    # Mark these changes, which are done outside the normal 'form' system
                    actiontype = "save"
                    changes = dict(author=obj.author.id, incipit=obj.incipit, explicit=obj.explicit)
                    details = dict(savetype="change", id=obj.id, changes=changes)
                    solemne_action_add(self, obj, details, actiontype)

                # And in all cases: make sure we redirect to the 'clean' GET page
                self.redirectpage = reverse('canwit_details', kwargs={'pk': self.object.id})
            else:
                context['sections'] = []

                # List of post-load objects
                context['postload_objects'] = []

                # Lists of related objects
                context['related_objects'] = []
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CanwitDetails/add_to_context")

        # Return the context we have made
        return context

    def before_save(self, form, instance):
        bResult = True
        msg = ""
        oErr = ErrHandle()
        try:
            # If needed, create an MsItem
            if instance.msitem == None:
                # Make sure we have the manuscript
                manu = form.cleaned_data.get("manu", None)
                msitem = MsItem.objects.create(manu=manu)
                # Now make sure to set the link from Manuscript to MsItem
                instance.msitem = msitem

                # If possible, also get the mtype
                mtype = self.qd.get("sermo-mtype", None)
                if mtype != None:
                    instance.mtype = mtype

            # Double check for the presence of manu and order
            if instance.msitem and instance.msitem.order < 0:
                # Calculate how many MSITEMS (!) there are
                msitem_count = instance.msitem.manu.manuitems.all().count()
                # Adapt the MsItem order
                msitem.order = msitem_count
                msitem.save()
                # Find out which is the one PRECEDING me (if any) at the HIGHEST level
                prec_list = instance.msitem.manu.manuitems.filter(parent__isnull=True, order__gt=msitem.order)
                if prec_list.count() > 0:
                    # Get the last item
                    prec_item = prec_list.last()
                    # Set the 'Next' here correctly
                    prec_item.next = msitem
                    prec_item.save()
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CanwitDetails/before_save")
            bResult = False

        return bResult, msg

    def process_formset(self, prefix, request, formset):
        return None

    def after_save(self, form, instance):
        return True, ""


class CanwitListView(BasicList):
    """Search and list sermons"""
    
    model = Canwit
    listform = CanwitForm
    has_select2 = True
    use_team_group = True
    paginate_by = PAGINATE_BY_VALUE
    bUseFilter = True
    prefix = "canwi"
    new_button = False      # Don't show the [Add new canwit] button here. It is shown under the Manuscript Details view.
    basketview = False
    plural_name = "Canon witnesses"
    basic_name = "canwit"
    template_help = "seeker/filter_help.html"

    order_cols = ['lilacodefull', 'author__name', 'srchftext', 'srchftrans', 
                  'msitem__manu__idno', '', 'stype']
    order_default = order_cols
    order_heads = [
        {'name': 'LiLaC',       'order': 'o=1', 'type': 'str', 'custom': 'lilacode', 'linkdetails': True}, 
        {'name': 'Author',      'order': 'o=2', 'type': 'str', 'custom': 'author', 'linkdetails': True}, 
        {'name': 'Full text',   'order': 'o=3', 'type': 'str', 'custom': 'ftext',  'linkdetails': True, 'main': True},
        {'name': 'Translation', 'order': 'o=4', 'type': 'str', 'custom': 'ftrans', 'linkdetails': True},
        {'name': 'Manuscript',  'order': 'o=5', 'type': 'str', 'custom': 'manuscript'},
        #{'name': 'Title',       'order': 'o=6', 'type': 'str', 'custom': 'title', 
        # 'allowwrap': True,           'autohide': "on", 'filter': 'filter_title'},
        #{'name': 'Section',     'order': 'o=7', 'type': 'str', 'custom': 'sectiontitle', 
        # 'allowwrap': True,    'autohide': "on", 'filter': 'filter_sectiontitle'},
        {'name': 'Locus',       'order': '',    'type': 'str', 'field':  'locus' },
        {'name': 'Status',      'order': 'o=9', 'type': 'str', 'custom': 'status'}]

    filters = [ {"name": "Author",           "id": "filter_author",         "enabled": False},
                {"name": "Author type",      "id": "filter_autype",         "enabled": False},
                {"name": "Full text",        "id": "filter_ftext",          "enabled": False},
                {"name": "Translation",      "id": "filter_ftrans",         "enabled": False},
                {"name": "Title",            "id": "filter_title",          "enabled": False},
                {"name": "Section",          "id": "filter_sectiontitle",   "enabled": False},
                {"name": "Keyword",          "id": "filter_keyword",        "enabled": False}, 
                #{"name": "Feast",            "id": "filter_feast",          "enabled": False},
                #{"name": "Bible",            "id": "filter_bibref",         "enabled": False},
                {"name": "Note",             "id": "filter_note",           "enabled": False},
                {"name": "Status",           "id": "filter_stype",          "enabled": False},
                {"name": "solemne code",        "id": "filter_lilacode",       "enabled": False},
                {"name": "Free",             "id": "filter_freetext",       "enabled": False},
                {"name": "Project",          "id": "filter_project",        "enabled": False},

                {"name": "Collection...",    "id": "filter_collection",     "enabled": False, "head_id": "none"},
                {"name": "Manuscript...",    "id": "filter_manuscript",     "enabled": False, "head_id": "none"},

                {"name": "Canon witness",    "id": "filter_collcanwit",     "enabled": False, "head_id": "filter_collection"},
                {"name": "Authoritative statement", "id": "filter_collaustat","enabled": False, "head_id": "filter_collection"},
                {"name": "Manuscript",       "id": "filter_collmanu",       "enabled": False, "head_id": "filter_collection"},
                {"name": "Historical",       "id": "filter_collhc",         "enabled": False, "head_id": "filter_collection"},

                {"name": "Shelfmark",        "id": "filter_manuid",         "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Country",          "id": "filter_country",        "enabled": False, "head_id": "filter_manuscript"},
                {"name": "City",             "id": "filter_city",           "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Library",          "id": "filter_library",        "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Origin",           "id": "filter_origin",         "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Provenance",       "id": "filter_provenance",     "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Date from",        "id": "filter_datestart",      "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Date until",       "id": "filter_datefinish",     "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Manuscript type",  "id": "filter_manutype",       "enabled": False, "head_id": "filter_manuscript"},
                ]
    
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'ftext',         'dbfield': 'srchftext',         'keyS': 'ftext',  'regex': adapt_regex_incexp},
            {'filter': 'ftrans',        'dbfield': 'srchftrans',        'keyS': 'ftrans', 'regex': adapt_regex_incexp},
            {'filter': 'title',         'dbfield': 'title',             'keyS': 'srch_title'},
            {'filter': 'sectiontitle',  'dbfield': 'sectiontitle',      'keyS': 'srch_sectiontitle'},
            {'filter': 'feast',         'fkfield': 'feast',             'keyFk': 'feast', 'keyList': 'feastlist', 'infield': 'id'},
            {'filter': 'note',          'dbfield': 'note',              'keyS': 'note'},
            {'filter': 'bibref',        'dbfield': '$dummy',            'keyS': 'bibrefbk'},
            {'filter': 'bibref',        'dbfield': '$dummy',            'keyS': 'bibrefchvs'},
            {'filter': 'freetext',      'dbfield': '$dummy',            'keyS': 'free_term'},
            {'filter': 'freetext',      'dbfield': '$dummy',            'keyS': 'free_include'},
            {'filter': 'freetext',      'dbfield': '$dummy',            'keyS': 'free_exclude'},
            # {'filter': 'lilacode',      'dbfield': 'lilacodefull', 'keyS': 'lilacode', 'keyFk': 'code', 'keyList': 'lilalist', 'infield': 'id'},
            {'filter': 'lilacode',      'dbfield': 'lilacodefull',      'keyS': 'lilacode', 'keyList': 'lilalist', 'infield': 'id'},
            {'filter': 'author',        'fkfield': 'author',            'keyS': 'authorname',
                                        'keyFk': 'name', 'keyList': 'authorlist', 'infield': 'id', 'external': 'sermo-authorname' },
            {'filter': 'autype',                                        'keyS': 'authortype',  'help': 'authorhelp'},
            {'filter': 'keyword',       'fkfield': 'keywords',          'keyFk': 'name', 'keyList': 'kwlist', 'infield': 'id' }, 
            {'filter': 'project',       'fkfield': 'projects',          'keyFk': 'name', 'keyList': 'projlist', 'infield': 'name'},
            {'filter': 'stype',         'dbfield': 'stype',             'keyList': 'stypelist', 'keyType': 'fieldchoice', 'infield': 'abbr' }
            ]},
        {'section': 'collection', 'filterlist': [
            {'filter': 'collmanu',      'fkfield': 'manu__collections',           'keyFk': 'name', 'keyList': 'collist_m',    'infield': 'id' }, 
            {'filter': 'collcanwit',    'fkfield': 'collections',                 'keyFk': 'name', 'keyList': 'collist_s',    'infield': 'id' }, 
            {'filter': 'collaustat',    'fkfield': 'austats__collections',        'keyFk': 'name', 'keyList': 'collist_ssg',  'infield': 'id' }, 
            {'filter': 'collhc',        'fkfield': 'austats__collections',        'keyFk': 'name', 'keyList': 'collist_hist', 'infield': 'id' }
            ]},
        {'section': 'manuscript', 'filterlist': [
            {'filter': 'manuid',        'fkfield': 'msitem__manu',                    
             'keyS': 'manuidno',     'keyList': 'manuidlist', 'keyFk': 'idno', 'infield': 'id'},
            {'filter': 'country',       'fkfield': 'msitem__manu__library__lcountry', 
             'keyS': 'country_ta',   'keyId': 'country',     'keyFk': "name"},
            {'filter': 'city',          'fkfield': 'msitem__manu__library__lcity',    
             'keyS': 'city_ta',      'keyId': 'city',        'keyFk': "name"},
            {'filter': 'library',       'fkfield': 'msitem__manu__library',           
             'keyS': 'libname_ta',   'keyId': 'library',     'keyFk': "name"},
            {'filter': 'origin',        'fkfield': 'msitem__codico__origins',          'keyS': 'origin_ta',    'keyId': 'origin',      'keyFk': "name"},
            {'filter': 'provenance',    'fkfield': 'msitem__codico__provenances|msitem__codico__provenances__location',     
             'keyS': 'prov_ta',      'keyId': 'prov',        'keyFk': "name"},
            {'filter': 'datestart',     'dbfield': 'msitem__codico__codico_dateranges__yearstart__gte',     'keyS': 'date_from'},
            {'filter': 'datefinish',    'dbfield': 'msitem__codico__codico_dateranges__yearfinish__lte',    'keyS': 'date_until'},
            {'filter': 'manutype',      'dbfield': 'msitem__manu__mtype',     'keyS': 'manutype',     'keyType': 'fieldchoice', 'infield': 'abbr'},
            ]},
        {'section': 'other', 'filterlist': [
            {'filter': 'mtype',     'dbfield': 'mtype',    'keyS': 'mtype'},
            {'filter': 'sigauto',   'fkfield': 'austats__equal_goldsermons__goldsignatures', 'keyList':  'siglist_a', 'infield': 'id'},
            {'filter': 'sigmanu',   'fkfield': 'canwitsignatures',                              'keyList':  'siglist_m', 'infield': 'id'},
            {'filter': 'atype',     'dbfield': 'canwit_austat__austat__atype',    'keyS': 'atype'}
            #{'filter': 'appr_type', 'fkfield': 'austats__', 'keyList':' ', 'infield': }
            ]}
         ]

    def initializations(self):
        oErr = ErrHandle()
        try:
            # ======== One-time adaptations ==============
            listview_adaptations("canwit_list")

            # Check if there are any sermons not connected to a manuscript: remove these
            delete_id = Canwit.objects.filter(Q(msitem__isnull=True)|Q(msitem__manu__isnull=True)).values('id')
            if len(delete_id) > 0:
                oErr.Status("Deleting {} sermons that are not connected".format(len(delete_id)))
                Canwit.objects.filter(id__in=delete_id).delete()

            # Make sure to set a basic filter
            self.basic_filter = Q(mtype="man")
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SermonListiew/initializations")
        return None

    def add_to_context(self, context, initial):
        # Find out who the user is
        profile = Profile.get_user_profile(self.request.user.username)
        context['basketsize'] = 0 if profile == None else profile.basketsize
        context['basket_show'] = reverse('basket_show')
        context['basket_update'] = reverse('basket_update')
        return context

    def get_basketqueryset(self):
        if self.basketview:
            profile = Profile.get_user_profile(self.request.user.username)
            qs = profile.basketitems.all()
        else:
            qs = Canwit.objects.all()
        return qs

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []
        if custom == "author":
            if instance.author:
                html.append("<span style='color: darkgreen; font-size: small;'>{}</span>".format(instance.author.name[:20]))
                sTitle = instance.author.name
            else:
                html.append("<span><i>(unknown)</i></span>")
        elif custom == "signature":
            html.append(instance.signature_string(include_auto=True, do_plain=False))
        elif custom == "ftext":
            html.append("<span>{}</span>".format(instance.get_ftext_markdown()))
        elif custom == "ftrans":
            html.append("<span>{}</span>".format(instance.get_ftrans_markdown()))
        elif custom == "manuscript":
            manu = instance.get_manuscript()
            if manu == None:
                html.append("-")
            else:
                if manu.idno == None:
                    sIdNo = "-"
                else:
                    sIdNo = manu.idno[:20]
                html.append("<a href='{}' class='nostyle'><span style='font-size: small;'>{}</span></a>".format(
                    reverse('manuscript_details', kwargs={'pk': manu.id}),
                    sIdNo))
                sTitle = manu.idno
        elif custom == "lilacode":
            html.append(instance.get_lilacode())
            # Against using the value that is checked and updated for each save():
            #   this might not work if a related manuscript or collection is changed

        # ==== issue #10: title and sectiontitle are not or almost not used ====
        #elif custom == "title":
        #    sTitle = ""
        #    if instance.title != None and instance.title != "":
        #        sTitle = instance.title
        #    html.append(sTitle)
        #elif custom == "sectiontitle":
        #    sSection = ""
        #    if instance.sectiontitle != None and instance.sectiontitle != "":
        #        sSection = instance.sectiontitle
        #    html.append(sSection)

        elif custom == "status":
            # Provide that status badge
            # html.append("<span class='badge' title='{}'>{}</span>".format(instance.get_stype_light(), instance.stype[:1]))
            html.append(instance.get_stype_light())


            
        # Combine the HTML code
        sBack = "\n".join(html)
        return sBack, sTitle

    def adapt_search(self, fields):
        # Adapt the search to the keywords that *may* be shown
        lstExclude=[]
        qAlternative = None
        oErr = ErrHandle()

        try:
            # Make sure we show MANUSCRIPTS (identifiers) as well as reconstructions
            lstExclude.append(Q(mtype='tem') )
            ## Make sure to only show mtype manifestations
            #fields['mtype'] = "man"

            manutype = fields.get('manutype')
            if manutype != None and manutype != "":
                if manutype.abbr == "rec":
                    # Restrict to sermons that are part of a codico that is in table Reconstruction
                    codicolist = [x.codico.id for x in Reconstruction.objects.all()]
                    fields['manutype'] = Q(msitem__codico__id__in=codicolist)

            # Check if a list of keywords is given
            if 'kwlist' in fields and fields['kwlist'] != None and len(fields['kwlist']) > 0:
                # Get the list
                kwlist = fields['kwlist']
                # Get the user
                username = self.request.user.username
                user = User.objects.filter(username=username).first()
                # Check on what kind of user I am
                if not user_is_ingroup(self.request, app_editor):
                    # Since I am not an app-editor, I may not filter on keywords that have visibility 'edi'
                    # issue #10, this doesn't work: kwlist = Keyword.objects.filter(id__in=kwlist).exclude(Q(visibility="edi")).values('id')
                    kwlist = Keyword.objects.filter(id__in=kwlist).exclude(Q(visibility="edi"))
                    fields['kwlist'] = kwlist
            
            # Check if a list of projects is given
            if 'projlist' in fields and fields['projlist'] != None and len(fields['projlist']) > 0:
                # Get the list
                projlist = fields['projlist']

            # Adapt the bible reference list
            bibrefbk = fields.get("bibrefbk", "")
            if bibrefbk != None and bibrefbk != "":
                bibrefchvs = fields.get("bibrefchvs", "")

                # Get the start and end of this bibref
                start, einde = Reference.get_startend(bibrefchvs, book=bibrefbk)
 
                # Find out which sermons have references in this range
                lstQ = []
                lstQ.append(Q(sermonbibranges__bibrangeverses__bkchvs__gte=start))
                lstQ.append(Q(sermonbibranges__bibrangeverses__bkchvs__lte=einde))
                sermonlist = [x.id for x in Canwit.objects.filter(*lstQ).order_by('id').distinct()]

                fields['bibrefbk'] = Q(id__in=sermonlist)

            # Adapt the search for empty authors
            if 'authortype' in fields:
                authortype = fields['authortype']
                if authortype == "non":
                    # lstExclude = []
                    lstExclude.append(Q(author__isnull=False))
                elif authortype == "spe":
                    # lstExclude = []
                    lstExclude.append(Q(author__isnull=True))
                else:
                    # Reset the authortype
                    fields['authortype'] = ""

            # Adapt according to the 'free' fields
            free_term = fields.get("free_term", "")
            if free_term != None and free_term != "":
                all_fields = "all_fields"

                free_include = fields.get("free_include", [])
                free_exclude = fields.get("free_exclude", [])

                # Look for include field(s)
                if len(free_include) == 1 and free_include[0].field == all_fields:
                    # Include all fields from Free
                    s_q_i_lst = ""
                    for obj in Free.objects.exclude(field=all_fields):
                        val = free_term
                        if "*" in val or "#" in val:
                            val = adapt_search(val)
                            s_q = Q(**{"{}__iregex".format(obj.field): val})
                        elif "@" in val:
                            val = val.replace("@", "").strip()
                            s_q = Q(**{"{}__icontains".format(obj.field): val})
                        else:
                            s_q = Q(**{"{}__iexact".format(obj.field): val})
                        if s_q_i_lst == "":
                            s_q_i_lst = s_q
                        else:
                            s_q_i_lst |= s_q
                else:
                    s_q_i_lst = ""
                    for obj in free_include:
                        if obj.field == all_fields:
                            # skip it
                            pass
                        else:
                            val = free_term
                            if "*" in val or "#" in val:
                                val = adapt_search(val)
                                s_q = Q(**{"{}__iregex".format(obj.field): val})
                            elif "@" in val:
                                val = val.replace("@", "").strip()
                                s_q = Q(**{"{}__icontains".format(obj.field): val})
                            else:
                                s_q = Q(**{"{}__iexact".format(obj.field): val})
                            if s_q_i_lst == "":
                                s_q_i_lst = s_q
                            else:
                                s_q_i_lst |= s_q

                # Look for exclude fields
                if len(free_exclude) == 1 and free_exclude[0].field == all_fields:
                    # Include all fields from Free
                    s_q_e_lst = ""
                    for obj in Free.objects.exclude(field=all_fields):
                        val = free_term
                        if "*" in val or "#" in val:
                            val = adapt_search(val)
                            s_q = Q(**{"{}__iregex".format(obj.field): val})
                        elif "@" in val:
                            val = val.replace("@", "").strip()
                            s_q = Q(**{"{}__icontains".format(obj.field): val})
                        else:
                            s_q = Q(**{"{}__iexact".format(obj.field): val})
                        if s_q_e_lst == "":
                            s_q_e_lst = s_q
                        else:
                            s_q_e_lst |= s_q
                else:
                    s_q_e_lst = ""
                    for obj in free_exclude:
                        if obj.field == all_fields:
                            # skip it
                            pass
                        else:
                            val = free_term
                            if "*" in val or "#" in val:
                                val = adapt_search(val)
                                s_q = Q(**{"{}__iregex".format(obj.field): val})
                            elif "@" in val:
                                val = val.replace("@", "").strip()
                                s_q = Q(**{"{}__icontains".format(obj.field): val})
                            else:
                                s_q = Q(**{"{}__iexact".format(obj.field): val})
                            if s_q_e_lst == "":
                                s_q_e_lst = s_q
                            else:
                                s_q_e_lst |= s_q

                if s_q_i_lst != "":
                    qAlternative = s_q_i_lst
                if s_q_e_lst != "":
                    lstExclude.append( s_q_e_lst )

                # CLear the fields
                fields['free_term'] = "yes"
                fields['free_include'] = ""
                fields['free_exclude'] = ""
            # Double check the length of the exclude list
            if len(lstExclude) == 0:
                lstExclude = None
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SermonListView/adapt_search")
        
        # Make sure we only use the Authoritative statements with accepted modifications
        # This means that atype should be 'acc' (and not: 'mod', 'rej' or 'def') 
        # With this condition we make sure ALL sermons are in de unfiltered listview
        if fields['lilacode'] != '':
            fields['atype'] = 'acc'

        return fields, lstExclude, qAlternative

    def view_queryset(self, qs):
        search_id = [x['id'] for x in qs.values('id')]
        profile = Profile.get_user_profile(self.request.user.username)
        profile.search_canwit = json.dumps(search_id)
        profile.save()
        return None

    def get_helptext(self, name):
        """Use the get_helptext function defined in models.py"""
        return get_helptext(name)


# ============= CANONICAL EDITION VIEWS ============================


class CanedEdit(BasicDetails):
    """The editable part of one caned description (manifestation)"""
    
    model = Caned
    mForm = CanedForm
    prefix = "caned"
    title = "Canon Edition" 
    rtype = "json"
    mainitems = []
    basic_name = "caned"
    use_team_group = True
    history_button = True
    prefix_type = "simple"

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        oErr = ErrHandle()
        try:
            # Need to know who this user (profile) is
            profile = Profile.get_user_profile(self.request.user.username)

            # Define the main items to show and edit
            collection = instance.collection
            austat = instance.austat
            order = instance.order
            context['mainitems'] = [
                # -------- HIDDEN field values ---------------
                {'type': 'plain', 'label': "Collection id", 'value': collection.id, 'field_key': "collection",  'empty': 'hide'},
                {'type': 'plain', 'label': "Austat id",     'value': austat.id,     'field_key': "austat",      'empty': 'hide'},
                {'type': 'plain', 'label': "Order",         'value': order,         'field_key': "order",       'empty': 'hide'},
                # --------------------------------------------
                {'type': 'safe',  'label': "Authoritative Statement:",  'value': instance.get_austat()                      }, 
                {'type': 'safe',  'label': "Historical Collection:",    'value': instance.get_collection()                  }, 
                {'type': 'plain', 'label': "LiLaC code:",               'value': instance.get_lilacode() ,
                    'title': 'The lilac code combines the [Historical Collection] and the [CanEd number]'}, 
                {'type': 'plain', 'label': "CanEd number:",             'value': instance.get_idno(),   'field_key': "idno",
                    'title': 'The [CanEd number] is the number of the canonical (authoritative) statement within the edition' }, 
                {'type': 'plain', 'label': "Order in HC:",              'value': instance.order                             }, 
                {'type': 'safe',  'label': "Full text:",                'value': instance.get_ftext_markdown(), 
                 'field_key': 'ftext',  'key_ta': 'srmincipit-key'}, 
                {'type': 'safe',  'label': "Translation:",              'value': instance.get_ftrans_markdown(),
                 'field_key': 'ftrans', 'key_ta': 'srmexplicit-key'}, 
                ]

            # Add something right to the CanEd title
            context['title_addition'] = instance.get_breadcrumb()

            # Signal that we have select2
            context['has_select2'] = True

        except:
            msg = oErr.get_error_message()
            oErr.DoError("CanedEdit/add_to_context")

        # Return the context we have made
        return context

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        solemne_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return solemne_get_history(instance)


class CanedDetails(CanedEdit):
    """The details of one sermon manifestation (Caned)"""

    rtype = "html"

    def process_formset(self, prefix, request, formset):
        return None

    def after_save(self, form, instance):
        return True, ""


class CanedListView(BasicList):
    """Search and list sermons"""
    
    model = Caned
    listform = CanedForm
    has_select2 = True
    use_team_group = True
    paginate_by = PAGINATE_BY_VALUE
    bUseFilter = True
    prefix = "caned"
    new_button = False      # Don't show the [Add new caned] button here. It is shown under the Manuscript Details view.
    basketview = False
    plural_name = "Canon editions"
    basic_name = "caned"
    template_help = "seeker/filter_help.html"

    order_cols = ['austat__auwork__key;austat__keycode', 'collection__lilacode', 'collection__lilacode;idno', '']
    order_default = order_cols
    order_heads = [
        {'name': 'Authoritative Statement', 'order': 'o=1', 'type': 'str', 'custom': 'austat', 'linkdetails': True}, 
        {'name': 'Historical Collection',   'order': 'o=2', 'type': 'str', 'custom': 'histcol','linkdetails': True}, 
        {'name': 'LiLaC code',              'order': 'o=3', 'type': 'str', 'custom': 'lilac',  'linkdetails': True, 'main': True},
        {'name': 'Canwit count',            'order': '', 'type': 'int', 'custom': 'canwit', 'linkdetails': True},
        ]

    filters = [ 
        {"name": "Authortative statement",  "id": "filter_austat",      "enabled": False},
        {"name": "Historical collection",   "id": "filter_collhist",    "enabled": False},
        {"name": "LiLaC code",              "id": "filter_lilac",       "enabled": False},
        {"name": "Full text",               "id": "filter_ftext",       "enabled": False},
        {"name": "Translation",             "id": "filter_ftrans",      "enabled": False},
                ]
    
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'collhist',      'fkfield': 'collection',        'keyFk': 'name',        
                'keyList': 'collist_hist',  'infield': 'name' },
            {'filter': 'austat',        'fkfield': 'austat',            'keyFk': 'keycodefull', 'keyS': 'ssgcode',
                'keyList': 'ssglilalist',  'infield': 'id'},
            {'filter': 'ftext',         'dbfield': 'srchftext',         'keyS': 'ftext',  'regex': adapt_regex_incexp},
            {'filter': 'ftrans',        'dbfield': 'srchftrans',        'keyS': 'ftrans', 'regex': adapt_regex_incexp}

            ]},
         {'section': 'other', 'filterlist': [
            {'filter': 'mtype',     'dbfield': 'mtype',    'keyS': 'mtype'},
            {'filter': 'sigauto',   'fkfield': 'austats__equal_goldsermons__goldsignatures', 'keyList':  'siglist_a', 'infield': 'id'},
            {'filter': 'sigmanu',   'fkfield': 'canedsignatures',                              'keyList':  'siglist_m', 'infield': 'id'},
            {'filter': 'atype',     'dbfield': 'caned_austat__austat__atype',    'keyS': 'atype'}
            #{'filter': 'appr_type', 'fkfield': 'austats__', 'keyList':' ', 'infield': }
            ]}
         ]

    def initializations(self):
        oErr = ErrHandle()
        try:
            # ======== One-time adaptations ==============
            listview_adaptations("caned_list")

        except:
            msg = oErr.get_error_message()
            oErr.DoError("CanedListView/initializations")
        return None

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []
        if custom == "austat":
            sAustat = instance.get_austat()
            url = reverse("austat_details", kwargs={'pk': instance.austat.id})
            html.append("<a href='{}' class='nostyle'><span style='font-size: small;'>{}</span></a>".format(
                url, sAustat))
        elif custom == "histcol":
            sHC = instance.get_collection()
            url = reverse("collhist_details", kwargs={'pk': instance.collection.id})
            html.append("<a href='{}' class='nostyle'><span style='font-size: small;'>{}</span></a>".format(
                url, sHC))
        elif custom == "lilac":
            html.append(instance.get_lilacode())
        elif custom == "canwit":
            # Count the number of canonical witnesses belonging to the austat
            count = CanwitAustat.objects.filter(austat= instance.austat).count()
            sTitle = "{}".format(count)
            html.append(sTitle)
            
        # Combine the HTML code
        sBack = "\n".join(html)
        return sBack, sTitle

    def get_helptext(self, name):
        """Use the get_helptext function defined in models.py"""
        return get_helptext(name)


# ============= AUTHORITATIVE STATEMENT VIEWS ============================


class AustatEdit(BasicDetails):
    model = Austat
    mForm = AustatForm
    prefix = 'as'
    title = "Authoritative Statement"
    rtype = "json"
    new_button = True
    mainitems = []
    use_team_group = True
    history_button = True
    
    EqgcolFormSet = inlineformset_factory(Austat, Caned,
                                       form=SuperSermonGoldCollectionForm, min_num=0,
                                       fk_name="austat", extra=0)
    SsgLinkFormSet = inlineformset_factory(Austat, AustatLink,
                                         form=AustatLinkForm, min_num=0,
                                         fk_name = "src",
                                         extra=0, can_delete=True, can_order=False)
    AlitFormSet = inlineformset_factory(Austat, LitrefAustat,
                                         form = LitrefAustatForm, min_num=0,
                                         fk_name = "austat",
                                         extra=0, can_delete=True, can_order=False)

        
    formset_objects = [
        {'formsetClass': EqgcolFormSet,  'prefix': 'eqgcol',  'readonly': False, 'noinit': True, 'linkfield': 'austat'},
        {'formsetClass': SsgLinkFormSet, 'prefix': 'ssglink', 'readonly': False, 'noinit': True, 'initial': [{'linktype': LINK_PARTIAL }], 'clean': True},     
        {'formsetClass': AlitFormSet,    'prefix': 'alit',    'readonly': False, 'noinit': True, 'linkfield': 'austat'},
        ]

    # Note: do not include [code] in here
    stype_edi_fields = ['author', 'number', 'ftext', 'ftrans',
                        'AustatLink', 'superlist',
                        'goldlist', 'projlist', 'litlist']

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        solemne_action_add(self, instance, details, actiontype)

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        oErr = ErrHandle()
        try:
            # One general item is the 'help-popup' to be shown when the user clicks on 'Author'
            info = render_to_string('seeker/author_info.html')

            # Need to know who this user (profile) is
            profile = Profile.get_user_profile(self.request.user.username)
            username = profile.user.username
            team_group = app_editor
            may_edit = (context['is_app_editor'])

            # Define the main items to show and edit
            author_id = None if instance.author is None else instance.author.id
            context['mainitems'] = [
                {'type': 'plain', 'label': "Status:",        'value': instance.get_stype_light(may_edit),'field_key': 'stype'},
                {'type': 'plain', 'label': "Author:",        'value': instance.author_help(info), 'field_key': 'newauthor'},

                # Issue #295: the [number] (number within author) must be there, though hidden, not editable
                {'type': 'plain', 'label': "Number:",        'value': instance.number,    'field_key': 'number',   'empty': 'hide'},
                {'type': 'plain', 'label': "Author id:",     'value': author_id,          'field_key': 'author',   'empty': 'hide'},
                {'type': 'plain', 'label': "Full text:",     'value': instance.ftext,     'field_key': 'ftext',    'empty': 'hide'},
                {'type': 'plain', 'label': "Translation:",   'value': instance.ftrans,    'field_key': 'ftrans',   'empty': 'hide'},

                # OLD {'type': 'plain', 'label': "Lilac Code:",   'value': instance.code,   'title': 'The lila Code is automatically determined'}, 
                {'type': 'plain', 'label': "Key Code:",     'value': instance.get_keycode(True), 'field_key': 'keycode'}, 
                {'type': 'plain', 'label': "Work:",         'value': instance.get_work(True), 
                 'field_key': 'auwork'}, 
                {'type': 'safe',  'label': "Full text:",    'value': instance.get_ftext_markdown("search"), 
                 'field_key': 'newftext',  'key_ta': 'gldftext-key', 'title': instance.get_ftext_markdown("actual")}, 
                {'type': 'safe',  'label': "Translation:",  'value': instance.get_ftrans_markdown("search"),
                 'field_key': 'newftrans', 'key_ta': 'gldftrans-key', 'title': instance.get_ftrans_markdown("actual")}, 

                # These three are taken over from Auwork (read only, non edit)
                {'type': 'plain', 'label': "Date:",         'value': instance.get_date()            }, 
                {'type': 'line',  'label': "Opus:",         'value': instance.get_opus()            }, 
                {'type': 'line',  'label': "Genre(s):",     'value': instance.get_genres_markdown() }, 
                {'type': 'line',  'label': "CPL:",          'value': instance.get_signatures(), 'title': 'Signature(s)' }, 

                {'type': 'line',  'label': "Keywords:",      'value': instance.get_keywords_markdown(), 'field_list': 'kwlist'},
                {'type': 'plain', 'label': "Keywords (user):", 'value': instance.get_keywords_user_markdown(profile),   'field_list': 'ukwlist',
                 'title': 'User-specific keywords. If the moderator accepts these, they move to regular keywords.'},
                {'type': 'bold',  'label': "Moved to:",      'value': instance.get_moved_code(), 'empty': 'hidenone', 'link': instance.get_moved_url()},
                {'type': 'bold',  'label': "Previous:",      'value': instance.get_previous_code(), 'empty': 'hidenone', 'link': instance.get_previous_url()},
                {'type': 'line',  'label': "Personal datasets:",   'value': instance.get_collections_markdown(username, team_group, settype="pd"), 
                    'multiple': True, 'field_list': 'collist_ssg', 'fso': self.formset_objects[0] },
                # Project HIER
                {'type': 'plain', 'label': "Project:",      'value': instance.get_project_markdown2(), 'field_list': 'projlist'},
            
                {'type': 'line',  'label': "Historical collections:",   'value': instance.get_collections_markdown(username, team_group, settype="hc"), 
                    'field_list': 'collist_hist', 'fso': self.formset_objects[0] },
                {'type': 'line',    'label': "Links:",  'title': "Authoritative statement links:",  'value': instance.get_superlinks_markdown(), 
                    'multiple': True,  'field_list': 'superlist',       'fso': self.formset_objects[1], 
                    'inline_selection': 'ru.solemne.as2as_template',   'template_selection': 'ru.solemne.ssg_template'},
                {'type': 'line', 'label': "Editions:",              'value': instance.get_editions_markdown(),
                 'title': 'All the editions associated with the Canonical Edition(s) in this equality set. Set the edition in the Authoritative Work!!'},
                {'type': 'line', 'label': "Literature:",            'value': instance.get_litrefs_markdown(), 
                 'title': 'All the literature references associated with the Gold Sermons in this equality set',
                 'multiple': True, 'field_list': 'litlist', 'fso': self.formset_objects[2], 'template_selection': 'ru.solemne.litref_template'},

                # Buttons to show or hide Manuscripts and (historical) Collections
                {'type': 'safe',   'label': 'Show/hide:',           'value': self.get_buttons(instance)}
                ]
            # Notes:
            # Collections: provide a link to the SSG-listview, filtering on those SSGs that are part of one particular collection

            # Possibly add 'field_key' for keycode, if no auwork
            if instance.auwork is None:
                # Find the Key Code
                for oItem in context['mainitems']:
                    if "Key Code" in oItem.get("label"):
                        oItem['field_key'] = "keycode"

            # Some tests can only be performed if this is *not* a new instance
            if not instance is None and not instance.id is None:

                # THe SSG items that have a value in *moved* may not be editable
                editable = (instance.moved == None)
                if not editable:
                    self.permission = "readonly"
                    context['permission'] = self.permission

                # Add comment modal stuff
                initial = dict(otype="austat", objid=instance.id, profile=profile)
                context['commentForm'] = CommentForm(initial=initial, prefix="com")
                context['comment_list'] = get_usercomments('austat', instance, profile)
                lhtml = []
                lhtml.append(render_to_string("seeker/comment_add.html", context, self.request))
                context['after_details'] = "\n".join(lhtml)

            # Add something right to the AustatDetails title
            context['title_addition'] = instance.get_breadcrumb()

            # Signal that we have select2
            context['has_select2'] = True

        except:
            msg = oErr.get_error_message()
            oErr.DoError("AustatEdit/add_to_context")

        # Return the context we have made
        return context

    def after_save(self, form, instance):

        def process_proj_adding(profile, projlist):
            oErr = ErrHandle()

            try:
                allow_adding = projlist
                if len(allow_adding) > 0:
                    # Some combinations of Project-SSG may be added right away
                    with transaction.atomic():
                        for oItem in allow_adding:
                            equal = oItem.get("austat")
                            project = oItem.get("project")
                            if not equal is None and not project is None:
                                obj = AustatProject.objects.create(equal=equal, project=project)
            except:
                msg = oErr.get_error_message()
                oErr.DoError("process_proj_adding")
            return True

        def process_proj_removing(profile, projlist):
            oErr = ErrHandle()

            try:
                if instance.projects.count() > 1:
                    allow_removing = projlist
                    if len(allow_removing) > 0:
                        # There are some project-SSG associations that may be removed right away
                        delete_id = []
                        for oItem in allow_removing:
                            equal = oItem.get("austat")
                            project = oItem.get("project")
                            if not equal is None and not project is None:
                                obj = AustatProject.objects.filter(equal=equal, project=project).first()
                                delete_id.append(obj.id)
                        # Now delete them
                        if len(delete_id) > 0:
                            AustatProject.objects.delete(delete_id)
            except:
                msg = oErr.get_error_message()
                oErr.DoError("process_proj_removing")
            return True

        def get_proj_add_remove(projlist):
            """Split the list of projects into those being added and those being removed"""
            oErr = ErrHandle()
            addprojlist = []
            delprojlist = []
            try:
                projcurrent = instance.projects.all()
                for obj in projlist:
                    # Check for addition
                    if obj in projlist and not obj in projcurrent:
                        addprojlist.append(obj)
                    elif obj in projcurrent and not obj in projlist:
                        delprojlist.append(obj)
            except:
                msg = oErr.get_error_message()
                oErr.DoError("get_proj_add_remove")
            return addprojlist, delprojlist


        msg = ""
        bResult = True
        oErr = ErrHandle()
                
        try:
            # Process many-to-many changes: Add and remove relations in accordance with the new set passed on by the user
            # (1) 'Personal Datasets' and 'Historical Collections'
            collist_ssg_id = form.cleaned_data['collist_ssg'].values('id') 
            collist_hist_id = form.cleaned_data['collist_hist'].values('id')
            collist_ssg = Collection.objects.filter(Q(id__in=collist_ssg_id) | Q(id__in=collist_hist_id))
            adapt_m2m(Caned, instance, "austat", collist_ssg, "collection")

            # (2) links from one SSG to another SSG
            superlist = form.cleaned_data['superlist']
            super_added = []
            super_deleted = []
            adapt_m2m(AustatLink, instance, "src", superlist, "dst", 
                      extra = ['linktype', 'alternatives', 'spectype', 'note'], related_is_through=True,
                      added=super_added, deleted=super_deleted)
            # Check for partial links in 'deleted'
            for obj in super_deleted:
                # This if-clause is not needed: anything that needs deletion should be deleted
                # if obj.linktype in LINK_BIDIR:
                # First find and remove the other link
                reverse = AustatLink.objects.filter(src=obj.dst, dst=obj.src, linktype=obj.linktype).first()
                if reverse != None:
                    reverse.delete()
                # Then remove myself
                obj.delete()
            # Make sure to add the reverse link in the bidirectionals
            for obj in super_added:
                if obj.linktype in LINK_BIDIR:
                    # Find the reversal
                    reverse = AustatLink.objects.filter(src=obj.dst, dst=obj.src, linktype=obj.linktype).first()
                    if reverse == None:
                        # Create the reversal 
                        reverse = AustatLink.objects.create(src=obj.dst, dst=obj.src, linktype=obj.linktype)
                        # Other adaptations
                        bNeedSaving = False
                        # Set the correct 'reverse' spec type
                        if obj.spectype != None and obj.spectype != "":
                          reverse.spectype = get_reverse_spec(obj.spectype)
                          bNeedSaving = True
                        # Possibly copy note
                        if obj.note != None and obj.note != "":
                          reverse.note = obj.note
                          bNeedSaving = True
                        # Need saving? Then save
                        if bNeedSaving:
                          reverse.save()

            # (3) 'genres'
            genrelist = form.cleaned_data['genrelist']
            adapt_m2m(AustatGenre, instance, "austat", genrelist, "genre")

            # (3) 'keywords'
            kwlist = form.cleaned_data['kwlist']
            adapt_m2m(AustatKeyword, instance, "equal", kwlist, "keyword")

            # (4) user-specific 'keywords'
            ukwlist = form.cleaned_data['ukwlist']
            profile = Profile.get_user_profile(self.request.user.username)
            adapt_m2m(UserKeyword, instance, "austat", ukwlist, "keyword", qfilter = {'profile': profile}, 
                      extrargs = {'profile': profile, 'type': 'austat'})

            # (5) 'literature'
            litlist = form.cleaned_data['litlist']
            adapt_m2m(LitrefAustat, instance, "austat", litlist, "reference", extra=['pages'], related_is_through = True)

            # (6) 'projects'
            projlist = form.cleaned_data['projlist']
            # === EK: CODE PRIOR TO #517 =====
            #equal_proj_deleted = []
            #adapt_m2m(AustatProject, instance, "equal", projlist, "project", deleted=equal_proj_deleted)
            #project_dependant_delete(self.request, equal_proj_deleted)
            # ================================
            # Figure out what are the proposed additions, and what are the proposed deletions
            addprojlist, delprojlist = get_proj_add_remove(projlist)
            process_proj_adding(profile, addprojlist)
            process_proj_removing(profile, delprojlist)

            # Issue #517: submit request to add this SSG to indicated project(s)
            # Process the line "Add to a project"
            addprojlist = form.cleaned_data.get("addprojlist")
            process_proj_adding(profile, addprojlist)

            # Process the line "Remove from a project"
            delprojlist = form.cleaned_data.get("delprojlist")
            process_proj_removing(profile, delprojlist)

            # Issue #473: default project assignment
            if instance.projects.count() == 0:
                # Need to know who is 'talking'...
                username = self.request.user.username
                profile = Profile.get_user_profile(username)

                # The user has not selected a project (yet): try default assignment
                user_projects = profile.projects.all()
                if user_projects.count() == 1:
                    project = profile.projects.first()
                    AustatProject.objects.create(equal=instance, project=project)

            # Process many-to-ONE changes

            # ADDED Take over any data from [instance] to [frm.data]
            #       Provided these fields are in the form's [initial_fields]
            if instance != None:

                # Walk the fields that need to be taken from the instance
                for key in form.initial_fields:
                    value = getattr(instance, key)

                    key_prf = '{}-{}'.format(form.prefix, key)
                    if isinstance(value, str) or isinstance(value, int):
                        form.data[key_prf] = value
                    elif isinstance(value, object):
                        form.data[key_prf] = str(value.id)

        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg

    def before_save(self, form, instance):
        oErr = ErrHandle()
        bBack = True
        msg = ""
        transfer_changes = [
            {'src': 'newftext',    'dst': 'ftext',    'type': 'text'},
            {'src': 'newftrans',   'dst': 'ftrans',   'type': 'text'},
            {'src': 'newauthor',   'dst': 'author',   'type': 'fk'},
            ]
        try:
            if instance != None:
                # Need to know who is 'talking'...
                username = self.request.user.username
                profile = Profile.get_user_profile(username)

                # Get the cleaned data: this is the new stuff
                cleaned_data = form.cleaned_data

                # This means that any changes may be implemented right away
                for oTransfer in transfer_changes:
                    type = oTransfer.get("type")
                    src_field = oTransfer.get("src")
                    dst_field = oTransfer.get("dst")
                    src_value = cleaned_data.get(src_field)
                        
                    # Transfer the value
                    if type == "fk" or type == "text":
                        # Is there any change?
                        prev_value = getattr(instance, dst_field)
                        if src_value != prev_value:
                            # Special cases
                            if dst_field == "author":
                                authornameLC = instance.author.name.lower()
                                # Determine what to do in terms of 'moved'.
                                if authornameLC != "undecided":
                                    # Create a copy of the object I used to be
                                    moved = Austat.create_moved(instance)

                            # Perform the actual change
                            setattr(form.instance, dst_field, src_value)

                # Check for author
                if instance.author == None:
                    # Set to "undecided" author if possible
                    author = Author.get_undecided()
                    instance.author = author

                # Issue #473: automatic assignment of project for particular editor(s)
                projlist = form.cleaned_data.get("projlist")
                bBack, msg = evaluate_projlist(profile, instance, projlist, "Authoritative statement")

        except:
            msg = oErr.get_error_message()
            oErr.DoError("AustatEdit/before_save")
            bBack = False
        return bBack, msg

    def get_buttons(self, instance):
        """Get HTML code for buttons to show the Manuscripts and/or Collections"""

        oErr = ErrHandle()
        sBack = "-"
        try:
            html = []
            # Button for Manuscripts (actually: canonical witnesses)
            count_m = CanwitAustat.objects.filter(austat=instance).exclude(canwit__mtype="tem").count()
            if count_m > 0:
                title = ""
                url = "basic_manu_set"
                # html.append("<span data-toggle='collapse' data-target='#{}'>".format(url))
                html.append("<span related-target='#{}'>".format(url))
                html.append("<a class='btn btn-xs jumbo-1'>Canonical witnesses <span style='color: red;'>{}</span></a></span> ".format(count_m))
                html.append("<span>&nbsp;</span>")

            # Button for Historical Collections
            count_hc = instance.collections.filter(settype="hc").count()
            if count_hc > 0:
                title = ""
                url = "basic_hist_set"
                # html.append("<span data-toggle='collapse' data-target='#{}'>".format(url))
                html.append("<span related-target='#{}'>".format(url))
                html.append("<a class='btn btn-xs jumbo-3'>Historical collections <span style='color: red;'>{}</span></a></span> ".format(count_hc))

            sBack = "\n".join(html)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("AustatEdit/get_buttons")

        return sBack

    def get_form_kwargs(self, prefix):
        # This is for ssglink

        oBack = None
        if prefix == "ssglink":
            if self.object != None:
                # Make sure to return the ID of the Austat
                oBack = dict(austat_id=self.object.id)

        return oBack

    def get_history(self, instance):
        return solemne_get_history(instance)

    def process_formset(self, prefix, request, formset):
        errors = []
        bResult = True
        instance = formset.instance
        # Need to know who is 'talking'...
        username = self.request.user.username
        profile = Profile.get_user_profile(username)
        for form in formset:
            if form.is_valid():
                oErr = ErrHandle()
                try:
                    cleaned = form.cleaned_data
                    # Action depends on prefix
                
                    # Note: eqgcol can be either settype 'pd' or 'hc'
                    if prefix == "eqgcol":
                        # Keyword processing
                        if 'newcol' in cleaned and cleaned['newcol'] != "":
                            newcol = cleaned['newcol']
                            profile = Profile.get_user_profile(request.user.username)
                            # Is the COL already existing?
                            obj = Collection.objects.filter(name=newcol).first()
                            if obj == None:
                                # TODO: add profile here
                                obj = Collection.objects.create(name=newcol, type='austat', owner=profile)
                            # once a collection has been created, make sure it gets assigned to a project
                            if not profile is None and obj.projects.count() == 0:
                                # Assign the default projects
                                projects = profile.get_defaults()
                                obj.set_projects(projects)
                            # Make sure we set the keyword
                            form.instance.collection = obj
                            # Note: it will get saved with formset.save()
                    elif prefix == "ssglink":
                        # Canwit-To-Austat processing
                        newaustat = cleaned.get("newsuper")
                        if not newaustat is None:
                            # There also must be a linktype
                            if 'newlinktype' in cleaned and cleaned['newlinktype'] != "":
                                linktype = cleaned['newlinktype']
                                # Get optional parameters
                                note = cleaned.get('note', None)
                                spectype = cleaned.get('newspectype', None)
                                # Alternatives: this is true if it is in there, and false otherwise
                                alternatives = cleaned.get("newalt", None)
                                # Check existence
                                obj = AustatLink.objects.filter(src=instance, dst=newaustat, linktype=linktype).first()
                                if obj == None:
                                    austat = Austat.objects.filter(id=newaustat.id).first()
                                    if austat != None:
                                        # Set the right parameters for creation later on
                                        form.instance.linktype = linktype
                                        form.instance.dst = austat
                                        if note != None and note != "": 
                                            form.instance.note = note
                                        if spectype != None and len(spectype) > 1:
                                            form.instance.spectype = spectype
                                        form.instance.alternatives = alternatives

                                        # Double check reverse
                                        if linktype in LINK_BIDIR:
                                            rev_link = AustatLink.objects.filter(src=austat, dst=instance).first()
                                            if rev_link == None:
                                                # Add it
                                                rev_link = AustatLink.objects.create(src=austat, dst=instance, linktype=linktype)
                                            else:
                                                # Double check the linktype
                                                if rev_link.linktype != linktype:
                                                    rev_link.linktype = linktype
                                            if note != None and note != "": 
                                                rev_link.note = note
                                            if spectype != None and len(spectype) > 1:
                                                rev_link.spectype = get_reverse_spec(spectype)
                                            rev_link.alternatives = alternatives
                                            rev_link.save()

                        # Note: it will get saved with form.save()
                    elif prefix == "alit":
                        # Edition processing
                        newpages = ""
                        if 'newpages' in cleaned and cleaned['newpages'] != "":
                            newpages = cleaned['newpages']
                        # Also get the litref
                        if 'oneref' in cleaned:
                            litref = cleaned['oneref']
                            # Check if all is in order
                            if litref:
                                form.instance.reference = litref
                                if newpages:
                                    form.instance.pages = newpages
                        # Note: it will get saved with form.save()

                except:
                    msg = oErr.get_error_message()
                    oErr.DoError("AustatEdit/process_formset")
            else:
                errors.append(form.errors)
                bResult = False
        return None


class AustatDetails(AustatEdit):
    rtype = "html"

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Start by executing the standard handling
        context = super(AustatDetails, self).add_to_context(context, instance)

        oErr = ErrHandle()
        related_objects = []
        resizable = True
        sort_start = ""
        sort_start_mix = ""
        sort_start_int = ""
        sort_end = ""
        try:
            if instance != None and instance.id != None:
                context['sections'] = []

                username = self.request.user.username
                team_group = app_editor
                profile = Profile.get_user_profile(username=username)

                # Authorization: only app-editors may edit!
                bMayEdit = user_is_ingroup(self.request, team_group)
            
                # Anyone may sort the contents, even non-editors
                # Lists of related objects
                sort_start = '<span class="sortable"><span class="fa fa-sort sortshow"></span>&nbsp;'
                sort_start_mix = '<span class="sortable mixed"><span class="fa fa-sort sortshow"></span>&nbsp;'
                sort_start_int = '<span class="sortable integer"><span class="fa fa-sort sortshow"></span>&nbsp;'
                sort_end = '</span>'


                # ============= List of manuscripts related to the Austat via canwit descriptions ==================
                manuscripts = dict(title="Canonical witnesses in their Manuscripts", prefix="manu", gridclass="resizable", classes="hidden")

                # Get all the Canwit instances linked with equality to SSG:
                # But make sure the EXCLUDE those with `mtype` = `tem`
                qs_s = CanwitAustat.objects.filter(austat=instance).exclude(canwit__mtype="tem").order_by(
                    'canwit__msitem__manu__idno', 'canwit__locus')
                rel_list =[]
                for canwitlink in qs_s:
                    canwit = canwitlink.canwit
                    # Get the 'item': the manuscript
                    item = canwit.msitem.manu
                    rel_item = []
                
                    # Shelfmark = CITY - LIBRARY - IDNO
                    url = reverse('manuscript_details', kwargs={'pk': item.id})
                    add_rel_item(rel_item, self.get_field_value("manu", item, "shelfmark"), resizable, link=url, main=True)

                    # Origin
                    add_rel_item(rel_item, self.get_field_value("manu", item, "origin"), resizable, 
                                 title="Origin (if known), followed by provenances (between brackets)")

                    # date range
                    add_rel_item(rel_item, self.get_field_value("manu", item, "daterange"), resizable, align="right")

                    # Collection(s)
                    #coll_info = item.get_collections_markdown(username, team_group)
                    #rel_item.append({'value': coll_info, 'initial': 'small'})
                    add_rel_item(rel_item, self.get_field_value("manu", item, "collections"), resizable)

                    # Location number and link to the correct point in the manuscript details view...
                    # itemloc = "{}/{}".format(canwit.msitem.order, item.get_canwit_count())
                    # link_on_manu_page = "{}#canwit_{}".format(reverse('manuscript_details', kwargs={'pk': item.id}), canwit.id)
                    url_canwit = reverse('canwit_details', kwargs={'pk': canwit.id})
                    #rel_item.append({'value': itemloc, 'align': "right", 'title': 'Jump to the canwit in the manuscript', 'initial': 'small',
                    #                    'link': link_to_canwit })
                    add_rel_item(rel_item, self.get_field_value("manu", canwit, "canwit"), resizable, align="right",
                                 title="Jump to the canwit in the manuscript", link=url_canwit)

                    # Folio number of the item
                    # rel_item.append({'value': canwit.locus, 'initial': 'small'})
                    add_rel_item(rel_item, self.get_field_value("manu", canwit, "locus"), resizable)

                    # Attributed author
                    # rel_item.append({'value': canwit.get_author(), 'initial': 'small'})
                    add_rel_item(rel_item, self.get_field_value("manu", canwit, "author"), resizable)

                    # Ftext
                    # rel_item.append({'value': canwit.get_ftext_markdown()}) #, 'initial': 'small'})
                    add_rel_item(rel_item, self.get_field_value("manu", canwit, "ftext"), resizable)

                    # Ftrans
                    #rel_item.append({'value': canwit.get_ftrans_markdown()}) #, 'initial': 'small'})
                    add_rel_item(rel_item, self.get_field_value("manu", canwit, "ftrans"), resizable)

                    # Keywords
                    #rel_item.append({'value': canwit.get_keywords_markdown(), 'initial': 'small'})
                    add_rel_item(rel_item, self.get_field_value("manu", canwit, "keywords"), resizable)

                    # Add this Manu/Canwit line to the list
                    rel_list.append(dict(id=item.id, cols=rel_item))
                manuscripts['rel_list'] = rel_list

                manuscripts['columns'] = [
                    '{}<span title="Shelfmark of the manuscript in which the Canonical witness is">Shelfmark</span>{}'.format(sort_start_int, sort_end), 
                    '{}<span title="Origin/Provenance">or./prov.</span>{}'.format(sort_start_int, sort_end), 
                    '{}<span title="Date range">date</span>{}'.format(sort_start_int, sort_end), 
                    '{}<span title="Collection name">coll.</span>{}'.format(sort_start_int, sort_end), 
                    '{}<span title="Item">item</span>{}'.format(sort_start_mix, sort_end), 
                    '{}<span title="Folio number">ff.</span>{}'.format(sort_start_int, sort_end), 
                    '{}<span title="Attributed author">auth.</span>{}'.format(sort_start_int, sort_end), 
                    '{}<span title="Full text">txt.</span>{}'.format(sort_start_int, sort_end), 
                    '{}<span title="Translation">trns.</span>{}'.format(sort_start_int, sort_end), 
                    '{}<span title="Keywords of the Sermon manifestation">keyw.</span>{}'.format(sort_start_int, sort_end), 
                    ]

                # Add the manuscript to the related objects
                related_objects.append(manuscripts)

                # ============= List of historical collections related to the Austat  ==============================
                collections = dict(title="Historical collections", prefix="hist", gridclass="resizable", classes="")

                # Get all historical collections (including private ones)
                qs_hc = instance.collections.filter(settype="hc").order_by("name")
                qs_hc = Caned.objects.filter(austat=instance).order_by(
                    "collection__name")
                rel_list = []
                for obj in qs_hc:
                    rel_item = []
                    # The [obj] is a Caned. Now get to the actual Collection
                    item = obj.collection
                    
                    # Make sure we have the link to the HC
                    url = reverse("collhist_details", kwargs={'pk': item.id})

                    # HC: Order of Austat within collection
                    add_rel_item(rel_item, obj.order, False, align="right")

                    # HC: Name of collection
                    add_rel_item(rel_item, self.get_field_value("collection", item, "name"), resizable, link=url)

                    # HC: Manuscript + Canonical witness linked to collection
                    add_rel_item(rel_item, self.get_field_value("collection", item, "canwits"), resizable, main=True, nowrap=False)

                    # HC: Owner of collection
                    add_rel_item(rel_item, self.get_field_value("collection", item, "owner"), resizable, link=url)

                    # HC: Scope of collection
                    add_rel_item(rel_item, self.get_field_value("collection", item, "scope"), resizable, link=url)

                    # HC: Number of authors
                    add_rel_item(rel_item, self.get_field_value("collection", item, "authnum"), resizable, link=url, align="right")


                    # Add this line to the list
                    rel_list.append(dict(id=item.id, cols=rel_item))

                collections['rel_list'] = rel_list

                collections['columns'] = [
                    '{}<span title="Order">Order<span>{}'.format(sort_start_int, sort_end),
                    '{}<span title="Name of the historical collection">Name</span>{}'.format(sort_start_int, sort_end), 
                    '{}<span title="Manuscripts with canonical witnesses in this collection">Manuscripts</span>{}'.format(sort_start_int, sort_end), 
                    '{}<span title="Owner">Owner</span>{}'.format(sort_start_int, sort_end), 
                    '{}<span title="Scope">Scope</span>{}'.format(sort_start_int, sort_end), 
                    '{}<span title="Number of Authoritative Statement authors">Authors</span>{}'.format(sort_start_int, sort_end), 
                    ]

                # Add the manuscript to the related objects
                related_objects.append(collections)

                context['related_objects'] = related_objects

                # =============== Visualizations ====================================================================

                # Use the 'graph' function or not?
                use_network_graph = True

                # THe graph also needs room in after details
                if use_network_graph:
                    context['austat_graph'] = reverse("austat_graph", kwargs={'pk': instance.id})
                    context['austat_trans'] = reverse("austat_trans", kwargs={'pk': instance.id})
                    context['austat_overlap'] = reverse("austat_overlap", kwargs={'pk': instance.id})
                context['austat_pca'] = reverse("austat_pca", kwargs={'pk': instance.id})
                context['manuscripts'] = qs_s.count()
                lHtml = []
                if 'after_details' in context:
                    lHtml.append(context['after_details'])
                if context['object'] == None:
                    context['object'] = instance

                # Note (EK): this must be here, see issue #508
                lHtml.append(render_to_string('seeker/super_graph.html', context, self.request))

                context['after_details'] = "\n".join(lHtml)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("AustatDetails/add_to_context")

        # Return the context we have made
        return context

    def after_new(self, form, instance):
        """Action to be performed after adding a new item"""

        bResult = True
        msg = ""

        # Make sure we change the [atype] from default into 'accepted'
        instance.atype = "acc"
        instance.save()

        return bResult, msg

    def after_save(self, form, instance):
        oErr = ErrHandle()
        bBack = True
        msg = ""
        try:
            if self.isnew:
                # Try default project assignment
                profile = Profile.get_user_profile(self.request.user.username)
                qs = profile.project_editor.filter(status="incl")
                for obj in qs:
                    AustatProject.objects.create(project=obj.project, equal=instance)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("AustatDetails/before_save")
            bBack = False
        return bBack, msg

    def before_save(self, form, instance):
        oErr = ErrHandle()
        bBack = True
        msg = ""
        try:
            self.isnew = False
            if not instance is None:
                if instance.id is None:
                    # This is a new SSG being created.
                    # Provide standard stuff:
                    instance.author = Author.get_undecided()
                    instance.stype = STYPE_MANUAL
                    self.isnew = True
        except:
            msg = oErr.get_error_message()
            oErr.DoError("AustatDetails/before_save")
            bBack = False
        return bBack, msg

    def get_field_value(self, type, instance, custom):
        username = self.request.user.username
        team_group = app_editor
        sBack = ""
        if type == "manu":
            if custom == "shelfmark":
                sBack = "{}, {}, <span class='signature'>{}</span>".format(instance.get_city(), instance.get_library(), instance.idno)
            elif custom == "origin":
                sBack = "{} ({})".format(instance.get_origins(), instance.get_provenance_markdown(table=False))
            elif custom == "daterange":
                sBack = "{}-{}".format(instance.yearstart, instance.yearfinish)
            elif custom == "collections":
                sBack = instance.get_collections_markdown(username, team_group)
            elif custom == "canwit":
                manu = instance.msitem.manu
                sBack = "{}/{}: {}".format(instance.msitem.order, manu.get_canwit_count(), instance.get_lilacode())
            elif custom == "locus":
                sBack = instance.locus
            elif custom == "author":
                sBack = instance.get_author()
            elif custom == "ftext":
                sBack = instance.get_ftext_markdown()
            elif custom == "ftrans":
                sBack = instance.get_ftrans_markdown()
            elif custom == "keywords":
                sBack = instance.get_keywords_markdown()
        elif type == "collection":
            lCombi = []
            if custom == "name":
                sBack = instance.name
            elif custom == "canwits":
                # austat_ids = [x.austat.id for x in Caned.objects.filter(collection=instance)]
                # We should just look at ONE SINGLE austat id!!
                austat_ids = [ instance.id ]
                canwits = [x.canwit for x in CanwitAustat.objects.filter(austat__id__in=austat_ids)]
                html = []
                for obj in canwits:
                    manu = obj.msitem.codico.manuscript
                    locus = obj.locus
                    sCanwit = "{}: {}".format(manu.get_full_name(), locus)
                    url = reverse('canwit_details', kwargs={'pk': obj.id})
                    html.append("<span class='badge signature ot'><a href='{}'>{}</a></span>".format(url, sCanwit))
                sBack = ", ".join(html)
            elif custom == "owner":
                sBack = instance.owner.user.username
            elif custom == "scope":
                sBack = instance.get_scope_display()
            elif custom == "authnum":
                sBack = str(instance.ssgauthornum)
        return sBack

    def process_formset(self, prefix, request, formset):
        return None


class AustatListView(BasicList):
    """List austat instances"""

    model = Austat
    listform = AustatForm
    has_select2 = True  # Check
    use_team_group = True
    template_help = "seeker/filter_help.html"
    prefix = "as"
    bUseFilter = True  
    plural_name = "Authoritative statements"
    sg_name = "Authoritative statement"
    # order_cols = ['code', 'author', 'firstsig', 'srchftext', '', 'scount', 'ssgcount', 'hccount', 'stype']
    order_cols = ['code', 'auwork__key;keycode', 'srchftext', '', 'scount', 'ssgcount', 'hccount', 'stype']
    order_default= order_cols
    order_heads = [
        {'name': 'Author',                  'order': 'o=1', 'type': 'str', 'custom': 'author', 'linkdetails': True},
        #{'name': 'Gryson/Clavis',           'order': 'o=3', 'type': 'str', 'custom': 'sig',    'allowwrap': True, 'options': "abcd",
        # 'title': "The Gryson/Clavis codes of all the Sermons Gold in this equality set"},
        {'name': 'Key',                     'order': 'o=2', 'type': 'str', 'custom': 'keycode'}, # 'linkdetails': True},
        {'name': 'Full text',               'order': 'o=3', 'type': 'str', 'custom': 'ftext',  'main': True, 'linkdetails': True,
         'title': "The full text that has been chosen for this Authoritative statement"},
        {'name': 'HC', 'title': "Historical collections associated with this Authoritative statement", 
         'order': '', 'allowwrap': True, 'type': 'str', 'custom': 'hclist'},
        {'name': 'Sermons',                 'order': 'o=5'   , 'type': 'int', 'custom': 'scount',
         'title': "Number of Sermon (manifestation)s that are connected with this Authoritative statement"},
        {'name': 'Authority',                   'order': 'o=6'   , 'type': 'int', 'custom': 'ssgcount',
         'title': "Number of other Authoritative statements this Authoritative statement links to"},
        {'name': 'HCs',                     'order': 'o=7'   , 'type': 'int', 'custom': 'hccount',
         'title': "Number of historical collections associated with this Authoritative statement"},
        {'name': 'Status',                  'order': 'o=8',   'type': 'str', 'custom': 'status'}        
        ]
    filters = [
        {"name": "Author",          "id": "filter_author",            "enabled": False},
        {"name": "Key code",        "id": "filter_keycode",           "enabled": False},
        {"name": "Work",            "id": "filter_work",              "enabled": False},
        {"name": "Full text",       "id": "filter_ftext",             "enabled": False},
        {"name": "Translation",     "id": "filter_ftrans",            "enabled": False},
        #{"name": "Number",          "id": "filter_number",            "enabled": False},
        {"name": "Genre",           "id": "filter_genre",             "enabled": False},
        {"name": "Keyword",         "id": "filter_keyword",           "enabled": False},
        {"name": "Status",          "id": "filter_stype",             "enabled": False},
        {"name": "CanWit count",    "id": "filter_scount",            "enabled": False},
        {"name": "Relation count",  "id": "filter_ssgcount",          "enabled": False},
        {"name": "Project",         "id": "filter_project",           "enabled": False},        
        {"name": "Collection...",   "id": "filter_collection",        "enabled": False, "head_id": "none"},
        {"name": "Manuscript",      "id": "filter_collmanu",          "enabled": False, "head_id": "filter_collection"},
        {"name": "Canon witness",   "id": "filter_collcanwit",        "enabled": False, "head_id": "filter_collection"},
        {"name": "Authoritative statement",  "id": "filter_collaustat",         "enabled": False, "head_id": "filter_collection"},
        {"name": "Historical",      "id": "filter_collhist",          "enabled": False, "head_id": "filter_collection"},
               ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'ftext',     'dbfield': 'srchftext',         'keyS': 'ftext',    'regex': adapt_regex_incexp},
            {'filter': 'ftrans',    'dbfield': 'srchftrans',        'keyS': 'ftrans',   'regex': adapt_regex_incexp},
            {'filter': 'keycode',   'dbfield': 'keycodefull',       'keyS': 'keycode',  'keyList': 'keycodelist', 'infield': 'id'},
            {'filter': 'work',      'fkfield': 'auwork',            'keyList': 'worklist', 'infield': 'id'},
            #{'filter': 'number',    'dbfield': 'number',            'keyS': 'number',
            # 'title': 'The per-author-canwit-number (these numbers are assigned automatically and have no significance)'},
            {'filter': 'scount',    'dbfield': 'soperator',         'keyS': 'soperator'},
            {'filter': 'scount',    'dbfield': 'scount',            'keyS': 'scount',
             'title': 'The number of canon witnesses belonging to this Authoritative statement'},
            {'filter': 'ssgcount',  'dbfield': 'ssgoperator',       'keyS': 'ssgoperator'},
            {'filter': 'ssgcount',  'dbfield': 'ssgcount',          'keyS': 'ssgcount',
             'title': 'The number of links an Authoritative statement has to other Authoritative statements'},
            {'filter': 'genre',     'fkfield': 'genres',            'keyFk': 'name', 'keyList': 'genrelist', 'infield': 'id'},
            {'filter': 'keyword',   'fkfield': 'keywords',          'keyFk': 'name', 'keyList': 'kwlist', 'infield': 'id'},
            {'filter': 'author',    'fkfield': 'author',            
             'keyS': 'authorname', 'keyFk': 'name', 'keyList': 'authorlist', 'infield': 'id', 'external': 'gold-authorname' },
            {'filter': 'stype',     'dbfield': 'stype',             'keyList': 'stypelist', 'keyType': 'fieldchoice', 'infield': 'abbr' },
            {'filter': 'project',   'fkfield': 'projects',   'keyFk': 'name', 'keyList': 'projlist', 'infield': 'name'}            
            ]},
        {'section': 'collection', 'filterlist': [
            {'filter': 'collmanu',  'fkfield': 'equal_goldsermons__canwit__manu__collections',  
             'keyS': 'collection','keyFk': 'name', 'keyList': 'collist_m', 'infield': 'name' }, 
            {'filter': 'collcanwit', 'fkfield': 'austat_canwits__canwit_col__collection',        
             'keyS': 'collection','keyFk': 'name', 'keyList': 'collist_s', 'infield': 'name' }, 
            {'filter': 'collaustat', 'fkfield': 'collections',                                        
             'keyS': 'collection','keyFk': 'name', 'keyList': 'collist_ssg', 'infield': 'name' }, 
            {'filter': 'collhist', 'fkfield': 'collections',                                        
             'keyS': 'collection','keyFk': 'name', 'keyList': 'collist_hist', 'infield': 'name' }
            ]},
        {'section': 'other', 'filterlist': [
            {'filter': 'atype', 'dbfield': 'atype', 'keyS': 'atype'}
            ]}
        ]
    custombuttons = [{"name": "scount_histogram", "title": "Sermon Histogram", 
                      "icon": "th-list", "template_name": "seeker/scount_histogram.html" }]

    def initializations(self):
        """Initializations to the Austat listview"""

        # ======== One-time adaptations ==============
        listview_adaptations("austat_list")

        self.uploads = []

        return None
    
    def add_to_context(self, context, initial):
        oErr = ErrHandle()
        try:
            # Find out who the user is
            profile = Profile.get_user_profile(self.request.user.username)
            context['basketsize'] = 0 if profile == None else profile.basketsize_austat
            context['basket_show'] = reverse('basket_show_austat')
            context['basket_update'] = reverse('basket_update_austat')
            context['histogram_data'] = self.get_histogram_data('d3')

            # Does this user have upload permissions?
            if context['is_app_uploader']:
                # Yes, user has upload permissions
                html = []
                html.append("Import Authoritative statements from one or more Excel files.")
                msg = "<br />".join(html)
                oExcel = dict(title="authoritative_statements", label="Excel",
                                url=reverse('austat_upload_excel'),
                                type="multiple", msg=msg)
                self.uploads.append(oExcel)

                context['uploads'] = self.uploads

        except:
            msg = oErr.get_error_message()
            oErr.DoError("AustatListview/add_to_context")

        return context

    def get_histogram_data(self, method='d3'):
        """Get data to make a histogram"""

        oErr = ErrHandle()
        histogram_data = []
        b_chart = None
        try:
            # Get the base url
            baseurl = reverse('austat_list')
            # Get the queryset for this view
            qs = self.get_queryset().order_by('scount').values('scount', 'id')
            scount_index = {}
            frequency = None
            for item in qs:
                scount = item['scount']
                if frequency == None or frequency != scount:
                    frequency = scount
                    histogram_data.append(dict(scount=scount, freq=1))
                else:
                    histogram_data[-1]['freq'] += 1

            # Determine the targeturl for each histogram bar
            other_list = []
            for item in self.param_list:
                if "-soperator" not in item and "-scount" not in item:
                    other_list.append(item)
            other_filters = "&".join(other_list)
            for item in histogram_data:
                targeturl = "{}?ssg-soperator=exact&ssg-scount={}".format(baseurl, item['scount'], other_filters)
                item['targeturl'] = targeturl

            if method == "d3":
                histogram_data = json.dumps(histogram_data)
            
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_histogram_data")
        return histogram_data
        
    def get_basketqueryset(self):
        if self.basketview:
            profile = Profile.get_user_profile(self.request.user.username)
            qs = profile.basketitems_austat.all()
        else:
            qs = Austat.objects.all()
        return qs

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []
        oErr = ErrHandle()
        try:
            if custom == "author": 
                # Get a good name for the author
                if instance.author:
                    html.append(instance.author.name)
                else:
                    html.append("<i>(not specified)</i>")
            elif custom == "scount":
                sCount = instance.scount
                if sCount == None: sCount = 0
                html.append("{}".format(sCount))
            elif custom == "ssgcount":
                sCount = instance.ssgcount
                if sCount == None: sCount = 0
                html.append("{}".format(sCount))
            elif custom == "hccount":
                html.append("{}".format(instance.hccount))
            elif custom == "hclist":
                html.append(instance.get_hclist_markdown())
            elif custom == "code":
                sCode = "-" if instance.code  == None else instance.code
                html.append("{}".format(sCode))
            elif custom == "ftext":
                html.append("<span>{}</span>".format(instance.get_ftext_markdown()))
            elif custom == "keycode":
                sKeyCode = "-" if instance.keycode  is None else instance.keycode
                if not instance.auwork is None:
                    # Prepend the AuWork, if needed
                    sAuWorkCode = instance.auwork.key
                    if not sAuWorkCode in sKeyCode:
                        sKeyCode = "{}.{}".format(sAuWorkCode, sKeyCode)
                # We also need the URL of austat details
                url = reverse("austat_details", kwargs={'pk': instance.id})

                maydragdrop = False if self is None else ( user_is_authenticated(self.request) and user_is_ingroup(self.request, app_editor) )

                if self is None or not maydragdrop:
                    html.append("<span class='badge signature'><a class='nostyle' href='{}'>{}</a></span>".format(
                        url, sKeyCode))
                else:
                    # Get the url to do what is needed
                    url_start = reverse('austat_drag_start', kwargs={'pk': instance.id})
                    url_end = reverse('austat_drag_end', kwargs={'pk': instance.id})
                    localcontext = dict(
                        keycode=sKeyCode,
                        url=url,
                        targeturl_start=url_start, 
                        targeturl_end=url_end, 
                        austatid=instance.id, 
                        austatkey=sKeyCode)
                    html.append(render_to_string("seeker/austat_exchange.html", localcontext, self.request))

            elif custom == "status":
                # Provide the status traffic light
                html.append(instance.get_stype_light())

            sBack = "\n".join(html) 
        except:
            msg = oErr.get_error_message()
            oErr.DoError("AustatListView/get_field_value")
        return sBack, sTitle

    def adapt_search(self, fields):
        # Adapt the search to the keywords that *may* be shown
        lstExclude= None
        qAlternative = None

        # Check if a list of keywords is given
        if 'kwlist' in fields and fields['kwlist'] != None and len(fields['kwlist']) > 0:
            # Get the list
            kwlist = fields['kwlist']
            # Get the user
            username = self.request.user.username
            user = User.objects.filter(username=username).first()
            # Check on what kind of user I am
            if not user_is_ingroup(self.request, app_editor):
                # Since I am not an app-editor, I may not filter on keywords that have visibility 'edi'
                kwlist = Keyword.objects.filter(id__in=kwlist).exclude(Q(visibility="edi")).values('id')
                fields['kwlist'] = kwlist

        scount = fields.get('scount', -1)
        soperator = fields.pop('soperator', None)
        if scount != None and scount >= 0 and soperator != None:
            # Action depends on the operator
            fields['scount'] = Q(**{"scount__{}".format(soperator): scount})

        ssgcount = fields.get('ssgcount', -1)
        ssgoperator = fields.pop('ssgoperator', None)
        if ssgcount != None and ssgcount >= 0 and ssgoperator != None:
            # Action depends on the operator
            fields['ssgcount'] = Q(**{"ssgcount__{}".format(ssgoperator): ssgcount})

        # Make sure we only show the SSG/AF's that have accepted modifications
        # (fields['atype'] = 'acc'), so exclude the others:
        lstExclude = [ Q(atype__in=['mod', 'def', 'rej']) ]      
       
        return fields, lstExclude, qAlternative        

    def view_queryset(self, qs):
        search_id = [x['id'] for x in qs.values('id')]
        profile = Profile.get_user_profile(self.request.user.username)
        profile.search_austat = json.dumps(search_id)
        profile.save()
        return None

    def get_helptext(self, name):
        """Use the get_helptext function defined in models.py"""
        return get_helptext(name)
        

class AustatScountDownload(BasicPart):
    MainModel = Austat
    template_name = "seeker/download_status.html"
    action = "download"
    dtype = "csv"       # downloadtype

    def custom_init(self):
        """Calculate stuff"""
        
        dt = self.qd.get('downloadtype', "")
        if dt != None and dt != '':
            self.dtype = dt

    def get_queryset(self, prefix):

        # Construct the QS
        #qs = TestsetUnit.objects.all().order_by('testset__round', 'testset__number').values(
        #    'testset__round', 'testset__number', 'testunit__speaker__name', 'testunit__fname',
        #    'testunit__sentence__name', 'testunit__ntype', 'testunit__speaker__gender')
        qs = Austat.objects.none()
        #TODO: this should be corrected

        return qs

    def get_data(self, prefix, dtype, response=None):
        """Gather the data as CSV, including a header line and comma-separated"""

        # Initialize
        lData = []
        sData = ""

        if dtype == "json":
            # Loop over all round/number combinations (testsets)
            for obj in self.get_queryset(prefix):
                round = obj.get('testset__round')               # obj.testset.round
                number = obj.get('testset__number')             # obj.testset.number
                speaker = obj.get('testunit__speaker__name')    # obj.testunit.speaker.name
                gender = obj.get('testunit__speaker__gender')   # obj.testunit.speaker.gender
                sentence = obj.get('testunit__sentence__name')  # obj.testunit.sentence.name
                ntype = obj.get('testunit__ntype')              # obj.testunit.ntype
                fname = obj.get('testunit__fname')              # Pre-calculated filename
                row = dict(round=round, testset=number, speaker=speaker, gender=gender,
                    filename=fname, sentence=sentence, ntype=ntype)
                lData.append(row)
            # convert to string
            sData = json.dumps(lData, indent=2)
        elif dtype == "csv" or dtype == "xlsx":
            # Create CSV string writer
            output = StringIO()
            delimiter = "\t" if dtype == "csv" else ","
            csvwriter = csv.writer(output, delimiter=delimiter, quotechar='"')
            # Headers
            headers = ['round', 'testset', 'speaker', 'gender', 'filename', 'sentence', 'ntype']
            csvwriter.writerow(headers)
            for obj in self.get_queryset(prefix):
                round = obj.get('testset__round')                # obj.testset.round
                number = obj.get('testset__number')             # obj.testset.number
                speaker = obj.get('testunit__speaker__name')    # obj.testunit.speaker.name
                gender = obj.get('testunit__speaker__gender')   # obj.testunit.speaker.gender
                sentence = obj.get('testunit__sentence__name')  # obj.testunit.sentence.name
                fname = obj.get('testunit__fname')              # Pre-calculated filename
                ntype = obj.get('testunit__ntype')              # obj.testunit.ntype
                row = [round, number, speaker, gender, fname, sentence, ntype]
                csvwriter.writerow(row)

            # Convert to string
            sData = output.getvalue()
            output.close()

        return sData


class AustatDragStart(BasicPart):
    """Start of dragging one particular Austat item"""

    MainModel = Austat

    def custom_init(self):
        """Figure out what is going on"""

        oErr = ErrHandle()
        try:
            # if all is right, the austat should be in [obj]
            austat = self.obj
            # User must be signed in and authenticated for editing
            if user_is_authenticated(self.request) and user_is_ingroup(self.request, app_editor):
                # Figure out who I am
                profile = self.request.user.user_profiles.first()
                if not profile is None:
                    if not austat is None:
                        # First delete any 'stale' objects
                        DraggingAustat.objects.filter(profile=profile).delete()

                        # Check if it already is available in my 'dragged_austat' items
                        if profile.dragged_austats.filter(id=austat.id).first() is None:
                            # Add it to my dragged austats
                            try:
                                DraggingAustat.objects.create(profile=profile, austat=austat)
                                # ========== DEBUGGING =====================
                                oErr.Status("Started dragging profile={} austat={}".format(profile.id, austat.id))
                                # ==========================================
                            except:
                                oErr.Status("Could not create combination profile={} austat={}".format(profile.id, austat.id))
                else:
                    iStop = 1
            else:
                # A non-user or non-editor attempts to change something
                username = "non-user" if self.request.user is None else self.request.user.username
                oErr.Status("User [{}] illegally tried to drag austat {}".format(username, austat.id))

        except:
            msg = oErr.get_error_message()
            oErr.DoError("AustatDragStart/custom_init")

        return None


class AustatDragDrop(BasicPart):
    """Drop the currently 'dragged' Austat item into this Collection"""

    MainModel = Collection

    def custom_init(self):
        """Figure out what is going on"""

        oErr = ErrHandle()
        try:
            # If all is right, the Collection should be in[obj]
            collection = self.obj

            # User must be signed in and authenticated for editing
            if user_is_authenticated(self.request) and user_is_ingroup(self.request, app_editor):
                # Figure out who I am
                profile = self.request.user.user_profiles.first()
                if not profile is None:

                    if collection is None:
                        # Something is wrong here
                        iStop = 2
                    else:
                        # Be sure to re-order the collection, if needed
                        collection.reorder()

                        # Get the url for this collection view, depending on what kind of collection this is
                        if collection.scope == "priv":
                            url = reverse('collpriv_details', kwargs={'pk': collection.id})
                        elif collection.scope == "publ":
                            url = reverse('collpubl_details', kwargs={'pk': collection.id})
                        else:
                            url = reverse('collany_details', kwargs={'pk': collection.id})

                        # And there should be one or more Austat items that have been dragged
                        austats = profile.dragged_austats.all()

                        # Walk all the austats
                        for austat in austats:
                            # Add them into the collection if they are not already there
                            caned = Caned.objects.filter(collection=collection, austat=austat).first()
                            if caned is None:
                                # add the austat to this collection (the [Caned]
                                order = Caned.objects.filter(collection=collection).count() + 1
                                caned = Caned.objects.create(collection=collection, austat=austat, order=order)
                                oErr.Status("Dragging added austat={} to collection={} as caned={}".format(austat.id, collection.id, caned.id))
                                # Make sure that we re-load the Collection view that we are part of
                                self.redirectpage = url
                else:
                    iStop = 1
            else:
                # A non-user or non-editor attempts to change something
                username = "non-user" if self.request.user is None else self.request.user.username
                oErr.Status("User [{}] illegally tried to drop into collection {}".format(username, collection.id))
        except:
            msg = oErr.get_error_message()
            oErr.DoError("AustatDragDrop/custom_init")

        return None


class AustatDragEnd(BasicPart):
    """Stop dragging the currently 'dragged' Austat item"""

    MainModel = Austat

    def custom_init(self):
        """Figure out what is going on"""

        oErr = ErrHandle()
        try:
            # Figure out who I am
            profile = self.request.user.user_profiles.first()
            if not profile is None:
                # if all is right, the austat should be in [obj]
                austat = self.obj
                if not austat is None:
                    # Delete all 'dragged' austats
                    DraggingAustat.objects.filter(profile=profile).delete()
                    # ========== DEBUGGING =====================
                    oErr.Status("Dragging: deleted all austats for profile={}".format(profile.id))
                    # ==========================================
            else:
                iStop = 1
        except:
            msg = oErr.get_error_message()
            oErr.DoError("AustatDragEnd/custom_init")

        return None
    

class AustatVisDownload(BasicPart):
    """Generic treatment of Visualization downloads for SSGs"""

    MainModel = Austat
    template_name = "seeker/download_status.html"
    action = "download"
    dtype = "hist-svg"
    vistype = ""

    def custom_init(self):
        """Calculate stuff"""
        
        dt = self.qd.get('downloadtype', "")
        if dt != None and dt != '':
            self.dtype = dt

    def get_data(self, prefix, dtype, response=None):
        """Gather the data as CSV, including a header line and comma-separated"""

        # Initialize
        lData = []
        sData = ""

        if dtype == "json":
            # Retrieve the actual data from self.data
            oData = dict(legend=self.data['legend'],
                         link_list=self.data['link_list'],
                         node_list=self.data['node_list'])
            sData = json.dumps(oData, indent=2)
        elif dtype == "hist-svg":
            pass
        elif dtype == "hist-png":
            pass
        elif dtype == "csv" or dtype == "xlsx":
            # Create CSV string writer
            output = StringIO()
            delimiter = "\t" if dtype == "csv" else ","
            csvwriter = csv.writer(output, delimiter=delimiter, quotechar='"')
            # Headers
            headers = ['round', 'testset', 'speaker', 'gender', 'filename', 'sentence', 'ntype']
            csvwriter.writerow(headers)
            pass

            # Convert to string
            sData = output.getvalue()
            output.close()

        return sData


class AustatGraphDownload(AustatVisDownload):
    """Network graph"""
    vistype = "graph"


class AustatTransDownload(AustatVisDownload):
    """Transmission graph"""
    vistype = "trans"


class AustatOverlapDownload(AustatVisDownload):
    """Overlap graph"""
    vistype = "overlap"


# ============= COLLECTION VIEWS ============================


class CollAnyEdit(BasicDetails):
    """Manu: Manuscript collections"""

    model = Collection
    mForm = CollectionForm
    prefix = "any"
    basic_name_prefix = "coll"
    rtype = "json"
    settype = "pd"
    title = "Any collection"
    history_button = True
    manu = None
    codico = None
    datasettype = ""
    use_team_group = True
    mainitems = []
    hlistitems = [
        {'type': 'manu',    'clsColl': CollectionMan,   'field': 'manuscript'},
        {'type': 'sermo',   'clsColl': CollectionCanwit,  'field': 'sermon'},
        {'type': 'austat',  'clsColl': Caned, 'field': 'austat'},
        ]

    ClitFormSet = inlineformset_factory(Collection, LitrefCol,
                                         form = CollectionLitrefForm, min_num=0,
                                         fk_name = "collection",
                                         extra=0, can_delete=True, can_order=False)

    formset_objects = []

    stype_edi_fields = ['name', 'owner', 'readonly', 'type', 'settype', 'descrip', 'url', 'path', 'scope', 
                        'LitrefMan', 'litlist', 'projlist']

    def custom_init(self, instance):
        if instance != None and instance.settype == "hc":
            # First check if the 'clit' is already in the formset_objects or not
            bFound = False
            for oItem in self.formset_objects:
                if oItem['prefix'] == "clit":
                    bFound = True
                    break
            if not bFound:
                self.formset_objects.append(
                    {'formsetClass': self.ClitFormSet,  'prefix': 'clit',  
                     'readonly': False, 'noinit': True, 'linkfield': 'collection'})
        if instance != None:
            self.datasettype = instance.type
        return None

    def check_hlist(self, instance):
        """Check if a hlist parameter is given, and hlist saving is called for"""

        oErr = ErrHandle()

        try:
            arg_hlist = instance.type + "-hlist"
            arg_savenew = instance.type + "-savenew"
            if arg_hlist in self.qd and arg_savenew in self.qd:
                # Interpret the list of information that we receive
                hlist = json.loads(self.qd[arg_hlist])
                # Interpret the savenew parameter
                savenew = self.qd[arg_savenew]

                # Make sure we are not saving
                self.do_not_save = True
                # But that we do a new redirect
                self.newRedirect = True

                # Action depends on the particular prefix
                for hlistitem in self.hlistitems:
                    # Is this the one?
                    if hlistitem['type'] == instance.type:
                        # THis is the type
                        clsColl = hlistitem['clsColl']
                        field = hlistitem['field']

                        # First check if this needs to be a *NEW* collection instance
                        if savenew == "true":
                            profile = Profile.get_user_profile(self.request.user.username)
                            # Yes, we need to copy the existing collection to a new one first
                            original = instance
                            instance = original.get_copy(owner=profile)

                        # Change the redirect URL
                        if self.redirectpage == "":
                            this_type = ""
                            if instance.settype == "hc": this_type = "hist"
                            elif instance.scope == "priv": this_type = "priv"
                            else: this_type = "publ"
                            self.redirectpage = reverse('coll{}_details'.format(this_type), kwargs={'pk': instance.id})
                        else:
                            self.redirectpage = self.redirectpage.replace(original.id, instance.id)

                        # What we have is the ordered list of Manuscript id's that are part of this collection
                        with transaction.atomic():
                            # Make sure the orders are correct
                            for idx, item_id in enumerate(hlist):
                                order = idx + 1
                                lstQ = [Q(collection=instance)]
                                lstQ.append(Q(**{"{}__id".format(field): item_id}))
                                obj = clsColl.objects.filter(*lstQ).first()
                                if obj != None:
                                    if obj.order != order:
                                        obj.order = order
                                        obj.save()
                        # See if any need to be removed
                        existing_item_id = [str(getattr(x, field).id) for x in clsColl.objects.filter(collection=instance)]
                        delete_id = []
                        for item_id in existing_item_id:
                            if not item_id in hlist:
                                delete_id.append(item_id)
                        if len(delete_id)>0:
                            lstQ = [Q(collection=instance)]
                            lstQ.append(Q(**{"{}__id__in".format(field): delete_id}))
                            clsColl.objects.filter(*lstQ).delete()

            return True
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollAnyEdit/check_hlist")
            return False
    
    def add_to_context(self, context, instance):
        """Add to the existing context"""

        prefix_scope = ['any', 'manu', 'sermo', 'gold', 'austat', 'priv', 'publ']
        prefix_type = ['any', 'manu', 'sermo', 'gold', 'austat', 'priv', 'publ']
        prefix_readonly = ['any', 'manu', 'sermo', 'gold', 'austat']
        prefix_elevate = ['any', 'austat', 'priv', 'publ']

        oErr = ErrHandle()
        try:
            # One general item is the 'help-popup' to be shown when the user clicks on 'Author'
            info = render_to_string('seeker/author_info.html')

            # Need to know who this is
            profile = Profile.get_user_profile(self.request.user.username)

            # Define the main items to show and edit
            context['mainitems'] = [
                {'type': 'plain', 'label': "Name:",        'value': instance.name,      'field_key': 'name'},
                {'type': 'safe',  'label': "Description:", 'value': instance.get_descr(), 'field_key': 'descrip'},
                {'type': 'plain', 'label': "URL:",         'value': instance.url,       'field_key': 'url'}, 
                {'type': 'plain', 'label': "LiLaC code:",  'value': instance.lilacode,  'field_key': 'lilacode'}, 
                {'type': 'plain', 'label': "Date:",        'value': instance.date,      'field_key': 'date'}, 
                {'type': 'plain', 'label': "Origin:",      'value': instance.origin,    'field_key': 'origin'}, 
                {'type': 'plain', 'label': "Author:",      'value': instance.author_help(info), 'field_key': 'newauthor'},
                ]

            # Optionally add Scope: but only for the actual *owner* of this one
            if self.prefix in prefix_scope and instance.owner.user == self.request.user:
                context['mainitems'].append(
                {'type': 'plain', 'label': "Scope:",       'value': instance.get_scope_display, 'field_key': 'scope'})

            # Always add Type, but its value may not be changed
            context['mainitems'].append(
                {'type': 'plain', 'label': "Type:",        'value': instance.get_type_display})

            # Optionally add Readonly
            if self.prefix in prefix_readonly:
                context['mainitems'].append(
                {'type': 'plain', 'label': "Readonly:",    'value': instance.readonly, 'field_key': 'readonly'})

            bMayEdit = user_is_ingroup(self.request, app_editor)

            # This is only for private PDs:
            if self.prefix == "priv" and instance != None and instance.settype == "pd" and instance.id != None:
                name_choice = dict(
                    manu=dict(sg_name="Manuscript", pl_name="Manuscripts"),
                    canwit=dict(sg_name="Canon witness", pl_name="Canon witnesses"),
                    austat=dict(sg_name="Authoritative statement", pl_name="Authoritative statements")
                    )
                # Add a button + text
                context['datasettype'] = instance.type
                context['sg_name'] = name_choice[instance.type]['sg_name']
                context['pl_name'] = name_choice[instance.type]['pl_name']
            
                context['size'] = instance.get_size_markdown()

                # Any target url for dropping an austat
                context['targeturl'] = reverse('austat_drag_drop', kwargs={'pk': instance.id})
                size_value = render_to_string("seeker/collpriv.html", context, self.request)
            else:
                if instance.settype == "hc" and bMayEdit:
                    context['size'] = instance.get_size_markdown()

                    # Any target url for dropping an austat
                    context['targeturl'] = reverse('austat_drag_drop', kwargs={'pk': instance.id})
                    size_value = render_to_string("seeker/collsize.html", context, self.request)
                else:
                    size_value = instance.get_size_markdown()
        
            # Always add Created and Size
            context['mainitems'].append( {'type': 'plain', 'label': "Created:",     'value': instance.get_created})
            context['mainitems'].append( {'type': 'line',  'label': "Size:",        'value': size_value})

            # If this is a historical collection,and an app-editor gets here, add a link to a button to create a manuscript
            if instance.settype == "hc" and context['is_app_editor']:
                # If 'manu' is set, then this procedure is called from 'collhist_compare'
                if self.manu == None:
                    context['mainitems'].append({'type': 'safe', 'label': 'Manuscript', 'value': instance.get_manuscript_link()})
            # Historical collections may have literature references
            if instance.settype == "hc" and len(self.formset_objects) > 0:
                oLitref = {'type': 'plain', 'label': "Literature:",   'value': instance.get_litrefs_markdown() }
                if context['is_app_editor']:
                    oLitref['multiple'] = True
                    oLitref['field_list'] = 'litlist'
                    oLitref['fso'] = self.formset_objects[0]
                    oLitref['template_selection'] = 'ru.solemne.litref_template'
                context['mainitems'].append(oLitref)        
        
            # Historical collections have a project assigned to them
            if instance.settype == "hc":
                oProject =  {'type': 'plain', 'label': "Project:",     'value': instance.get_project_markdown2()}
                if may_edit_project(self.request, profile, instance):
                    oProject['field_list'] = 'projlist'
                context['mainitems'].append(oProject)        
                        

            # Any dataset may optionally be elevated to a historical collection
            # BUT: only if a person has permission
            if instance.settype == "pd" and self.prefix in prefix_elevate and instance.type in prefix_elevate and \
                context['authenticated'] and context['is_app_editor']:
                context['mainitems'].append(
                    {'type': 'safe', 'label': "Historical:", 'value': instance.get_elevate()}
                    )
            # Buttons to switch to a listview of M/S/SG/SSG based on this collection
            context['mainitems'].append(
                    {'type': 'safe', 'label': "Listviews:", 'value': self.get_listview_buttons(instance),
                     'title': 'Open a listview that is filtered on this dataset'}
                    )
            # For HC: buttons to switch between related listviews
            if instance.settype == "hc" and context['is_app_editor'] and self.manu == None and self.codico == None:
                context['mainitems'].append(
                        {'type': 'safe', 'label': "Show/hide:", 'value': self.get_hc_buttons(instance),
                         'title': 'Optionally show and edit the Authoritative statements in this collection'}
                        )


            # Signal that we have select2
            context['has_select2'] = True

            # Determine what the permission level is of this collection for the current user
            # (1) Is this user a different one than the one who created the collection?
            profile_owner = instance.owner
            profile_user = Profile.get_user_profile(self.request.user.username)
            # (2) Set default permission
            permission = ""
            if profile_owner.id == profile_user.id:
                # (3) Any creator of the collection may write it
                permission = "write"
            else:
                # (4) permission for different users
                if context['is_app_editor']:
                    # (5) what if the user is an app_editor?
                    if instance.scope == "publ":
                        # Editors may read/write collections with 'public' scope
                        permission = "write"
                    elif instance.scope == "team":
                        # Editors may read collections with 'team' scope
                        permission = "read"
                else:
                    # (5) any other users
                    if instance.scope == "publ":
                        # All users may read collections with 'public' scope
                        permission = "read"

            context['permission'] = permission
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollAnyEdit/add_to_context")
          
        # Return the context we have made
        return context    

    def get_hc_buttons(self, instance):
        """Get buttons to show/hide the two HC listviews (SSGs and codicological unites)"""

        sBack = ""
        lHtml = []
        abbr = None
        button_list = [
            ]
        oErr = ErrHandle()
        try:
            # Button to show AuStat
            count_austat = instance.collections_austat.count()
            button_list.append(dict(label='Authoritative statements ({})'.format(count_austat),
                                    id='basic_austat_set',
                                    show=True))
            # Button to show Codico
            count_codico = instance.get_count_codico()
            button_list.append(dict(label='Codicological units ({})'.format(count_codico),
                                    id='basic_codi_set',
                                    show=False))

            for oButton in button_list:
                lHtml.append("<a role='button' class='btn btn-xs jumbo-1' data-toggle='collapse' data-target='#{}' >{}</a>".format(
                    oButton['id'], oButton['label']))
            sBack = "<span>&nbsp;</span>".join(lHtml)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollAnyEdit/get_hc_buttons")

        return sBack

    def get_listview_buttons(self, instance):
        """Create an HTML list of buttons for M/S/SG/SSG listviews filtered on this collection"""

        sBack = ""
        context = {}
        abbr = None
        oErr = ErrHandle()
        try:
            url_m = reverse("manuscript_list")
            url_s = reverse("canwit_list")
            url_ssg = reverse("austat_list")
            if instance.type == "manu":
                # collection of manuscripts
                abbr = "m"
            elif instance.type == "canwit":
                # collection of canwits
                abbr = "s"
            elif instance.type == "austat":
                # collection of SSG
                abbr = "ssg"
                if self.settype == "hc": abbr = "hist"
            if url_m != None and url_s != None and url_ssg != None and abbr != None:
                context['url_manu'] = "{}?manu-collist_{}={}".format(url_m, abbr, instance.id)
                context['url_canwit'] = "{}?canwit-collist_{}={}".format(url_s, abbr, instance.id)
                context['url_austat'] = "{}?au-collist_{}={}".format(url_ssg, abbr, instance.id) # TH zit het hier?

                sBack = render_to_string('seeker/coll_buttons.html', context, self.request)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollAnyEdit/get_listview_buttons")

        return sBack
    
    def process_formset(self, prefix, request, formset):
        errors = []
        bResult = True
        instance = formset.instance
        for form in formset:
            if form.is_valid():
                cleaned = form.cleaned_data
                # Action depends on prefix

                if prefix == "clit":
                    # Literature reference processing
                    newpages = cleaned.get("newpages")
                    # Also get the litref
                    oneref = cleaned.get("oneref")
                    if oneref:
                        litref = cleaned['oneref']
                        # Check if all is in order
                        if not litref is None:
                            # Check that this link is not there already
                            obj = LitrefCol.objects.filter(reference=oneref, collection=instance, pages=newpages).first()

                            if obj is None:
                                # Continue
                                form.instance.reference = litref
                                if newpages:
                                    form.instance.pages = newpages
                    # Note: it will get saved with form.save()

            else:
                errors.append(form.errors)
                bResult = False
        return None

    def before_save(self, form, instance):
        oErr = ErrHandle()
        bBack = True
        msg = ""
        transfer_changes = [
            # {'src': 'newftext',    'dst': 'ftext',    'type': 'text'},
            # {'src': 'newftrans',   'dst': 'ftrans',   'type': 'text'},
            {'src': 'newauthor',   'dst': 'author',   'type': 'fk'},
            ]
        try:
            if form != None and instance != None:
                # Search the user profile
                profile = Profile.get_user_profile(self.request.user.username)
                form.instance.owner = profile

                # Get the cleaned data: this is the new stuff
                cleaned_data = form.cleaned_data

                # This means that any changes may be implemented right away
                for oTransfer in transfer_changes:
                    type = oTransfer.get("type")
                    src_field = oTransfer.get("src")
                    dst_field = oTransfer.get("dst")
                    src_value = cleaned_data.get(src_field)
                        
                    # Transfer the value
                    if type == "fk" or type == "text":
                        # Is there any change?
                        prev_value = getattr(instance, dst_field)
                        if src_value != prev_value:
                            # Special cases
                            if dst_field == "author" and not instance.author is None:
                                authornameLC = instance.author.name.lower()
                                # Determine what to do in terms of 'moved'.
                                if authornameLC != "undecided":
                                    # Create a copy of the object I used to be
                                    moved = Austat.create_moved(instance)

                            # Perform the actual change
                            setattr(form.instance, dst_field, src_value)

                # Check for author
                if instance.author == None:
                    # Set to "undecided" author if possible
                    author = Author.get_undecided()
                    instance.author = author


                # The collection type is now a parameter
                type = form.cleaned_data.get("type", "")
                if type == "":
                    if self.prefix == "hist":
                        form.instance.type = "austat"
                    elif self.prefix == "publ":
                        form.instance.type = self.datasettype
                    elif self.prefix == "priv":
                        type = self.qd.get("datasettype", "")
                        if type == "": type = self.datasettype
                        if type == "": type = "austat"
                        form.instance.type = type
                    else:
                        form.instance.type = self.prefix

                # Check out the name, if this is not in use elsewhere
                if instance.id != None:
                    name = form.instance.name
                    if Collection.objects.filter(name__iexact=name).exclude(id=instance.id).exists():
                        # The name is already in use, so refuse it.
                        msg = "The name '{}' is already in use for a dataset. Please chose a different one".format(name)
                        return False, msg
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollAnyEdit/before_save")
            bBack = False
        return bBack, msg

    def after_save(self, form, instance):
        msg = ""
        bResult = True
        oErr = ErrHandle()
        
        try:
            # Process many-to-many changes: Add and remove relations in accordance with the new set passed on by the user
            # (1) 'literature'
            litlist = form.cleaned_data['litlist']
            adapt_m2m(LitrefCol, instance, "collection", litlist, "reference", extra=['pages'], related_is_through = True)
            
            # (2) 'projects'
            projlist = form.cleaned_data['projlist']
            col_proj_deleted = []
            adapt_m2m(CollectionProject, instance, "collection", projlist, "project", deleted=col_proj_deleted)
            project_dependant_delete(self.request, col_proj_deleted)

        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        solemne_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return solemne_get_history(instance)

    def get_histogram_data(self, instance=None, qs=None, listview="collist_hist", method='d3'):
        """Get data to make a histogram"""

        oErr = ErrHandle()
        histogram_data = []
        try:
            # Get the queryset for this view
            if instance != None and qs != None:
                # Get the base url
                baseurl = reverse('austat_list')
                # Determine the list
                qs = qs.order_by('scount').values('scount', 'id')
                scount_index = {}
                frequency = None
                for item in qs:
                    scount = item['scount']
                    if frequency == None or frequency != scount:
                        # Initialize the frequency
                        frequency = scount
                        # Add to the histogram data
                        histogram_data.append(dict(scount=scount, freq=1))
                    else:
                        histogram_data[-1]['freq'] += 1

                # Determine the targeturl for each histogram bar
                for item in histogram_data:
                    targeturl = "{}?ssg-{}={}&ssg-soperator=exact&ssg-scount={}".format(baseurl, listview, instance.id, item['scount'])
                    item['targeturl'] = targeturl
                # D3-specific
                if method == "d3":
                    histogram_data = json.dumps(histogram_data)
            
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_histogram_data")
        return histogram_data


class CollPrivEdit(CollAnyEdit):
    prefix = "priv"
    basic_name = "collpriv"
    title = "My Dataset"


class CollPublEdit(CollAnyEdit):
    prefix = "publ"
    basic_name = "collpubl"
    title = "Public Dataset"


class CollHistEdit(CollAnyEdit):
    prefix = "austat"
    settype = "hc"
    basic_name = "collhist"
    title = "Historical collection"

    def before_save(self, form, instance):
        # Make sure the [CollAnyEdit] is executed
        response = super(CollHistEdit, self).before_save(form, instance)

        # Now do the remainder
        oErr = ErrHandle()
        bBack = True
        msg = ""
        try:
            if not instance is None and not instance.id is None:
                # Need to know who is 'talking'...
                username = self.request.user.username
                profile = Profile.get_user_profile(username)

                # Issue #473: automatic assignment of project for particular editor(s)
                projlist = form.cleaned_data.get("projlist")
                bBack, msg = evaluate_projlist(profile, instance, projlist, "Historical collection")
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollHistEdit/before_save")
            bBack = False
        return bBack, msg

    def after_save(self, form, instance):
        # Make sure the [CollAnyEdit] is executed
        response = super(CollHistEdit, self).after_save(form, instance)

        # Now do the remainder
        msg = ""
        bResult = True
        oErr = ErrHandle()
        
        try:
            # Issue #473: default project assignment
            if instance.projects.count() == 0:
                # Need to know who is 'talking'...
                username = self.request.user.username
                profile = Profile.get_user_profile(username)

                # The user has not selected a project (yet): try default assignment
                user_projects = profile.projects.all()
                if user_projects.count() == 1:
                    project = profile.projects.first()
                    CollectionProject.objects.create(collection=instance, project=project)
        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg


class CollSuperEdit(CollAnyEdit):
    """Super: Austat collections = super sermon gold """

    prefix = "austat"
    title = "Authoritative Statement collection"


class CollAnyDetails(CollAnyEdit):
    """Like CollAnyEdit, but then with html"""
    rtype = "html"


class CollPrivDetails(CollAnyEdit):
    """Like CollPrivEdit, but then with html"""

    prefix = "priv"
    basic_name = "collpriv"
    title = "My Dataset"
    rtype = "html"
    custombuttons = []

    def custom_init(self, instance):
        if instance != None:
            # Check if someone acts as if this is a public dataset, whil it is not
            if instance.settype == "pd":
                # Determine what kind of dataset/collection this is
                if instance.owner != Profile.get_user_profile(self.request.user.username):
                    # It is a public dataset after all!
                    self.redirectpage = reverse("collpubl_details", kwargs={'pk': instance.id})
            elif instance.settype == "hc":
                # This is a historical collection
                self.redirectpage = reverse("collhist_details", kwargs={'pk': instance.id})

            if instance.type == "austat":
                self.custombuttons = [{"name": "scount_histogram", "title": "Sermon Histogram", 
                      "icon": "th-list", "template_name": "seeker/scount_histogram.html" }]

            # Check for hlist saving
            self.check_hlist(instance)
        return None

    def add_to_context(self, context, instance):
        # Perform the standard initializations:
        context = super(CollPrivDetails, self).add_to_context(context, instance)

        def add_one_item(rel_item, value, resizable=False, title=None, align=None, link=None, main=None, draggable=None):
            oAdd = dict(value=value)
            if resizable: oAdd['initial'] = 'small'
            if title != None: oAdd['title'] = title
            if align != None: oAdd['align'] = align
            if link != None: oAdd['link'] = link
            if main != None: oAdd['main'] = main
            if draggable != None: oAdd['draggable'] = draggable
            rel_item.append(oAdd)
            return True

        def check_order(qs):
            with transaction.atomic():
                for idx, obj in enumerate(qs):
                    if obj.order < 0:
                        obj.order = idx + 1
                        obj.save()

        oErr = ErrHandle()
        related_objects = []
        lstQ = []
        rel_list =[]
        resizable = True
        index = 1
        sort_start = ""
        sort_start_mix = ""
        sort_start_int = ""
        sort_end = ""

        try:
            username = self.request.user.username
            team_group = app_editor

            # Authorization: only app-editors may edit!
            bMayEdit = user_is_ingroup(self.request, team_group)
            
            # All PDs: show the content
            if bMayEdit:
                sort_start = '<span class="sortable"><span class="fa fa-sort sortshow"></span>&nbsp;'
                sort_start_mix = '<span class="sortable mixed"><span class="fa fa-sort sortshow"></span>&nbsp;'
                sort_start_int = '<span class="sortable integer"><span class="fa fa-sort sortshow"></span>&nbsp;'
                sort_end = '</span>'


            # Action depends on instance.type: M/S/SG/SSG
            if instance.type == "manu":
                # Get all non-template manuscripts part of this PD
                manuscripts = dict(title="Manuscripts within this dataset", prefix="manu")
                if resizable: manuscripts['gridclass'] = "resizable dragdrop"
                manuscripts['savebuttons'] = True
                manuscripts['saveasbutton'] = True

                # Check ordering
                qs_manu = instance.manuscript_col.all().order_by(
                        'order', 'manuscript__lcity__name', 'manuscript__library__name', 'manuscript__idno')
                check_order(qs_manu)

                for obj in qs_manu:
                    rel_item = []
                    item = obj.manuscript

                    # S: Order in Manuscript
                    #add_one_item(rel_item, index, False, align="right", draggable=True)
                    #index += 1
                    add_one_item(rel_item, obj.order, False, align="right", draggable=True)

                    # Shelfmark = IDNO
                    add_one_item(rel_item,  self.get_field_value("manu", item, "shelfmark"), False, title=item.idno, main=True, 
                                 link=reverse('manuscript_details', kwargs={'pk': item.id}))

                    # Just the name of the manuscript
                    add_one_item(rel_item, self.get_field_value("manu", item, "name"), resizable)

                    # Origin
                    add_one_item(rel_item, self.get_field_value("manu", item, "orgprov"), False, 
                                 title="Origin (if known), followed by provenances (between brackets)")

                    # date range
                    add_one_item(rel_item, self.get_field_value("manu", item, "daterange"), False, align="right")

                    # Number of sermons in this manuscript
                    add_one_item(rel_item, self.get_field_value("manu", item, "sermons"), False, align="right")

                    if bMayEdit:
                        # Actions that can be performed on this item
                        add_one_item(rel_item, self.get_actions())

                    # Add this line to the list
                    rel_list.append(dict(id=item.id, cols=rel_item))

                manuscripts['rel_list'] = rel_list

                manuscripts['columns'] = [
                    '{}<span title="Default order">Order<span>{}'.format(sort_start_int, sort_end),
                    '{}<span title="City/Library/Shelfmark">Shelfmark</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Name">Name</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Origin/Provenance">or./prov.</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Date range">date</span>{}'.format(sort_start_int, sort_end), 
                    '{}<span title="Sermons in this manuscript">sermons</span>{}'.format(sort_start_int, sort_end)
                    ]
                if bMayEdit:
                    manuscripts['columns'].append("")
                related_objects.append(manuscripts)

            elif instance.type == "sermo":
                # Get all sermons that are part of this PD
                sermons = dict(title="Sermon manifestations within this dataset", prefix="sermo")
                if resizable: sermons['gridclass'] = "resizable dragdrop"
                sermons['savebuttons'] = True
                sermons['saveasbutton'] = True

                qs_sermo = instance.canwit_col.all().order_by(
                        'order', 'sermon__author__name', 'sermon__siglist', 'sermon__srchftext', 'sermon__srchftrans')
                check_order(qs_sermo)

                # Walk these collection sermons
                for obj in qs_sermo:
                    rel_item = []
                    item = obj.sermon

                    # S: Order in Sermon
                    #add_one_item(rel_item, index, False, align="right")
                    #index += 1
                    add_one_item(rel_item, obj.order, False, align="right", draggable=True)

                    # S: Author
                    add_one_item(rel_item, self.get_field_value("sermo", item, "author"), False, main=True)

                    # S: Signature
                    add_one_item(rel_item, self.get_field_value("sermo", item, "signature"), False)

                    # S: Inc+Expl
                    add_one_item(rel_item, self.get_field_value("sermo", item, "incexpl"), resizable)

                    # S: Manuscript
                    add_one_item(rel_item, self.get_field_value("sermo", item, "manuscript"), False)

                    # S: Locus
                    add_one_item(rel_item, item.locus, False)

                    if bMayEdit:
                        # Actions that can be performed on this item
                        add_one_item(rel_item, self.get_actions())

                    # Add this line to the list
                    rel_list.append(dict(id=item.id, cols=rel_item))
            
                sermons['rel_list'] = rel_list
                sermons['columns'] = [
                    '{}<span title="Default order">Order<span>{}'.format(sort_start_int, sort_end),
                    '{}<span title="Attributed author">Author</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Gryson or Clavis code">Signature</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Incipit and explicit">inc...expl</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Manuscript shelfmark">Manuscript</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Location within the manuscript">Locus</span>{}'.format(sort_start_int, sort_end)
                    ]
                if bMayEdit: sermons['columns'].append("")
                related_objects.append(sermons)

            elif instance.type == "austat":
                # Get all sermons that are part of this PD
                supers = dict(title="Authoritative statements within this dataset", prefix="austat")   #prefix="sermo")
                if resizable: supers['gridclass'] = "resizable dragdrop"
                supers['savebuttons'] = True
                supers['saveasbutton'] = True

                qs_sermo = instance.austat_col.all().order_by(
                        'order', 'austat__author__name', 'austat__srchftext', 'austat__srchftrans')
                check_order(qs_sermo)

                # Walk these collection sermons
                for obj in qs_sermo:
                    rel_item = []
                    item = obj.austat

                    # SSG: Order in Manuscript
                    add_one_item(rel_item, obj.order, False, align="right", draggable=True)

                    # SSG: Author
                    add_one_item(rel_item, self.get_field_value("austat", item, "author"), False, main=True)

                    # SSG: solemne code
                    add_one_item(rel_item, self.get_field_value("austat", item, "keycode"), False)

                    # SSG: Full text
                    add_one_item(rel_item, self.get_field_value("austat", item, "ftext"), resizable)

                    # SSG: Size (number of SG in equality set)
                    add_one_item(rel_item, self.get_field_value("austat", item, "scount"), False)

                    if bMayEdit:
                        # Actions that can be performed on this item
                        add_one_item(rel_item, self.get_actions())

                    # Add this line to the list
                    rel_list.append(dict(id=item.id, cols=rel_item))
            
                supers['rel_list'] = rel_list
                supers['columns'] = [
                    '{}<span title="Default order">Order</span>{}'.format(sort_start_int, sort_end),
                    '{}<span title="Author">Author</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Key code">lila</span>{}'.format(sort_start_mix, sort_end), 
                    '{}<span title="Full text">ftext</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Number of Sermons Gold part of this set">Size</span>{}'.format(sort_start_int, sort_end)
                    ]
                if bMayEdit: supers['columns'].append("")
                related_objects.append(supers)

                context['histogram_data'] = self.get_histogram_data(instance, 
                                                                    instance.collections_austat.all(), 
                                                                    'collist_{}'.format(self.prefix), 
                                                                    'd3')

            context['related_objects'] = related_objects
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollPrivDetails/add_to_context")

        # REturn the total context
        return context

    def get_actions(self):
        html = []
        buttons = ['remove']    # This contains all the button names that need to be added

        # Start the whole spane
        html.append("<div class='blinded'>")
        
        # Add components
        if 'up' in buttons: 
            html.append("<a class='related-up' ><span class='glyphicon glyphicon-arrow-up'></span></a>")
        if 'down' in buttons: 
            html.append("<a class='related-down'><span class='glyphicon glyphicon-arrow-down'></span></a>")
        if 'remove' in buttons: 
            html.append("<a class='related-remove'><span class='glyphicon glyphicon-remove'></span></a>")

        # Finish up the span
        html.append("&nbsp;</span>")

        # COmbine the list into a string
        sHtml = "\n".join(html)
        # Return out HTML string
        return sHtml

    def get_field_value(self, type, instance, custom):
        sBack = ""
        if type == "manu":
            if custom == "shelfmark":
                sBack = "{}, {}, <span class='signature'>{}</span>".format(instance.get_city(), instance.get_library(), instance.idno)
            elif custom == "name":
                sBack = instance.name
            elif custom == "origprov":
                sBack = "{} ({})".format(instance.get_origins(), instance.get_provenance_markdown())
            elif custom == "daterange":
                sBack = "{}-{}".format(instance.yearstart, instance.yearfinish)
            elif custom == "sermons":
                sBack = instance.get_canwit_count()
        elif type == "canwit":
            sBack, sTitle = CanwitListView.get_field_value(None, instance, custom)
        elif type == "austat":
            sBack, sTitle = AustatListView.get_field_value(None, instance, custom)
        return sBack


class CollPublDetails(CollPrivDetails):
    """Like CollPrivDetails, but then with public"""

    prefix = "publ"
    basic_name = "collpubl"
    title = "Public Dataset"

    def custom_init(self, instance):
        if instance != None:
            # Check if someone acts as if this is a public dataset, whil it is not
            if instance.settype == "pd":
                # Determine what kind of dataset/collection this is
                if instance.owner == Profile.get_user_profile(self.request.user.username):
                    # It is a private dataset after all!
                    self.redirectpage = reverse("collpriv_details", kwargs={'pk': instance.id})
            elif instance.settype == "hc":
                # This is a historical collection
                self.redirectpage = reverse("collhist_details", kwargs={'pk': instance.id})
            if instance.type == "austat":
                self.custombuttons = [{"name": "scount_histogram", "title": "Sermon Histogram", 
                      "icon": "th-list", "template_name": "seeker/scount_histogram.html" }]
            # Check for hlist saving
            self.check_hlist(instance)
        return None


class CollHistDetails(CollHistEdit):
    """Like CollHistEdit, but then with html"""
    rtype = "html"
    custombuttons = [{"name": "scount_histogram", "title": "Sermon Histogram", 
                      "icon": "th-list", "template_name": "seeker/scount_histogram.html" }]

    def custom_init(self, instance):
        # First do the original custom init
        response = super(CollHistDetails, self).custom_init(instance)
        if not instance is None:
            # Now continue
            if instance.settype != "hc":
                # Someone does as if this is a historical collection...
                # Determine what kind of dataset/collection this is
                if instance.owner == Profile.get_user_profile(self.request.user.username):
                    # Private dataset
                    self.redirectpage = reverse("collpriv_details", kwargs={'pk': instance.id})
                else:
                    # Public dataset
                    self.redirectpage = reverse("collpubl_details", kwargs={'pk': instance.id})

            # Check for hlist saving
            self.check_hlist(instance)
        return None

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Start by executing the standard handling
        super(CollHistDetails, self).add_to_context(context, instance)

        def add_one_item(rel_item, value, resizable=False, title=None, align=None, link=None, main=None, draggable=None):
            oAdd = dict(value=value)
            if resizable: oAdd['initial'] = 'small'
            if title != None: oAdd['title'] = title
            if align != None: oAdd['align'] = align
            if link != None: oAdd['link'] = link
            if main != None: oAdd['main'] = main
            if draggable != None: oAdd['draggable'] = draggable
            rel_item.append(oAdd)
            return True

        def check_order(qs):
            with transaction.atomic():
                for idx, obj in enumerate(qs):
                    if obj.order < 0:
                        obj.order = idx + 1
                        obj.save()

        oErr = ErrHandle()
        try:

            context['sections'] = []

            username = self.request.user.username
            team_group = app_editor

            # Authorization: only app-editors may edit!
            bMayEdit = user_is_ingroup(self.request, team_group)
            
            # Lists of related objects and other initializations
            related_objects = []
            resizable = True
            lstQ = []
            rel_list =[]
            index = 1
            sort_start = ""
            sort_start_int = ""
            sort_start_mix = ""
            sort_end = ""
            show_codico = True  # See issue #363
            show_manu = False   # See issue #363

            if bMayEdit:
                sort_start = '<span class="sortable"><span class="fa fa-sort sortshow"></span>&nbsp;'
                sort_start_int = '<span class="sortable integer"><span class="fa fa-sort sortshow"></span>&nbsp;'
                sort_start_mix = '<span class="sortable mixed"><span class="fa fa-sort sortshow"></span>&nbsp;'
                sort_end = '</span>'

            # In all cases: Get all the SSGs that are part of this historical collection:
            qs_ssg = instance.collections_austat.all().values("id")

            # Check what kind of comparison we need to make
            if self.manu == None and self.codico == None:
                # This is the plain historical collection details view

                # Get all Austats that are part of this PD/HC
                supers = dict(title="Canon editions within this historical collection", prefix="austat")   #prefix="sermo")
                if resizable: supers['gridclass'] = "resizable dragdrop"
                supers['savebuttons'] = True
                supers['saveasbutton'] = True
                supers['classes'] = 'collapse in'

                qs_sermo = instance.austat_col.all().order_by(
                        'order', 'austat__author__name', 'austat__srchftext', 'austat__srchftrans')
                check_order(qs_sermo)

                # Walk these collection austats
                for obj in qs_sermo:
                    rel_item = []
                    item = obj.austat

                    # SSG: Order in Manuscript
                    add_one_item(rel_item, obj.order, False, align="right", draggable=True)

                    # SSG: Author
                    add_one_item(rel_item, self.get_field_value("austat", item, "author"), resizable)

                    # CanEd: key code
                    add_one_item(rel_item, self.get_field_value("collection", obj, "caned"), False)

                    # Austat key code
                    add_one_item(rel_item, self.get_field_value("austat", item, "keycode"), False)

                    # SSG: Inc/Expl
                    add_one_item(rel_item, self.get_field_value("austat", item, "ftext"), False, main=True)

                    # SSG: Size (number of SG in equality set)
                    add_one_item(rel_item, self.get_field_value("austat", item, "scount"), False)

                    # Actions that can be performed on this item
                    if bMayEdit:
                        add_one_item(rel_item, self.get_actions())

                    # Add this line to the list
                    rel_list.append(dict(id=item.id, cols=rel_item))
            
                supers['rel_list'] = rel_list
                supers['columns'] = [
                    '{}<span title="Default order">Order<span>{}'.format(sort_start_int, sort_end),
                    '{}<span title="Author">Author</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="CanEd code">CanEd</span>{}'.format(sort_start_mix, sort_end), 
                    '{}<span title="Austat code">Austat</span>{}'.format(sort_start_mix, sort_end), 
                    '{}<span title="Full text">ftext</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Number of Canon witnesses part of this set">Size</span>{}'.format(sort_start_int, sort_end), 
                    ''
                    ]
                related_objects.append(supers)

                if show_codico:
                    # NEW (see issue #363): List of codicological unites contained in this collection
                    sTitle = "codicological unites with sermons connected to this collection"
                    codicos = dict(title=sTitle, prefix="codi")
                    if resizable:
                        codicos['gridclass'] = "resizable"
                    codicos['classes'] = 'collapse'

                    # Get the codicos linked to these SSGs (in this historical collection)
                    lstQ = []
                    lstQ.append(Q(codicoitems__itemsermons__austats__collections=instance))
                    lstQ.append(Q(manuscript__mtype="man"))
                    qs_codi = Codico.objects.filter(*lstQ).order_by(
                        'id').distinct().order_by('manuscript__lcity__name', 'manuscript__library__name', 'manuscript__idno')

                    rel_list =[]
                    for item in qs_codi:
                        rel_item = []

                        # Shelfmark = IDNO
                        add_one_item(rel_item,  self.get_field_value("codicos", item, "manuscript"), False, title=item.manuscript.idno, main=True, 
                                     link=reverse('manuscript_details', kwargs={'pk': item.manuscript.id}))

                        # Origin
                        add_one_item(rel_item, self.get_field_value("codicos", item, "origprov"), resizable, 
                                     title="Origin (if known), followed by provenances (between brackets)")

                        # date range
                        add_one_item(rel_item, self.get_field_value("codicos", item, "daterange"), resizable, align="right")

                        # Number of sermons in this codicological unit
                        add_one_item(rel_item, self.get_field_value("codicos", item, "sermons"), resizable, align="right")

                        # Linked SSG(s)
                        ssg_info = item.get_ssg_count(compare_link=True, collection=instance)
                        add_one_item(rel_item, ssg_info, resizable, align="right")

                        # Add this Manu line to the list
                        rel_list.append(dict(id=item.id, cols=rel_item))

                    codicos['rel_list'] = rel_list

                    codicos['columns'] = [
                        'Manuscript', 
                        '<span title="Origin/Provenance">or./prov.</span>', 
                        '<span title="Date range">date</span>', 
                        '<span title="Sermons in this codicological unit">sermons</span>',
                        '<span title="Authoritative statement links">ssgs.</span>', 
                        ]
                    related_objects.append(codicos)

                if show_manu:
                    # OLD (see issue #363): List of Manuscripts contained in this collection
                    sTitle = "Manuscripts with sermons connected to this collection"
                    manuscripts = dict(title=sTitle, prefix="manu")
                    if resizable:
                        manuscripts['gridclass'] = "resizable"
                    manuscripts['classes'] = 'collapse in'

                    # Get the manuscripts linked to these SSGs (in this historical collection)
                    lstQ = []
                    lstQ.append(Q(manuitems__itemsermons__austats__collections=instance))
                    lstQ.append(Q(mtype="man"))
                    qs_manu = Manuscript.objects.filter(*lstQ).order_by(
                        'id').distinct().order_by('lcity__name', 'library__name', 'idno')

                    rel_list =[]
                    for item in qs_manu:
                        rel_item = []

                        # Get the codico's for this manuscript
                        codico_lst = item.manuscriptcodicounits.all().order_by('order')

                        # Shelfmark = IDNO
                        add_one_item(rel_item,  self.get_field_value("manu", item, "shelfmark"), False, title=item.idno, main=True, 
                                     link=reverse('manuscript_details', kwargs={'pk': item.id}))

                        # Origin
                        add_one_item(rel_item, self.get_field_value("manucodicos", codico_lst, "origprov"), resizable, 
                                     title="Origin (if known), followed by provenances (between brackets)")

                        # date range
                        add_one_item(rel_item, self.get_field_value("manucodicos", codico_lst, "daterange"), resizable, align="right")

                        # Number of sermons in this manuscript
                        add_one_item(rel_item, self.get_field_value("manu", item, "sermons"), resizable, align="right")

                        # Linked SSG(s)
                        ssg_info = item.get_ssg_count(compare_link=True, collection=instance)
                        add_one_item(rel_item, ssg_info, resizable, align="right")

                        # Add this Manu line to the list
                        rel_list.append(dict(id=item.id, cols=rel_item))

                    manuscripts['rel_list'] = rel_list

                    manuscripts['columns'] = [
                        'Shelfmark', 
                        '<span title="Origin/Provenance">or./prov.</span>', 
                        '<span title="Date range">date</span>', 
                        '<span title="Sermons in this manuscript">sermons</span>',
                        '<span title="Authoritative statement links">ssgs.</span>', 
                        ]
                    related_objects.append(manuscripts)
            elif self.manu != None:
                # This is a comparison between the SSGs in the historical collection and the sermons in the manuscript
                # (1) Start making a comparison table
                title = "Comparison with manuscript [{}]".format(self.manu.get_full_name())
                sermons = dict(title=title, prefix="sermo", gridclass="resizable")
                # (2) Get a list of sermons
                qs_s = Canwit.objects.filter(msitem__manu=self.manu).order_by('msitem__order')

                # Build the related list
                rel_list =[]
                equal_list = []
                index = 1
                for item in qs_s:
                    rel_item = []
                    # Determine the matching SSG from the Historical Collection
                    equal = Austat.objects.filter(canwit_austat__austat__in=qs_ssg, canwit_austat__canwit__id=item.id).first()
                    # List of SSGs that have been dealt with already
                    if equal != None: equal_list.append(equal.id)

                    # S: Order in Manuscript
                    rel_item.append({'value': index, 'initial': 'small'})
                    index += 1

                    # S: Locus
                    rel_item.append({'value': item.get_locus()})

                    # S: TItle
                    rel_item.append({'value': item.title, 'initial': 'small'})

                    # SSG: solemne code
                    if equal:
                        rel_item.append({'value': equal.get_lilacode_markdown(), 'initial': 'small'})
                    else:
                        rel_item.append({'value': "(none)", 'initial': 'small'})

                    ratio = 0.0
                    if equal:
                        # S: incipit + explicit compared
                        s_equal, ratio_equal = equal.get_incexp_match()
                        comparison, ratio = item.get_incexp_match(s_equal)
                        rel_item.append({'value': comparison, 'initial': 'small'})

                        # SSG: incipit + explicit compared
                        s_sermon, ratio_sermon = item.get_incexp_match()
                        comparison, ratio2 = equal.get_incexp_match(s_sermon)
                        rel_item.append({'value': comparison, 'initial': 'small'})
                    else:
                        # S: incipit + explicit compared
                        s_sermon, ratio_sermon = item.get_incexp_match()
                        rel_item.append({'value': s_sermon, 'initial': 'small'})

                        # SSG: incipit + explicit compared
                        rel_item.append({'value': "", 'initial': 'small'})

                    # Ratio of equalness
                    rel_item.append({'value': "{:.1%}".format(ratio), 'initial': 'small'})

                    rel_list.append(dict(id=item.id, cols=rel_item))

                # Check if there are any SSGs in the collection that have not been dealt with yet
                qs_ssg = instance.collections_austat.exclude(id__in=equal_list)
                for item in qs_ssg:
                    rel_item = []
                    equal = item
                    # S: Order in Manuscript
                    rel_item.append({'value': "-", 'initial': 'small'})

                    # S: Locus
                    rel_item.append({'value': "-"})

                    # S: TItle
                    rel_item.append({'value': "-", 'initial': 'small'})

                    # SSG: solemne code
                    rel_item.append({'value': equal.get_lilacode_markdown(), 'initial': 'small'})

                    # S: incipit + explicit compared
                    ratio = 0.0
                    rel_item.append({'value': "", 'initial': 'small'})

                    # SSG: incipit + explicit compared
                    s_equal, ratio_equal = equal.get_incexp_match()
                    rel_item.append({'value': s_equal, 'initial': 'small'})

                    # Ratio of equalness
                    rel_item.append({'value': ratio, 'initial': 'small'})

                    rel_list.append(dict(id=item.id, cols=rel_item))


                # Add the related list
                sermons['rel_list'] = rel_list

                # Set the columns
                sermons['columns'] = ['Order', 'Locus', 'Title', 
                                      '<span title="Authoritative statement">ssg</span>',
                                      '<span title="Incipit + explicit of sermon manifestation">inc/exp. s</span>', 
                                      '<span title="Incipit + explicit of Authoritative statement">inc/exp. ssg</span>',
                                      '<span title="Comparison ratio between inc/exp of S and SSG">ratio</span>']
                # Add to related objects
                related_objects.append(sermons)

            elif self.codico != None:
                # This is a comparison between the SSGs in the historical collection and the sermons in a codicological  unit
                # (1) Start making a comparison table
                title = "Comparison with codicological unit [{}]".format(self.codico.get_full_name())
                sermons = dict(title=title, prefix="sermo", gridclass="resizable")
                # (2) Get a list of sermons
                qs_s = Canwit.objects.filter(msitem__codico=self.codico).order_by('msitem__order')

                # Build the related list
                rel_list =[]
                equal_list = []
                index = 1
                for item in qs_s:
                    rel_item = []
                    # Determine the matching SSG from the Historical Collection
                    equal = Austat.objects.filter(canwit_austat__austat__in=qs_ssg, canwit_austat__canwit__id=item.id).first()
                    # List of SSGs that have been dealt with already
                    if equal != None: equal_list.append(equal.id)

                    # S: Order in Manuscript
                    rel_item.append({'value': index, 'initial': 'small'})
                    index += 1

                    # S: Locus
                    rel_item.append({'value': item.get_locus()})

                    # S: TItle
                    rel_item.append({'value': item.title, 'initial': 'small'})

                    # SSG: solemne code
                    if equal:
                        rel_item.append({'value': equal.get_lilacode_markdown(), 'initial': 'small'})
                    else:
                        rel_item.append({'value': "(none)", 'initial': 'small'})

                    ratio = 0.0
                    if equal:
                        # S: incipit + explicit compared
                        s_equal, ratio_equal = equal.get_incexp_match()
                        comparison, ratio = item.get_incexp_match(s_equal)
                        rel_item.append({'value': comparison, 'initial': 'small'})

                        # SSG: incipit + explicit compared
                        s_sermon, ratio_sermon = item.get_incexp_match()
                        comparison, ratio2 = equal.get_incexp_match(s_sermon)
                        rel_item.append({'value': comparison, 'initial': 'small'})
                    else:
                        # S: incipit + explicit compared
                        s_sermon, ratio_sermon = item.get_incexp_match()
                        rel_item.append({'value': s_sermon, 'initial': 'small'})

                        # SSG: incipit + explicit compared
                        rel_item.append({'value': "", 'initial': 'small'})

                    # Ratio of equalness
                    rel_item.append({'value': "{:.1%}".format(ratio), 'initial': 'small'})

                    rel_list.append(dict(id=item.id, cols=rel_item))

                # Check if there are any SSGs in the collection that have not been dealt with yet
                qs_ssg = instance.collections_austat.exclude(id__in=equal_list)
                for item in qs_ssg:
                    rel_item = []
                    equal = item
                    # S: Order in Manuscript
                    rel_item.append({'value': "-", 'initial': 'small'})

                    # S: Locus
                    rel_item.append({'value': "-"})

                    # S: TItle
                    rel_item.append({'value': "-", 'initial': 'small'})

                    # SSG: solemne code
                    rel_item.append({'value': equal.get_lilacode_markdown(), 'initial': 'small'})

                    # S: incipit + explicit compared
                    ratio = 0.0
                    rel_item.append({'value': "", 'initial': 'small'})

                    # SSG: incipit + explicit compared
                    s_equal, ratio_equal = equal.get_incexp_match()
                    rel_item.append({'value': s_equal, 'initial': 'small'})

                    # Ratio of equalness
                    rel_item.append({'value': ratio, 'initial': 'small'})

                    rel_list.append(dict(id=item.id, cols=rel_item))


                # Add the related list
                sermons['rel_list'] = rel_list

                # Set the columns
                sermons['columns'] = ['Order', 'Locus', 'Title', 
                                      '<span title="Authoritative statement">ssg</span>',
                                      '<span title="Incipit + explicit of sermon manifestation">inc/exp. s</span>', 
                                      '<span title="Incipit + explicit of Authoritative statement">inc/exp. ssg</span>',
                                      '<span title="Comparison ratio between inc/exp of S and Authoritative statementSSG">ratio</span>']
                # Add to related objects
                related_objects.append(sermons)

            context['related_objects'] = related_objects
            context['histogram_data'] = self.get_histogram_data(instance, instance.collections_austat.all(), 'collist_hist', 'd3')
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollHistDetails/add_to_context")

        # Return the context we have made
        return context

    def get_actions(self):
        html = []
        buttons = ['remove']    # This contains all the button names that need to be added

        # Start the whole spane
        html.append("<div class='blinded'>")
        
        # Add components
        if 'up' in buttons: 
            html.append("<a class='related-up' ><span class='glyphicon glyphicon-arrow-up'></span></a>")
        if 'down' in buttons: 
            html.append("<a class='related-down'><span class='glyphicon glyphicon-arrow-down'></span></a>")
        if 'remove' in buttons: 
            html.append("<a class='related-remove'><span class='glyphicon glyphicon-remove'></span></a>")

        # Finish up the span
        html.append("&nbsp;</span>")

        # COmbine the list into a string
        sHtml = "\n".join(html)
        # Return out HTML string
        return sHtml

    def get_field_value(self, type, instance, custom):
        sBack = ""
        if type == "manu":
            if custom == "shelfmark":
                sBack = "{}, {}, <span class='signature'>{}</span>".format(instance.get_city(), instance.get_library(), instance.idno)
            elif custom == "name":
                sBack = instance.name
            #elif custom == "origprov":
            #    sBack = "origin: {} (provenance[s]: {})".format(instance.get_origin(), instance.get_provenance_markdown(table=False))
            #elif custom == "daterange":
            #    sBack = "{}-{}".format(instance.yearstart, instance.yearfinish)
            elif custom == "sermons":
                sBack = instance.get_canwit_count()
        elif type == "manucodicos":
            # Note that [instance] now is a list of Codico elements
            lCombi = []
            if custom == "origprov":
                for obj in instance:
                    # Now [obj] is a Codico
                    lCombi.append( "origin: {} (provenance[s]: {})".format(obj.get_origins(), obj.get_provenance_markdown(table=False)))
                sBack = "; ".join(lCombi)
            elif custom == "daterange":
                for obj in instance:
                    lCombi.append( "{}-{}".format(obj.yearstart, obj.yearfinish))
                sBack = "; ".join(lCombi)
        elif type == "codicos":
            lCombi = []
            if custom == "origprov":
                sBack = "origin[s]: {} (provenance[s]: {})".format(instance.get_origins(), instance.get_provenance_markdown(table=False))
            elif custom == "daterange":
                sBack = "{}-{}".format(instance.yearstart, instance.yearfinish)
            elif custom == "manuscript":
                sBack = "<span class='signature'>{}</span>".format(instance.manuscript.get_full_name())
            elif custom == "sermons":
                sBack = instance.get_canwit_count()
        elif type == "collection":
            if custom == "caned":
                # Get a link to the Caned
                url = reverse('caned_details', kwargs={'pk': instance.id})
                # Get the lilacode for this CanEd
                lilacode = instance.get_lilacode()
                sBack = "<span class='badge signature'><a href='{}' class='nostyle'>{}</a></span>".format(url, lilacode)
        elif type == "austat":
            sBack, sTitle = AustatListView.get_field_value(None, instance, custom)
        return sBack


class CollHistCompare(CollHistDetails):
    """Compare the SSGs in a historical collection with the sermons in a manuscript"""

    def custom_init(self, instance):
        # FIrst perform the standard initialization
        response = super(CollHistCompare, self).custom_init(instance)

        # Make sure to get the Manuscript for comparison
        if user_is_authenticated(self.request) and user_is_ingroup(self.request, app_editor):
            manu_id = self.qd.get('manu')
            codico_id = self.qd.get('codico')
            if manu_id != None:
                manu = Manuscript.objects.filter(id=manu_id).first()
                if manu != None:
                    # We have the manuscript: the comparison can continue
                    self.manu = manu
                    self.codico = None
            elif codico_id != None:
                codico = Codico.objects.filter(id=codico_id).first()
                if codico != None:
                    # We have the codico: the comparison can continue
                    self.codico = codico
                    self.manu = None
        return None


class CollHistElevate(CollHistDetails):
    """ELevate this dataset to be a historical collection"""

    def custom_init(self, instance):
        if user_is_authenticated(self.request):
            # Double check if I have the right to do this...
            if user_is_ingroup(self.request, app_editor):
                # Change the settype to hc
                instance.settype = "hc"
                instance.save()
                self.redirectpage = reverse("collhist_details", kwargs={'pk': instance.id})
            elif instance.settype == "pd":
                # Determine what kind of dataset/collection this is
                if instance.owner == Profile.get_user_profile(self.request.user.username):
                    # Private dataset
                    self.redirectpage = reverse("collpriv_details", kwargs={'pk': instance.id})
                else:
                    # Public dataset
                    self.redirectpage = reverse("collpubl_details", kwargs={'pk': instance.id})
        else:
            self.redirectpage = reverse("home")
        return None


class CollHistApply(CollHistDetails):
    """Apply the historical collection to create a manuscript with sermons from the SSGs"""

    apply_type = ""

    def custom_init(self, instance):
        # Create a new manuscript that is based on this historical collection
        item_new = instance.get_hctemplate_copy(self.request.user.username, self.apply_type)
        if item_new == None:
            # THis wasn't successful: redirect to the details view
            self.redirectpage = reverse("collhist_details", kwargs={'pk': instance.id})
        elif self.apply_type == "tem":
            # A template has been created
            self.redirectpage = reverse("template_details", kwargs={'pk': item_new.id})
        else:
            # Manuscript created: re-direct to this manuscript
            self.redirectpage = reverse("manuscript_details", kwargs={'pk': item_new.id})
        return None


class CollHistManu(CollHistApply):
    """Apply the historical collection to create a manuscript with sermons from the SSGs"""

    apply_type = "man"


class CollHistTemp(CollHistApply):
    """Apply the historical collection to create a manuscript with sermons from the SSGs"""

    apply_type = "tem"


class CollSuperDetails(CollSuperEdit):
    """Like CollSuperEdit, but then with html"""
    rtype = "html"


class CollectionListView(BasicList):
    """Search and list collections"""

    model = Collection
    listform = CollectionForm
    prefix = "any"
    paginate_by = PAGINATE_BY_VALUE
    bUseFilter = True
    has_select2 = True
    basic_name_prefix = "coll"
    settype = "pd"              # Personal Dataset versus Historical Collection
    use_team_group = True
    plural_name = ""
    order_cols = ['scope', 'name', 'created', 'owner__user__username', '']
    order_default = order_cols
    order_heads = [{'name': 'Scope',        'order': 'o=1', 'type': 'str', 'custom': 'scope'},
                   {'name': 'Collection',   'order': 'o=2', 'type': 'str', 'field': 'name', 'linkdetails': True, 'main': True},
                   {'name': 'Created',      'order': 'o=3', 'type': 'str', 'custom': 'created'},
                   {'name': 'Owner',        'order': 'o=4', 'type': 'str', 'custom': 'owner'},
                   {'name': 'Frequency',    'order': '',    'type': 'str', 'custom': 'links'}
                   ]
    filters = [ {"name": "Collection", "id": "filter_collection", "enabled": False},
                {"name": "Owner",      "id": "filter_owner",      "enabled": False}]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'owner',     'fkfield': 'owner',  'keyS': 'owner', 'keyFk': 'id', 'keyList': 'ownlist', 'infield': 'id' },
            {'filter': 'collection','dbfield': 'name',   'keyS': 'collection_ta', 'keyList': 'collist', 'infield': 'name'}]},
        {'section': 'other', 'filterlist': [
            {'filter': 'coltype',   'dbfield': 'type',   'keyS': 'type',  'keyList': 'typelist' },
            {'filter': 'settype',   'dbfield': 'settype','keyS': 'settype'},
            {'filter': 'scope',     'dbfield': 'scope',  'keyS': 'scope'}]}
        ]

    def initializations(self):
        oErr = ErrHandle()
        try:
            if self.prefix == "sermo":
                self.plural_name = "Sermon collections"
                self.sg_name = "Sermon collection"
                self.searches[0]['filterlist'][1]['keyList'] = "collist_s"
            elif self.prefix == "manu":
                self.plural_name = "Manuscript Collections"
                self.sg_name = "Manuscript collection"
                self.searches[0]['filterlist'][1]['keyList'] = "collist_m"
            elif self.prefix == "gold":
                self.plural_name = "Gold sermons Collections"
                self.sg_name = "Sermon gold collection"
                self.searches[0]['filterlist'][1]['keyList'] = "collist_sg"
            elif self.prefix == "austat":
                self.plural_name = "Authoritative statement Collections"
                self.sg_name = "Authoritative statement collection"        
                self.searches[0]['filterlist'][1]['keyList'] = "collist_ssg"
            elif self.prefix == "any":
                self.new_button = False
                self.plural_name = "All types Collections"
                self.sg_name = "Collection"  
                self.order_cols = ['type', 'scope', 'name', 'created', 'owner__user__username', '']
                self.order_default = self.order_cols
                self.order_heads  = [
                    {'name': 'Type',        'order': 'o=1', 'type': 'str', 'custom': 'type'},
                    {'name': 'Scope',       'order': 'o=2', 'type': 'str', 'custom': 'scope'},
                    {'name': 'Dataset',     'order': 'o=3', 'type': 'str', 'field': 'name', 'linkdetails': True, 'main': True},
                    {'name': 'Created',     'order': 'o=4', 'type': 'str', 'custom': 'created'},
                    {'name': 'Owner',       'order': 'o=5', 'type': 'str', 'custom': 'owner'},
                    {'name': 'Frequency',   'order': '',    'type': 'str', 'custom': 'links'}
                ]  
            elif self.prefix == "priv":
                self.new_button = False
                self.titlesg = "Personal Dataset"
                self.plural_name = "My Datasets"
                self.sg_name = "My Dataset"  
                self.order_cols = ['type', 'name', 'scope', 'owner__user__username', 'created', '']
                self.order_default = self.order_cols
                self.order_heads  = [
                    {'name': 'Type',        'order': 'o=1', 'type': 'str', 'custom': 'type'},
                    {'name': 'Dataset',     'order': 'o=2', 'type': 'str', 'field': 'name', 'linkdetails': True, 'main': True},
                    {'name': 'Scope',       'order': 'o=3', 'type': 'str', 'custom': 'scope'},
                    {'name': 'Owner',       'order': 'o=4', 'type': 'str', 'custom': 'owner'},
                    {'name': 'Created',     'order': 'o=5', 'type': 'str', 'custom': 'created'},
                    {'name': 'Frequency',   'order': '',    'type': 'str', 'custom': 'links'}
                ]  
                self.filters = [ {"name": "My dataset", "id": "filter_collection", "enabled": False}]
                self.searches = [
                    {'section': '', 'filterlist': [
                        {'filter': 'collection','dbfield': 'name',   'keyS': 'collection_ta', 'keyList': 'collist', 'infield': 'name'}]},
                    {'section': 'other', 'filterlist': [
                        {'filter': 'owner',     'fkfield': 'owner',  'keyS': 'owner', 'keyFk': 'id', 'keyList': 'ownlist', 'infield': 'id' },
                        {'filter': 'coltype',   'dbfield': 'type',   'keyS': 'type',  'keyList': 'typelist' },
                        {'filter': 'settype',   'dbfield': 'settype','keyS': 'settype'},
                        {'filter': 'scope',     'dbfield': 'scope',  'keyS': 'scope'}]}
                    ]
            elif self.prefix == "publ":
                self.new_button = False
                self.plural_name = "Public Datasets"
                self.sg_name = "Public Dataset"  
                self.order_cols = ['type', 'name', 'created',  'owner__user__username', '']
                self.order_default = self.order_cols
                self.order_heads  = [
                    {'name': 'Type',        'order': 'o=1', 'type': 'str', 'custom': 'type'},
                    {'name': 'Dataset',     'order': 'o=2', 'type': 'str', 'field': 'name', 'linkdetails': True, 'main': True},
                    {'name': 'Created',     'order': 'o=3', 'type': 'str', 'custom': 'created'},
                    {'name': 'Owner',       'order': 'o=4', 'type': 'str', 'custom': 'owner'},
                    {'name': 'Frequency',   'order': '',    'type': 'str', 'custom': 'links'}
                ]  
                self.filters = [ {"name": "Public dataset", "id": "filter_collection", "enabled": False}]
                self.searches = [
                    {'section': '', 'filterlist': [
                        {'filter': 'collection','dbfield': 'name',   'keyS': 'collection_ta', 'keyList': 'collist', 'infield': 'name'},
                        {'filter': 'owner',     'fkfield': 'owner',  'keyS': 'owner', 'keyFk': 'id', 'keyList': 'ownlist', 'infield': 'id' }]},
                    {'section': 'other', 'filterlist': [
                        {'filter': 'coltype',   'dbfield': 'type',   'keyS': 'type',  'keyList': 'typelist' },
                        {'filter': 'settype',   'dbfield': 'settype','keyS': 'settype'},
                        {'filter': 'scope',     'dbfield': 'scope',  'keyS': 'scope'}]}
                    ]
            elif self.prefix == "hist":
                # OLD: self.new_button = False
                # See issue #47
                self.new_button = True
                self.settype = "hc"
                self.plural_name = "Historical Collections"
                self.sg_name = "Historical Collection"  
                self.order_cols = ['name', 'lilacode', '', 'ssgauthornum', 'owner__user__username', 'created']
                self.order_default = self.order_cols
                self.order_heads  = [
                    {'name': 'Historical Collection',   'order': 'o=1', 'type': 'str', 'field': 'name',     'linkdetails': True},
                    {'name': 'LiLaC',                   'order': 'o=2', 'type': 'str', 'custom': 'lilacode', 'linkdetails': True},
                    {'name': 'Authors',                 'order': '',    'type': 'str', 'custom': 'authors', 'allowwrap': True, 'main': True},
                    {'name': 'Author count',            'order': 'o=4', 'type': 'int', 'custom': 'authcount'},
                    {'name': 'Owner',                   'order': 'o=5', 'type': 'str', 'custom': 'owner'},
                    {'name': 'Created',                 'order': 'o=6', 'type': 'str', 'custom': 'created'},
                ]  
                # Add if user is app editor
                if user_is_authenticated(self.request) and user_is_ingroup(self.request, app_editor):
                    self.order_heads.append({'name': 'Manuscript', 'order': '', 'type': 'str', 'custom': 'manuscript'})
                    # Must also add to the order_cols and he order_default
                    if len(self.order_default) < len(self.order_heads):
                        self.order_default.append("")
                self.filters = [ 
                    {"name": "Collection",                  "id": "filter_collection",  "enabled": False},
                    {"name": "Project",                     "id": "filter_project",     "enabled": False},
                    {"name": "Authoritative statement...",  "id": "filter_austat",      "enabled": False, "head_id": "none"},
                    {"name": "Canon witness...",            "id": "filter_canwit",      "enabled": False, "head_id": "none"},
                    {"name": "Manuscript...",               "id": "filter_manu",        "enabled": False, "head_id": "none"},
                    # Section SSG = AuStat
                    {"name": "Author",          "id": "filter_ssgauthor",       "enabled": False, "head_id": "filter_austat"},
                    {"name": "Full text",       "id": "filter_ssgftext",        "enabled": False, "head_id": "filter_austat"},
                    {"name": "Translation",     "id": "filter_ssgftrans",       "enabled": False, "head_id": "filter_austat"},
                    {"name": "Lila code",       "id": "filter_ssgcode",         "enabled": False, "head_id": "filter_austat"},
                    {"name": "Number",          "id": "filter_ssgnumber",       "enabled": False, "head_id": "filter_austat"},
                    # {"name": "Gryson/Clavis",   "id": "filter_ssgsignature",  "enabled": False, "head_id": "filter_austat"},
                    {"name": "Keyword",         "id": "filter_ssgkeyword",      "enabled": False, "head_id": "filter_austat"},
                    {"name": "Status",          "id": "filter_ssgstype",        "enabled": False, "head_id": "filter_austat"},
                    # Section S
                    {"name": "Lila code",       "id": "filter_canwitcode",      "enabled": False, "head_id": "filter_canwit"},
                    {"name": "Author",          "id": "filter_canwitauthor",    "enabled": False, "head_id": "filter_canwit"},
                    {"name": "Full text",       "id": "filter_canwitftext",     "enabled": False, "head_id": "filter_canwit"},
                    {"name": "Translation",     "id": "filter_canwitftrans",    "enabled": False, "head_id": "filter_canwit"},
                    {"name": "Keyword",         "id": "filter_canwitkeyword",   "enabled": False, "head_id": "filter_canwit"}, 
                    #{"name": "Feast",           "id": "filter_canwitfeast",    "enabled": False, "head_id": "filter_canwit"},
                    #{"name": "Bible reference", "id": "filter_bibref",         "enabled": False, "head_id": "filter_canwit"},
                    {"name": "Note",            "id": "filter_canwitnote",      "enabled": False, "head_id": "filter_canwit"},
                    {"name": "Status",          "id": "filter_canwitstype",     "enabled": False, "head_id": "filter_canwit"},
                    # Section M
                    {"name": "Shelfmark",       "id": "filter_manuid",          "enabled": False, "head_id": "filter_manu"},
                    {"name": "Country",         "id": "filter_manucountry",     "enabled": False, "head_id": "filter_manu"},
                    {"name": "City",            "id": "filter_manucity",        "enabled": False, "head_id": "filter_manu"},
                    {"name": "Library",         "id": "filter_manulibrary",     "enabled": False, "head_id": "filter_manu"},
                    {"name": "Origin",          "id": "filter_manuorigin",      "enabled": False, "head_id": "filter_manu"},
                    {"name": "Provenance",      "id": "filter_manuprovenance",  "enabled": False, "head_id": "filter_manu"},
                    {"name": "Date range",      "id": "filter_manudaterange",   "enabled": False, "head_id": "filter_manu"},
                    {"name": "Keyword",         "id": "filter_manukeyword",     "enabled": False, "head_id": "filter_manu"},
                    {"name": "Status",          "id": "filter_manustype",       "enabled": False, "head_id": "filter_manu"},
                    ]
                self.searches = [
                    {'section': '', 'filterlist': [
                        {'filter': 'collection',    'dbfield': 'name',   'keyS': 'collection_ta', 'keyList': 'collist', 'infield': 'name'},
                        {'filter': 'project',       'fkfield': 'projects', 'keyFk': 'name', 'keyList': 'projlist', 'infield': 'name'},
                        ]},
                    # Section SSG = Austat
                    {'section': 'austat', 'filterlist': [
                        {'filter': 'ssgauthor',    'fkfield': 'austat_col__austat__author',            
                         'keyS': 'ssgauthorname', 'keyFk': 'name', 'keyList': 'ssgauthorlist', 'infield': 'id', 'external': 'gold-authorname' },
                        {'filter': 'ssgftext',     'dbfield': 'austat_col__austat__srchftext',   'keyS': 'ssgftext'},
                        {'filter': 'ssgftrans',    'dbfield': 'austat_col__austat__srchftrans',  'keyS': 'ssgftrans'},
                        {'filter': 'ssgcode',      'fkfield': 'austat_col__austat', 'keyFk': 'keycodefull',           
                         'keyS': 'ssgcode', 'keyList': 'ssglilalist', 'infield': 'id'},
                        {'filter': 'ssgnumber',    'dbfield': 'austat_col__austat__number',       'keyS': 'ssgnumber'},
                        {'filter': 'ssgkeyword',   'fkfield': 'austat_col__austat__keywords',          
                         'keyFk': 'name', 'keyList': 'ssgkwlist', 'infield': 'id'},
                        {'filter': 'ssgstype',     'dbfield': 'austat_col__austat__stype',             
                         'keyList': 'ssgstypelist', 'keyType': 'fieldchoice', 'infield': 'abbr' },
                        ]},
                    # Section S
                    {'section': 'canwit', 'filterlist': [
                        {'filter': 'canwitftext',    'dbfield': 'austat_col__austat__austat_canwits__srchftext',   'keyS': 'canwitftext'},
                        {'filter': 'canwitftrans',   'dbfield': 'austat_col__austat__austat_canwits__srchftrans',  'keyS': 'canwitftrans'},
                        {'filter': 'canwitcode',     'dbfield': 'austat_col__austat__austat_canwits__lilacodefull',
                         'keyS': 'canwitcode', 'keyList': 'canwitlilalist', 'infield': 'id'},
                        #{'filter': 'canwitfeast',    'dbfield': 'austat_col__austat__austat_canwits__feast',         'keyS': 'canwitfeast'},
                        #{'filter': 'bibref',        'dbfield': '$dummy',                                             'keyS': 'bibrefbk'},
                        #{'filter': 'bibref',        'dbfield': '$dummy',                                             'keyS': 'bibrefchvs'},
                        {'filter': 'canwitnote',     'dbfield': 'austat_col__austat__austat_canwits__additional',    'keyS': 'canwitnote'},
                        {'filter': 'canwitauthor',   'fkfield': 'austat_col__austat__austat_canwits__author',            
                         'keyS': 'canwitauthorname', 'keyFk': 'name', 'keyList': 'canwitauthorlist', 'infield': 'id', 'external': 'canwit-authorname' },
                        {'filter': 'canwitkeyword',  'fkfield': 'austat_col__austat__austat_canwits__keywords',          
                         'keyFk': 'name', 'keyList': 'canwitkwlist', 'infield': 'id' }, 
                        {'filter': 'canwitstype',    'dbfield': 'austat_col__austat__austat_canwits__stype',             
                         'keyList': 'canwitstypelist', 'keyType': 'fieldchoice', 'infield': 'abbr' }                    ]},
                    # Section M
                    {'section': 'manu', 'filterlist': [
                        {'filter': 'manuid',        'fkfield': 'austat_col__austat__austat_canwits__msitem__manu',                   
                         'keyS': 'manuidno',    'keyFk': "idno", 'keyList': 'manuidlist', 'infield': 'id'},
                        {'filter': 'manulibrary',       'fkfield': 'austat_col__austat__austat_canwits__msitem__manu__library',                
                         'keyS': 'libname_ta',    'keyId': 'library',     'keyFk': "name"},
                        {'filter': 'manukeyword',       'fkfield': 'austat_col__austat__austat_canwits__msitem__manu__keywords',               
                         'keyFk': 'name', 'keyList': 'manukwlist', 'infield': 'name' },
                        {'filter': 'manustype',         'dbfield': 'austat_col__austat__austat_canwits__msitem__manu__stype',                  
                         'keyList': 'manustypelist', 'keyType': 'fieldchoice', 'infield': 'abbr' },
                        {'filter': 'manuprovenance',    'fkfield': 'austat_col__austat__austat_canwits__msitem__codico__provenances__location',  
                         'keyS': 'prov_ta',       'keyId': 'prov',        'keyFk': "name"},
                        {'filter': 'manuorigin',        'fkfield': 'austat_col__austat__austat_canwits__msitem__codico__origin',                 
                         'keyS': 'origin_ta',     'keyId': 'origin',      'keyFk': "name"},
                        {'filter': 'manudaterange',     'dbfield': 'austat_col__austat__austat_canwits__msitem__codico__codico_dateranges__yearstart__gte',         
                         'keyS': 'date_from'},
                        {'filter': 'manudaterange',     'dbfield': 'austat_col__austat__austat_canwits__msitem__codico__codico_dateranges__yearfinish__lte',        
                         'keyS': 'date_until'},
                        ]},
                    # Section Other
                    {'section': 'other', 'filterlist': [
                        {'filter': 'owner',     'fkfield': 'owner',  'keyS': 'owner', 'keyFk': 'id', 'keyList': 'ownlist', 'infield': 'id' },
                        {'filter': 'coltype',   'dbfield': 'type',   'keyS': 'type',  'keyList': 'typelist' },
                        {'filter': 'settype',   'dbfield': 'settype','keyS': 'settype'},
                        {'filter': 'atype',    'dbfield': 'austat_col__austat__atype',    'keyS': 'atype'}, 
                        {'filter': 'scope',     'dbfield': 'scope',  'keyS': 'scope'}]}
                    ]
                    # ======== One-time adaptations ==============
        
            listview_adaptations("collhist_list")
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollectionListView/initializations")
        
        return None

    def add_to_context(self, context, initial):
        if self.prefix == "priv" and context['is_app_editor']:
            context['prefix'] = self.prefix
            context['user_button'] = render_to_string('seeker/dataset_add.html', context, self.request)
        return context

    def get_own_list(self):
        oErr = ErrHandle()
        qs = None
        try:
            # Get the user
            username = self.request.user.username
            user = User.objects.filter(username=username).first()
            # Get to the profile of this user
            if user is None:
                qs = Profile.objects.none()
                oErr.Status("CollectionListView/get_own_list: unknown user is [{}]".format(username))
            else:
                qs = Profile.objects.filter(user=user)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollectionListView/get_own_list")
        return qs

    def adapt_search(self, fields):
        lstExclude=None
        qAlternative = None
        if self.prefix == "hist":
            # The settype should be specified
            fields['settype'] = "hc"

            # The collection type is 'austat'
            fields['type'] = "austat"

            # The scope of a historical collection to be shown should be 'public'
            if user_is_authenticated(self.request) and user_is_ingroup(self.request, app_editor):
                profile = Profile.get_user_profile(self.request.user.username)
                fields['scope'] = ( ( Q(scope="priv") & Q(owner=profile) ) | Q(scope="team") | Q(scope="publ") )
            else:
                fields['scope'] = "publ"

            # Adapt the bible reference list
            bibrefbk = fields.get("bibrefbk", "")
            if bibrefbk != None and bibrefbk != "":
                bibrefchvs = fields.get("bibrefchvs", "")

                # Get the start and end of this bibref
                start, einde = Reference.get_startend(bibrefchvs, book=bibrefbk)

                # Find out which sermons have references in this range
                lstQ = []
                lstQ.append(Q(austat_col__austat__austat_canwits__sermonbibranges__bibrangeverses__bkchvs__gte=start))
                lstQ.append(Q(austat_col__austat__austat_canwits__sermonbibranges__bibrangeverses__bkchvs__lte=einde))
                collectionlist = [x.id for x in Collection.objects.filter(*lstQ).order_by('id').distinct()]

                fields['bibrefbk'] = Q(id__in=collectionlist)
            
            # Make sure we only use the Authoritative statements with accepted modifications
            # This means that atype should be 'acc' (and not: 'mod', 'rej' or 'def') 
            # With this condition we make sure ALL historical collections are in de unfiltered listview
            if fields['ssgcode'] != '':
                fields['atype'] = 'acc'
        elif self.prefix == "priv":
            # Show private datasets as well as those with scope "team", provided the person is in the team
            fields['settype'] = "pd"
            ownlist = self.get_own_list()
            if user_is_ingroup(self.request, app_editor):
                fields['scope'] = ( ( Q(scope="priv") & Q(owner__in=ownlist)  ) | Q(scope="team") )
            else:
                fields['scope'] = ( Q(scope="priv") & Q(owner__in=ownlist)  )
        elif self.prefix == "publ":
            # Show only public datasets
            fields['settype'] = "pd"
            # qAlternative = Q(scope="publ")
            fields['scope'] = "publ"
        else:
            # Check if the collist is identified
            if fields['ownlist'] == None or len(fields['ownlist']) == 0:
                # Get the user
                #username = self.request.user.username
                #user = User.objects.filter(username=username).first()
                ## Get to the profile of this user
                #qs = Profile.objects.filter(user=user)
                #profile = qs[0]
                #fields['ownlist'] = qs
                fields['ownlist'] = self.get_own_list()

                # Check on what kind of user I am
                if user_is_ingroup(self.request, app_editor):
                    # This is an editor: may see collections in the team
                    qAlternative = Q(scope="team") | Q(scope="publ")
                else:
                    # Common user: may only see those with public scope
                    # fields['scope'] = "publ"
                    qAlternative = Q(scope="publ")

            # Also make sure that we add the collection type, which is specified in "prefix"
            if self.prefix != "any":
                fields['type'] = self.prefix
            # The settype should be specified
            fields['settype'] = "pd"
        return fields, lstExclude, qAlternative

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        if custom == "links":
            html = []
            # Get the HTML code for the links of this instance
            number = instance.freqcanwit()
            if number > 0:
                url = reverse('canwit_list')
                html.append("<a href='{}?sermo-collist_s={}'>".format(url, instance.id))
                html.append("<span class='badge jumbo-1 clickable' title='Frequency in manifestation sermons'>{}</span></a>".format(number))
            number = instance.freqmanu()
            if number > 0:
                url = reverse('manuscript_list')
                html.append("<a href='{}?manu-collist_m={}'>".format(url, instance.id))
                html.append("<span class='badge jumbo-3 clickable' title='Frequency in manuscripts'>{}</span></a>".format(number))
            number = instance.freqsuper()
            if number > 0:
                url = reverse('austat_list')
                html.append("<a href='{}?ssg-collist_hist={}'>".format(url, instance.id))
                html.append("<span class='badge jumbo-3 clickable' title='Frequency in manuscripts'>{}</span></a>".format(number))
            # Combine the HTML code
            sBack = "\n".join(html)
        elif custom == "lilacode":
            sBack = instance.get_lilacode()
        elif custom == "type":
            sBack = instance.get_type_display()
        elif custom == "scope":
            sBack = instance.get_scope_display()
        elif custom == "created":
            sBack = get_crpp_date(instance.created, True)
        elif custom == "owner":
            if instance.owner is None:
                sBack = "(no user)"
            else:
                sBack = instance.owner.user.username
        elif custom == "authors":
            sBack = instance.get_authors_markdown()
        elif custom == "authcount":
            sBack = "{}".format(instance.ssgauthornum)
        elif custom == "manuscript":
            html = []
            url = reverse('collhist_manu', kwargs={'pk': instance.id})
            html.append("<a href='{}' title='Create a manuscript based on this historical collection'><span class='glyphicon glyphicon-open'></span></a>".format(url))
            url = reverse('collhist_temp', kwargs={'pk': instance.id})
            html.append("<a href='{}' title='Create a template based on this historical collection'><span class='glyphicon glyphicon-open' style='color: darkblue;'></span></a>".format(url))
            sBack = "\n".join(html)
        return sBack, sTitle
    

# ================== BASKET VIEWS AND OPERATIONS =========================

class BasketView(CanwitListView):
    """Like SermonListView, but then with the basket set true"""
    basketview = True


class BasketViewManu(ManuscriptListView):
    """Like ManuscriptListView, but then with the basket set true"""
    basketview = True


class BasketViewSuper(AustatListView):
    """Like AustatListView, but then with the basket set true"""
    basketview = True


class BasketUpdate(BasicPart):
    """Update contents of the canwit basket"""

    MainModel = Canwit
    clsBasket = Basket
    template_name = "seeker/basket_choices.html"
    entrycount = 0
    bFilter = False
    s_view = CanwitListView
    s_form = CanwitForm
    s_field = "canwit"
    colltype = "canwit"
    form_objects = [{'form': CollectionForm, 'prefix': colltype, 'readonly': True}]

    def add_to_context(self, context):
        # Reset the redirect page
        self.redirectpage = ""

        method = "use_profile_search_id_list"

        # Get the operation
        if 'operation' in self.qd:
            operation = self.qd['operation']
        else:
            return context

        username=self.request.user.username
        team_group=app_editor

        # Note: only operations in either of these two lists will be executed
        lst_basket_target = ["create", "add", "remove", "reset"]
        lst_basket_source = ["collcreate", "colladd", "rsetcreate", "dctlaunch"]

        # Get our profile
        profile = Profile.get_user_profile(self.request.user.username)
        if profile != None:

            # Obligatory initialization
            rset = None

            # Get the queryset
            self.filters, self.bFilter, qs, ini, oFields = search_generic(self.s_view, self.MainModel, self.s_form, self.qd, username, team_group)

            # Action depends on operations
            if operation in lst_basket_target:
                if method == "use_profile_search_id_list":
                    # Get the latest search results
                    search_s = getattr(profile, "search_{}".format(self.colltype))
                    search_id = []
                    if search_s != None and search_s != "" and search_s[0] == "[":
                        search_id = json.loads(search_s)
                    search_count = len(search_id)

                    kwargs = {'profile': profile}

                    # NOTE PROBLEM - we don't have the [oFields] at this point...

                    # Action depends on the operation specified
                    if search_count > 0 and operation == "create":
                        # Remove anything there
                        self.clsBasket.objects.filter(profile=profile).delete()
                        # Add
                        with transaction.atomic():
                            for item in search_id:
                                kwargs["{}_id".format(self.s_field)] = item
                                self.clsBasket.objects.create(**kwargs)
                        # Process history
                        profile.history(operation, self.colltype, oFields)
                    elif search_count > 0  and operation == "add":
                        # Add
                        with transaction.atomic():
                            for item in search_id:
                                kwargs["{}_id".format(self.s_field)] = item
                                obj = self.clsBasket.objects.filter(**kwargs).first()
                                if obj == None:
                                    self.clsBasket.objects.create(**kwargs)
                        # Process history
                        profile.history(operation, self.colltype, oFields)
                    elif search_count > 0  and operation == "remove":
                        # Add
                        with transaction.atomic():
                            for item in search_id:
                                kwargs["{}_id".format(self.s_field)] = item
                                self.clsBasket.objects.filter(**kwargs).delete()
                        # Process history
                        profile.history(operation, self.colltype, oFields)
                    elif operation == "reset":
                        # Remove everything from our basket
                        self.clsBasket.objects.filter(profile=profile).delete()
                        # Reset the history for this one
                        profile.history(operation, self.colltype)

                else:
                    
                    kwargs = {'profile': profile}

                    # Action depends on the operation specified
                    if qs and operation == "create":
                        # Remove anything there
                        self.clsBasket.objects.filter(profile=profile).delete()
                        # Add
                        with transaction.atomic():
                            for item in qs:
                                kwargs[self.s_field] = item
                                self.clsBasket.objects.create(**kwargs)
                        # Process history
                        profile.history(operation, self.colltype, oFields)
                    elif qs and operation == "add":
                        # Add
                        with transaction.atomic():
                            for item in qs:
                                kwargs[self.s_field] = item
                                self.clsBasket.objects.create(**kwargs)
                        # Process history
                        profile.history(operation, self.colltype, oFields)
                    elif qs and operation == "remove":
                        # Add
                        with transaction.atomic():
                            for item in qs:
                                kwargs[self.s_field] = item
                                self.clsBasket.objects.filter(**kwargs).delete()
                        # Process history
                        profile.history(operation, self.colltype, oFields)
                    elif operation == "reset":
                        # Remove everything from our basket
                        self.clsBasket.objects.filter(profile=profile).delete()
                        # Reset the history for this one
                        profile.history(operation, self.colltype)

            elif operation in lst_basket_source:
                # Queryset: the basket contents
                qs = self.clsBasket.objects.filter(profile=profile)

                # Get the history string
                history = getattr(profile, "history{}".format(self.colltype))

                # New collection or existing one?
                coll = None
                bChanged = False
                if operation == "collcreate":
                    # Save the current basket as a collection that needs to receive a name
                    # Note: this assumes [scope='priv'] default
                    coll = Collection.objects.create(path=history, settype="pd",
                            descrip="Created from a {} listview basket".format(self.colltype), 
                            owner=profile, type=self.colltype)
                    # Assign it a name based on its ID number and the owner
                    name = "{}_{}_{}".format(profile.user.username, coll.id, self.colltype)
                    coll.name = name
                    coll.save()
                    # once a collection has been created, make sure it gets assigned to a project
                    if not profile is None and coll.projects.count() == 0:
                        # Assign the default projects
                        projects = profile.get_defaults()
                        coll.set_projects(projects)

                elif operation == "rsetcreate":
                    # Save the current basket as a research-set that needs to receive a name
                    rset = ResearchSet.objects.create(
                        name="tijdelijk",
                        notes="Created from a {} listview basket".format(self.colltype),
                        profile=profile)
                    # Assign it a name based on its ID number and the owner
                    name = "{}_{}_{}".format(profile.user.username, rset.id, self.colltype)
                    rset.name = name
                    rset.save()
                elif operation == "dctlaunch":
                    # Save the current basket as a research-set that needs to receive a name
                    rset = ResearchSet.objects.create(
                        name="tijdelijk",
                        notes="Created from a {} listview basket for direct DCT launching".format(self.colltype),
                        profile=profile)
                    # Assign it a name based on its ID number and the owner
                    name = "{}_{}_{}".format(profile.user.username, rset.id, self.colltype)
                    rset.name = name
                    rset.save()
                elif oFields['collone']:
                    coll = oFields['collone']

                # Process the basket elements into the ResearchSet or into the Collection
                if rset != None:
                    with transaction.atomic():
                        for idx, item in enumerate(qs):
                            # Check if it doesn't exist yet
                            obj = SetList.objects.filter(researchset=rset, manuscript=item.manu).first()
                            if obj == None:
                                # Create this
                                order = idx + 1
                                SetList.objects.create(researchset=rset, 
                                                       order = order,
                                                       setlisttype="manu",
                                                       manuscript=item.manu)

                    # Make sure to redirect to this instance -- but only for RSETCREATE and DCTLAUNCH
                    if operation == "rsetcreate":
                        self.redirectpage = reverse('researchset_details', kwargs={'pk': rset.id})
                    elif operation == "dctlaunch":
                        # Get the default DCT for this ad-hoc ResearchSet
                        dct = rset.researchset_setdefs.first()
                        self.redirectpage = reverse('setdef_details', kwargs={'pk': dct.id})
                elif coll == None:
                    # TODO: provide some kind of error??
                    pass
                else:
                    # Link the collection with the correct model
                    kwargs = {'collection': coll}
                    if self.colltype == "canwit":
                        clsColl = CollectionCanwit
                        field = "canwit"
                    elif self.colltype == "manu":
                        clsColl = CollectionMan
                        field = "manuscript"
                    elif self.colltype == "austat":
                        clsColl = Caned
                        field = "austat"

                    # THis is only needed for collections
                    with transaction.atomic():
                        for item in qs:
                            kwargs[field] = getattr( item, self.s_field)
                            # Check if it doesn't exist yet
                            obj = clsColl.objects.filter(**kwargs).first()
                            if obj == None:
                                clsColl.objects.create(**kwargs)
                                # Note that some changes have been made
                                bChanged = True

                    # Make sure to redirect to this instance -- but only for COLLCREATE
                    if operation == "collcreate":
                        self.redirectpage = reverse('collpriv_details', kwargs={'pk': coll.id})
                    else:
                        # We are adding to an existing Collecion that is either public or private (or 'team' in scope)
                        if coll.settype == "pd":
                            if coll.scope == "publ":
                                # Public dataset
                                urltype = "publ"
                            else:
                                # Team or Priv
                                urltype = "priv"
                        elif coll.settype == "hc":
                            urltype = "hist"
                        collurl = reverse('coll{}_details'.format(urltype), kwargs={'pk': coll.id})
                        collname = coll.name
                        context['data'] = dict(collurl=collurl, collname=collname)
                        # Have changes been made?
                        if bChanged:
                            # Add the current basket history to the collection's path
                            lst_history_basket = json.loads(history)
                            lst_history_coll = json.loads(coll.path)
                            for item in lst_history_basket:
                                lst_history_coll.append(item)
                            coll.path = json.dumps(lst_history_coll)
                            coll.save()

            # Adapt the basket size
            context['basketsize'] = self.get_basketsize(profile)

            # Set the other context parameters
            if self.colltype == "canwit":
                context['basket_show'] = reverse('basket_show' )
                context['basket_update'] = reverse('basket_update')
            else:
                context['basket_show'] = reverse('basket_show_{}'.format(self.colltype))
                context['basket_update'] = reverse('basket_update_{}'.format(self.colltype))
            context['colltype'] = self.colltype

        # Return the updated context
        return context

    def get_basketsize(self, profile):
        # Adapt the basket size
        basketsize = profile.basketitems.count()
        profile.basketsize = basketsize
        profile.save()
        # Return the basketsize
        return basketsize

 
class BasketUpdateManu(BasketUpdate):
    """Update contents of the manuscript basket"""

    MainModel = Manuscript
    clsBasket = BasketMan 
    s_view = ManuscriptListView
    s_form = SearchManuForm
    s_field = "manu"
    colltype = "manu"
    form_objects = [{'form': CollectionForm, 'prefix': colltype, 'readonly': True}]

    def get_basketsize(self, profile):
        # Adapt the basket size
        basketsize = profile.basketitems_manu.count()
        profile.basketsize_manu = basketsize
        profile.save()
        # Return the basketsize
        return basketsize


class BasketUpdateSuper(BasketUpdate):
    """Update contents of the Austat basket"""

    MainModel = Austat
    clsBasket = BasketAustat
    s_view = AustatListView
    s_form = AustatForm
    s_field = "austat"
    colltype = "austat"
    form_objects = [{'form': CollectionForm, 'prefix': colltype, 'readonly': True}]

    def get_basketsize(self, profile):
        # Adapt the basket size
        basketsize = profile.basketitems_austat.count()
        profile.basketsize_austat = basketsize
        profile.save()
        # Return the basketsize
        return basketsize
